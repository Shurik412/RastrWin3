# -*- coding: utf-8 -*-
import csv
import os
import shutil
import time

import win32com.client
from openpyxl import (Workbook, load_workbook)
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.utils import get_column_letter


def createSummaryFile(kol_file, dir_file_1, sezon_regim):
    wb_summary = Workbook()
    ws_summary_activ = wb_summary.active
    ws_summary_activ.title = 'Summary File'
    ws_summary = wb_summary['Summary File']

    ws_summary.cell(column=1, row=1, value='Номер режима')
    ws_summary.cell(column=2, row=1, value='Степень демпфирования D')

    for val_for in range(1, kol_file + 1):
        Excel = win32com.client.Dispatch("Excel.Application")
        Excel.Visible = False
        DirFileExcel = f'{dir_file_1}/{sezon_regim}_{val_for}.xlsx'
        wb_set = Excel.Workbooks.Open(DirFileExcel)
        sheet = wb_set.ActiveSheet
        val_set = sheet.Cells(11, 12).value
        val_D = val_set
        wb_set.Close()
        print(f'{val_for} - ValD => {val_D}')
        ws_summary.cell(column=2, row=val_for + 1, value=val_D)
        ws_summary.cell(column=1, row=val_for + 1, value=f'{sezon_regim}_{val_for}')
        ws_summary[get_column_letter(2) + str(val_for + 1)].number_format = '0.0000000'

    wb_summary.save('{0}/{1}_0_SummaryFile.xlsx'.format(dir_file_1, sezon_regim))


def change_variable(ws, num_time_to_row):
    for col_index in range(1, ws.max_column):
        for row_index in range(num_time_to_row, ws.max_row + 1):
            val = str(ws[get_column_letter(col_index) + str(row_index)].value)
            val2 = val.replace("+", "")
            ws.cell(row=row_index,
                    column=col_index,
                    value=val2)

    for col_index in range(1, ws.max_column):
        for row_index in range(num_time_to_row, ws.max_row + 1):
            val = str(ws[get_column_letter(col_index) + str(row_index)].value)
            ind = val.find("+")
            index = val.find("E")
            if ind != -1 or index != -1:
                arr = val.split("E")
                a = float(arr[0])
                b = int(arr[1])
                c = float(a * 10 ** b)
                ws.cell(row=row_index,
                        column=col_index,
                        value=c)


def delete_files(sezon_regim, Unom_kV, kol_file):
    for num_regim_csv in range(1, kol_file + 1):
        file_1 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_Y_{Unom_kV}kV.csv"
        file_2 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_N_{Unom_kV_exp2}kV.csv"
        file_3 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_Y_{Unom_kV}kV.xlsx"
        file_4 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_N_{Unom_kV_exp2}kV.xlsx"

        os.remove(file_1)
        os.remove(file_2)
        os.remove(file_3)
        os.remove(file_4)


def convert_csv_to_xlsx(file, file_save):
    wb = Workbook()
    sheet = wb.active
    CSV_SEPARATOR = ";"
    with open(file) as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                    cell_exp = sheet.cell(row=r + 1, column=idx + 1)
                    cell_exp.value = val
    wb.save(file_save)


def createExcelFile(dir_file_1, sezon_regim):
    wb = Workbook()
    ws = wb.active
    ws1 = wb.create_sheet("Mysheet")
    ws2 = wb.create_sheet("Mysheet")
    ws.title = "exp_path"
    ws1.title = "1_exp"
    ws2.title = "2_exp"
    wb.save(f"{dir_file_1}/{sezon_regim}_1.xlsx")


def copy_files():
    createExcelFile(dir_file_1, sezon_regim)
    for num_regim_csv in range(1, kol_file + 1):
        file1 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_Y_{Unom_kV}kV.exp"
        file2 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_Y_{Unom_kV}kV.csv"
        file5 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_Y_{Unom_kV}kV.xlsx"

        shutil.copyfile(file1, file2)
        print(f"{num_regim_csv}-File CSV", file2)
        convert_csv_to_xlsx(file=file2, file_save=file5)

        file3 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_N_{Unom_kV_exp2}kV.exp"
        file4 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_N_{Unom_kV_exp2}kV.csv"
        file6 = dir_file_1 + f"/{sezon_regim}_{num_regim_csv}_N_{Unom_kV_exp2}kV.xlsx"

        shutil.copyfile(file3, file4)
        print("   -File CSV", file4)
        convert_csv_to_xlsx(file=file4, file_save=file6)

        if num_regim_csv > 1:
            shutil.copyfile(dir_file_1 + f"/{sezon_regim}_1.xlsx",
                            dir_file_1 + f"/{sezon_regim}_{num_regim_csv}.xlsx")
            print("   -File CSV", file4)


def max_var_object(ws_Excel_Unite_1_exp, col_gen, num_time_to_row):
    var_2 = float(0)
    for i in range(num_time_to_row, ws_Excel_Unite_1_exp.max_row):
        var_1 = float(ws_Excel_Unite_1_exp[get_column_letter(col_gen) + str(i)].value)
        if var_2 < var_1:
            var_2 = var_1
    return var_2


def min_var_object(ws_Excel_Unite_1_exp, col_gen, num_time_to_row):
    var_2 = float(10000)
    for i in range(num_time_to_row, ws_Excel_Unite_1_exp.max_row):
        var_1 = float(ws_Excel_Unite_1_exp[f'{get_column_letter(col_gen)}{str(i)}'].value)
        if var_2 > var_1:
            var_2 = var_1
    return var_2


def max_Ch60(ws_Excel_Unite_1_exp, col_gen, num_time_to_row):
    global max_col_Ch60
    for jj in range(num_time_to_row, ws_Excel_Unite_1_exp.max_row):
        var_3 = ws_Excel_Unite_1_exp[f'{get_column_letter(1)}{str(jj)}'].value
        if var_3 >= 58 and var_3 <= 65:
            max_col_Ch60 = jj
            break
        else:
            max_col_Ch60 = 0
    maxCh60 = max_var_object(ws_Excel_Unite_1_exp, col_gen, max_col_Ch60)
    return maxCh60


def min_Ch60(ws_Excel_Unite_1_exp, col_gen, num_time_to_row):
    global min_col_Ch60
    for j in range(num_time_to_row, ws_Excel_Unite_1_exp.max_row):
        var_3 = ws_Excel_Unite_1_exp[get_column_letter(1) + str(j)].value
        if var_3 >= 58 and var_3 <= 65:
            min_col_Ch60 = j
            break
        else:
            min_col_Ch60 = 0
    minCh60 = min_var_object(ws_Excel_Unite_1_exp, col_gen, min_col_Ch60)
    return minCh60


def change_date(ws_Excel_Unite_1_exp, ws_Excel_Unite_2_exp):
    num_time_row_exp1 = 0
    num_time_row_exp2 = 0
    c = 0
    for i in range(1, 50):
        value_to_time_exp1 = ws_Excel_Unite_1_exp[f'A{i}'].value
        if value_to_time_exp1 == ' TIME':
            num_time_row_exp1 = i + 1

        if value_to_time_exp1 == 50:
            c = c + 1
            if c > 1:
                _ = ws_Excel_Unite_1_exp.cell(column=1,
                                              row=i,
                                              value=value_to_time_exp1 + 0.0000001 * i)
    b = 0
    for j in range(1, 50):
        value_to_time_exp2 = ws_Excel_Unite_2_exp[f'A{j}'].value
        if value_to_time_exp2 == ' TIME':
            num_time_row_exp2 = j + 1
        else:
            num_time_row_exp2 = 0

        if value_to_time_exp2 == 50:
            b = b + 1
            if b > 1:
                _ = ws_Excel_Unite_2_exp.cell(column=1,
                                              row=j,
                                              value=value_to_time_exp2 + 0.0000001 * j)
            else:
                _ = 0
        else:
            b = 0


def plot(ws_Excel_Unite_1_exp,
         ws_Excel_Unite_2_exp,
         ws_Excel_Unite_exp_path,
         dir_file_Excel_Unite,
         NameGenInvestigated):
    global n, value_to_time, name_obj_gen, h

    ws_Excel_Unite_exp_path.cell(
        column=1,
        row=1,
        value=f"Загруженный файл {dir_file_Excel_Unite}"
    )
    num_time_to_row = 0
    col_count = ws_Excel_Unite_1_exp.max_column - 1
    col_gen = 0

    for n in range(1, 50):
        value_to_time = ws_Excel_Unite_1_exp[f"A{n}"].value
        if value_to_time == ' TIME':
            num_time_to_row = int(n) + 1

    for h in range(2, col_count + 1):
        name_obj_gen = ws_Excel_Unite_1_exp[f"{get_column_letter(h)}{str(num_time_to_row - 1)}"].value
        if name_obj_gen == f"P/{NameGenInvestigated}":
            col_gen = h
            break

    for g in range(2, col_count + 1):
        nameObjExp1 = ws_Excel_Unite_1_exp[f"{get_column_letter(g)}{str(num_time_to_row - 1)}"].value
        nameObjExp2 = ws_Excel_Unite_2_exp[f"{get_column_letter(g)}{str(num_time_to_row - 1)}"].value
        nameObjExp1_dict = dict_set[str(nameObjExp1)]
        nameObjExp2_dict = dict_set[str(nameObjExp2)]
        ws_Excel_Unite_1_exp.cell(column=g,
                                  row=num_time_to_row - 1,
                                  value=f"с PSS {nameObjExp1_dict}")

        ws_Excel_Unite_2_exp.cell(column=g,
                                  row=num_time_to_row - 1,
                                  value=f"без PSS {nameObjExp2_dict}")

    max_Chart = max_var_object(ws_Excel_Unite_1_exp, col_gen, num_time_to_row)
    min_Chart = min_var_object(ws_Excel_Unite_1_exp, col_gen, num_time_to_row)

    maxCh60 = max_Ch60(ws_Excel_Unite_1_exp, col_gen, num_time_to_row)
    minCh60 = min_Ch60(ws_Excel_Unite_1_exp, col_gen, num_time_to_row)

    # График 1 активной мощности СГ на интервале от 0 до 15
    ch4 = ScatterChart()

    xvalues_exp1 = Reference(ws_Excel_Unite_1_exp,
                             min_col=1,
                             min_row=num_time_to_row,
                             max_row=ws_Excel_Unite_1_exp.max_row)

    values_exp1 = Reference(ws_Excel_Unite_1_exp,
                            min_col=int(col_gen),
                            min_row=num_time_to_row - 1,
                            max_row=ws_Excel_Unite_1_exp.max_row)

    series_exp1 = Series(values_exp1, xvalues_exp1, title_from_data=True)
    name_object = ws_Excel_Unite_1_exp[f'{get_column_letter(col_gen)}{str(num_time_to_row - 1)}'].value
    remove_list = ['с', 'PSS']
    edit_str_as_list = name_object.split()
    final_list = [word for word in edit_str_as_list if word not in remove_list]
    final_str = ' '.join(final_list)
    final_list.clear()
    ch4.title = f"{final_str}"
    ch4.x_axis.title = "Время, с"  # название оси Х
    ch4.y_axis.title = "Активная мощность, МВт"  # название оси У
    # ch4.style = 12 # стиль диаграммы
    ch4.series.append(series_exp1)

    ch4.x_axis.scaling.min = 49
    ch4.y_axis.scaling.min = 0
    ch4.y_axis.scaling.max = max_Chart + 10
    ch4.x_axis.scaling.max = 65

    ws_Excel_Unite_exp_path.add_chart(ch4, f'{get_column_letter(1)}{str(5)}')

    # График 2 активной мощности СГ на интервале от 0 до 15
    ch2 = ScatterChart()

    xvalues_exp1 = Reference(ws_Excel_Unite_1_exp,
                             min_col=1,
                             min_row=num_time_to_row,
                             max_row=ws_Excel_Unite_1_exp.max_row)

    values_exp1 = Reference(ws_Excel_Unite_1_exp,
                            min_col=int(col_gen),
                            min_row=num_time_to_row - 1,
                            max_row=ws_Excel_Unite_1_exp.max_row)

    xvalues_exp2 = Reference(ws_Excel_Unite_2_exp,
                             min_col=1,
                             min_row=num_time_to_row,
                             max_row=ws_Excel_Unite_2_exp.max_row)

    values_exp2 = Reference(ws_Excel_Unite_2_exp,
                            min_col=int(col_gen),
                            min_row=num_time_to_row - 1,
                            max_row=ws_Excel_Unite_2_exp.max_row)

    series_exp1 = Series(values_exp1, xvalues_exp1, title_from_data=True)
    series_exp2 = Series(values_exp2, xvalues_exp2, title_from_data=True)

    name_object = ws_Excel_Unite_1_exp[f'{get_column_letter(col_gen)}{str(num_time_to_row - 1)}'].value
    remove_list = ['с', 'PSS']
    edit_str_as_list = name_object.split()
    final_list1 = [word for word in edit_str_as_list if word not in remove_list]
    final_str = ' '.join(final_list1)
    final_list1.clear()
    # ch2.style = 12 # стиль диаграммы
    ch2.title = f"{final_str}"
    ch2.x_axis.title = "Время, с"  # название оси Х
    ch2.y_axis.title = "Активная мощность, МВт"  # название оси У

    series_exp2.smooth = False  # сграживание кривой
    series_exp2.graphicalProperties.line.width = 350  # толщина кривой линии
    series_exp2.graphicalProperties.line.solidFill = 'FF0000'  # цвет кривой Red

    ch2.series.append(series_exp1)
    ch2.series.append(series_exp2)

    ch2.x_axis.scaling.min = 49
    ch2.y_axis.scaling.min = 0
    ch2.x_axis.scaling.max = 65

    ws_Excel_Unite_exp_path.add_chart(ch2, get_column_letter(11) + str(22))

    # График 3 активной мощности СГ на интервале от 1 до 6
    ch_PSS = ScatterChart()

    xvalues_exp1_pss = Reference(ws_Excel_Unite_1_exp,
                                 min_col=1,
                                 min_row=num_time_to_row,
                                 max_row=ws_Excel_Unite_1_exp.max_row)

    values_exp1_pss = Reference(ws_Excel_Unite_1_exp,
                                min_col=int(col_gen),
                                min_row=num_time_to_row - 1,
                                max_row=ws_Excel_Unite_1_exp.max_row)

    xvalues_exp2_pss = Reference(ws_Excel_Unite_2_exp,
                                 min_col=1,
                                 min_row=num_time_to_row,
                                 max_row=ws_Excel_Unite_2_exp.max_row)

    values_exp2_pss = Reference(ws_Excel_Unite_2_exp,
                                min_col=int(col_gen),
                                min_row=num_time_to_row - 1,
                                max_row=ws_Excel_Unite_2_exp.max_row)

    series_exp1_pss = Series(values_exp1_pss,
                             xvalues_exp1_pss,
                             title_from_data=True)

    series_exp2_pss = Series(values_exp2_pss,
                             xvalues_exp2_pss,
                             title_from_data=True)

    name_object = ws_Excel_Unite_1_exp[f'{get_column_letter(col_gen)}{str(num_time_to_row - 1)}'].value

    remove_list = ['с', 'PSS']
    edit_str_as_list = name_object.split()
    final_list2 = [word for word in edit_str_as_list if word not in remove_list]
    final_str = ' '.join(final_list2)
    final_list2.clear()
    ch_PSS.title = f"Эффективность PSS(стаб.) {final_str}"
    ch_PSS.x_axis.title = "Время, с"  # название оси Х
    ch_PSS.y_axis.title = "Активная мощность, МВт"  # название оси У

    series_exp2_pss.smooth = False  # сграживание кривой
    series_exp2_pss.graphicalProperties.line.width = 350  # сграживание кривой
    series_exp2_pss.graphicalProperties.line.solidFill = 'FF0000'  # цвет кривой Red

    ch_PSS.series.append(series_exp1_pss)
    ch_PSS.series.append(series_exp2_pss)

    ch_PSS.x_axis.scaling.min = 51
    ch_PSS.y_axis.scaling.min = int(min_Chart)
    ch_PSS.y_axis.scaling.max = int(max_Chart + 4)
    ch_PSS.x_axis.scaling.max = 56

    ws_Excel_Unite_exp_path.add_chart(ch_PSS, f'{get_column_letter(21)}{str(22)}')

    # График 4 активной мощности СГ на интервале от 11 до 26
    ch3 = ScatterChart()
    xvalues = Reference(ws_Excel_Unite_1_exp,
                        min_col=1,
                        min_row=num_time_to_row,
                        max_row=ws_Excel_Unite_1_exp.max_row)

    values = Reference(ws_Excel_Unite_1_exp,
                       min_col=col_gen,
                       min_row=num_time_to_row - 1,
                       max_row=ws_Excel_Unite_1_exp.max_row)

    series = Series(values, xvalues, title_from_data=True)

    name_object = ws_Excel_Unite_1_exp[f'{get_column_letter(col_gen)}{str(num_time_to_row - 1)}'].value

    remove_list = ['с', 'PSS']
    edit_str_as_list = name_object.split()
    final_list3 = [word for word in edit_str_as_list if word not in remove_list]
    final_str = ' '.join(final_list3)
    final_list3.clear()
    # ch3.style = 12    # стиль диаграммы
    ch3.title = f"{final_str}"
    ch3.x_axis.title = "Время, с"  # название оси Х
    ch3.y_axis.title = "Активная мощность, МВт"  # название оси У
    ch3.series.append(series)

    ch3.x_axis.scaling.min = 61
    ch3.y_axis.scaling.min = int(minCh60)
    ch3.y_axis.scaling.max = int(maxCh60 + 1)
    ch3.x_axis.scaling.max = 76

    ws_Excel_Unite_exp_path.add_chart(ch3, get_column_letter(1) + str(22))

    for k in range(2, col_count + 1):
        ch1 = ScatterChart()
        xvalues_exp1 = Reference(ws_Excel_Unite_1_exp,
                                 min_col=1,
                                 min_row=num_time_to_row,
                                 max_row=ws_Excel_Unite_1_exp.max_row)

        values_exp1 = Reference(ws_Excel_Unite_1_exp,
                                min_col=k,
                                min_row=num_time_to_row - 1,
                                max_row=ws_Excel_Unite_1_exp.max_row)

        xvalues_exp2 = Reference(ws_Excel_Unite_2_exp,
                                 min_col=1,
                                 min_row=num_time_to_row,
                                 max_row=ws_Excel_Unite_2_exp.max_row)

        values_exp2 = Reference(ws_Excel_Unite_2_exp,
                                min_col=k,
                                min_row=num_time_to_row - 1,
                                max_row=ws_Excel_Unite_2_exp.max_row)

        series_exp1 = Series(values_exp1, xvalues_exp1, title_from_data=True)
        series_exp2 = Series(values_exp2, xvalues_exp2, title_from_data=True)

        name_object = ws_Excel_Unite_1_exp[get_column_letter(k) + str(num_time_to_row - 1)].value
        remove_list = ['с', 'PSS']
        edit_str_as_list = name_object.split()
        final_list4 = [word for word in edit_str_as_list if word not in remove_list]
        final_str = ' '.join(final_list4)
        final_list4.clear()
        # ch1.style = 12    # стиль диаграммы
        ch1.title = f"{final_str}"
        ch1.x_axis.title = "Время, с"  # название оси Х
        ch1.y_axis.title = "Активная мощность, МВт"  # название оси У

        series_exp2.smooth = False  # сграживание кривой
        series_exp2.graphicalProperties.line.width = 350  # сграживание кривой
        series_exp2.graphicalProperties.line.solidFill = 'FF0000'  # цвет кривой Red

        series_exp2.smooth = False  # сграживание кривой
        series_exp2.graphicalProperties.line.width = 350  # сграживание кривой
        series_exp2.graphicalProperties.line.solidFill = 'FF0000'  # цвет кривой Red

        ch1.series.append(series_exp1)
        ch1.series.append(series_exp2)

        ch1.x_axis.scaling.min = 48
        ch1.y_axis.scaling.min = 0
        ch1.x_axis.scaling.max = 80
        # ch1.y_axis.scaling.max = 650 # диапазоны на графике

        size_plot = 10
        if k < int(col_count / 2):
            if k == 2:
                ws_Excel_Unite_exp_path.add_chart(ch1, get_column_letter(1) + str(40))
            elif k != 2:
                ws_Excel_Unite_exp_path.add_chart(ch1, get_column_letter((k - 2) * size_plot) + str(40))
        else:
            ws_Excel_Unite_exp_path.add_chart(ch1, get_column_letter((k - 2) * size_plot) + str(40))


def dempfir(ws_Excel_Unite_exp_path,
            ws_Excel_Unite_1_exp,
            NameGenInvestigated,
            col_count,
            num_time_to_row):
    # col_gen = 1
    # row_point_50 = 1
    global col_gen, row_point_50, row_point_65, row_point_75

    for h in range(2, col_count + 1):
        name_obj_gen = ws_Excel_Unite_1_exp[get_column_letter(h) + str(num_time_to_row - 1)].value
        if name_obj_gen == f"P/{NameGenInvestigated}":
            col_gen = h
            break

    for i_point_65 in range(num_time_to_row, ws_Excel_Unite_1_exp.max_row + 1):
        val_point_65 = float(ws_Excel_Unite_1_exp[get_column_letter(1) + str(i_point_65)].value)
        if val_point_65 >= float(65):
            row_point_65 = i_point_65
            break

    for i_point_75 in range(num_time_to_row, ws_Excel_Unite_1_exp.max_row + 1):
        val_point_75 = float(ws_Excel_Unite_1_exp[get_column_letter(1) + str(i_point_75)].value)
        if val_point_75 >= float(75):
            row_point_75 = i_point_75
            break

    for i_point_50 in range(num_time_to_row, ws_Excel_Unite_1_exp.max_row):
        val_point_50 = float(ws_Excel_Unite_1_exp[get_column_letter(1) + str(i_point_50)].value)
        if val_point_50 >= float(50):
            row_point_50 = i_point_50
            break

    ws_Excel_Unite_exp_path.cell(column=11, row=6, value="Степень демпфирования:")
    ws_Excel_Unite_exp_path.cell(column=11, row=8, value="dPос= ")
    ws_Excel_Unite_exp_path.cell(column=11, row=9, value="dPто= ")
    ws_Excel_Unite_exp_path.cell(column=11, row=11, value="D=dPто/dPос= ")
    ws_Excel_Unite_exp_path.cell(column=13, row=7, value="Max")
    ws_Excel_Unite_exp_path.cell(column=14, row=7, value="Min")

    ws_Excel_Unite_exp_path[get_column_letter(13) + str(8)] = "=MAX('1_exp'!{0}{1}:{0}{2})".format(
        get_column_letter(col_gen), row_point_50, ws_Excel_Unite_1_exp.max_row)  # коэфф dPос -> Max
    ws_Excel_Unite_exp_path[
        get_column_letter(14) + str(8)] = f"='1_exp'!{get_column_letter(col_gen)}{row_point_50}"  # коэфф dPос -> Min

    ws_Excel_Unite_exp_path[get_column_letter(13) + str(9)] = "=MAX('1_exp'!{0}{1}:{0}{2})".format(
        get_column_letter(col_gen), row_point_65, row_point_75)  # коэфф dPто -> Max
    ws_Excel_Unite_exp_path[get_column_letter(14) + str(9)] = "=MIN('1_exp'!{0}{1}:{0}{2})".format(
        get_column_letter(col_gen), row_point_65, row_point_75)  # коэфф dPто -> Min

    ws_Excel_Unite_exp_path[get_column_letter(12) + str(8)] = "=M8-N8"
    ws_Excel_Unite_exp_path[get_column_letter(12) + str(9)] = "=M9-N9"

    ws_Excel_Unite_exp_path[get_column_letter(12) + str(11)] = "=L9/L8"  # коэфф D

    ws_Excel_Unite_exp_path[get_column_letter(12) + str(11)].number_format = '0.0000000'


def main(dir_file_1, kol_file, sezon_regim, Unom_kV, NameGenInvestigated):
    global num_time_to_row
    copy_files()  # вызов функции копирования файлов
    for num in range(1, kol_file + 1):
        dir_file_Excel_Unite = f"{dir_file_1}/{sezon_regim}_{num}.xlsx"
        wb_Excel_Unite = load_workbook(filename=dir_file_Excel_Unite)

        print(num, ".Load Workbook:", dir_file_Excel_Unite)

        ws_Excel_Unite_1_exp = wb_Excel_Unite['1_exp']
        ws_Excel_Unite_2_exp = wb_Excel_Unite['2_exp']
        ws_Excel_Unite_exp_path = wb_Excel_Unite['exp_path']
        flag = 1
        if flag == 1:
            wb_exp1 = load_workbook(
                filename=dir_file_1 + f"/{sezon_regim}_{num}_Y_{Unom_kV}kV.xlsx")
            wb_exp2 = load_workbook(
                filename=dir_file_1 + f"/{sezon_regim}_{num}_N_{Unom_kV_exp2}kV.xlsx")

            ws_exp1 = wb_exp1.worksheets[0]
            ws_exp2 = wb_exp2.worksheets[0]

            for row in ws_exp1:
                for cell in row:
                    ws_Excel_Unite_1_exp[cell.coordinate].value = cell.value

            for row in ws_exp2:
                for cell in row:
                    ws_Excel_Unite_2_exp[cell.coordinate].value = cell.value

        col_count = ws_Excel_Unite_1_exp.max_column - 1

        for n in range(1, 50):
            value_to_time = ws_Excel_Unite_1_exp[f'A{n}'].value
            if value_to_time == ' TIME':
                num_time_to_row = int(n) + 1

        change_variable(ws_Excel_Unite_1_exp, num_time_to_row)

        change_variable(ws_Excel_Unite_2_exp, num_time_to_row)

        change_date(ws_Excel_Unite_1_exp,
                    ws_Excel_Unite_2_exp)  # вызывается функция для корректировки данных времени

        dempfir(ws_Excel_Unite_exp_path,
                ws_Excel_Unite_1_exp,
                NameGenInvestigated,
                col_count,
                num_time_to_row)

        plot(ws_Excel_Unite_1_exp,
             ws_Excel_Unite_2_exp,
             ws_Excel_Unite_exp_path,
             dir_file_Excel_Unite,
             NameGenInvestigated)  # вызов функции построения графиков

        wb_Excel_Unite.save(dir_file_1 + f"/{sezon_regim}_{num}.xlsx")


start_time = time.time()

dir_wb1 = input("Dir File EXP (C:\\Users\\EXP): ")
kol_file = int(input("Количество режимов -> (Пример: 51): "))
sezon_regim = input("ZimaMax=1, ZimaMin=2, LetoMax=3, LetoMin=4: ")
Unom_kV = input("Unom KZ(3) -> 500 or 14_500: ")
Unom_kV_exp2 = input("Unom KZ(3) EXP_2 -> 500 or 14_500: ")
NameGenInvestigated = input(
    "Название исследуемого генератора, example (P/52601222) or (52601222),if 0 -> object #1: ")

dir_file_1 = dir_wb1.replace('\\', '/')

dict_set = dict()
wb_set = load_workbook(f"{dir_file_1}/Settings.xlsx")
ws_set = wb_set['Set']
for i in range(2, ws_set.max_row + 1):
    kay_setting = ws_set[get_column_letter(1) + str(i)].value
    value_set = ws_set[get_column_letter(2) + str(i)].value
    dict_set[f'P/{kay_setting}'] = value_set
wb_set.close()

main(dir_file_1, kol_file, sezon_regim, Unom_kV, NameGenInvestigated)  # function

delete_files(sezon_regim, Unom_kV, kol_file)

createSummaryFile(kol_file, dir_file_1, sezon_regim)

print("--- %s seconds ---" % (time.time() - start_time))
print("The End.")
