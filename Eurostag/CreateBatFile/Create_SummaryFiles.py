# -*- coding: utf-8 -*-
from openpyxl import load_workbook

dict_col_one = [3, 4]
dict_col_two = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
dict_col_three = [17, 19,
                  21, 23, 25, 27, 29,
                  31, 33, 35, 37, 39,
                  41, 43, 45, 47, 49,
                  51, 53, 55, 57, 59,
                  61, 63, 65, 67, 69,
                  71, 73, 75, 77, 79,
                  81, 83, 85, 87, 89,
                  91, 93, 95, 97, 99,
                  101, 103, 105]


def summary(h=0, col='B', col_main='D'):
    if i == h:
        for j in dict_col_one:  # 3, 4
            value_wb2_i = ws2[f'{col}{j - 1}'].value
            ws[f'{col_main}{j}'] = value_wb2_i

        for u in dict_col_two:  # 6, 15
            value_wb2_i = ws2[f'{col}{u - 2}'].value
            ws[f'{col_main}{u}'] = value_wb2_i

        for h, elem in enumerate(dict_col_three):  # 2, 45
            value_wb2_i = ws2[f'{col}{h + 14}'].value
            ws[f'{col_main}{str(elem)}'] = value_wb2_i


file_name = fr'P:\RASTRWIN\12-Динамика\01-АРВ\07-Проверка АРВ\ОДУ Центра\3. ТЭЦ-26\Результаты расчетов  ТЭЦ-26 Г-3,4,6,7 — акт.xlsx'
wb = load_workbook(filename=file_name)
ws = wb['Лето MIN']

file_name_G = fr'P:\RASTRWIN\12-Динамика\01-АРВ\07-Проверка АРВ\ОДУ Центра\3. ТЭЦ-26\4. Лето мин\01. EXCEL\SummaryFile'

name_file_dict = ['G3', 'G4', 'G6', 'G7']

for i in range(0, 4):
    wb2 = load_workbook(filename=fr'{file_name_G}\{name_file_dict[i]}.xlsx')
    ws2 = wb2['Summary File']
    print(f'\ti={i} => {file_name_G}\\{name_file_dict[i]}.xlsx')

    summary(h=0, col='B', col_main='D')
    summary(h=1, col='B', col_main='E')
    summary(h=2, col='B', col_main='F')
    summary(h=3, col='B', col_main='G')

wb.save(filename=file_name)
