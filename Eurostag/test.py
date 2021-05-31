# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def chengeVariable(ws, num_time_to_row):
    print(ws.max_column)
    for col_index in range(1, ws.max_column + 1):
        for row_index in range(num_time_to_row, ws.max_row + 1):
            val = str(ws[get_column_letter(col_index) + str(row_index)].value)
            val2 = val.replace("+", "")
            ws.cell(row=row_index,
                    column=col_index,
                    value=val2)

    for col_index in range(1, ws.max_column + 1):
        for row_index in range(num_time_to_row, ws.max_row + 1):
            val = str(ws[get_column_letter(col_index) + str(row_index)].value)
            ind = val.find("+")
            index = val.find("E")
            if ind != -1 or index != -1:
                arr = val.split("E")
                a = float(arr[0])
                b = int(arr[1])
                c = float(a * 10 ** b)
                ws.cell(row=row_index,
                        column=col_index,
                        value=c)


file_name = r'C:\Users\Ohrimenko_AG\Desktop\test\test_20.xlsx'
wb = load_workbook(file_name)
ws = wb.active
print(ws)
print(wb.sheetnames)

chengeVariable(ws=ws, num_time_to_row=15)

wb.save(filename=r'C:\Users\Ohrimenko_AG\Desktop\test\test_new.xlsx')
