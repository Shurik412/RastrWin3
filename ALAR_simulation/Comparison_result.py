from openpyxl import load_workbook
from openpyxl.chart import ScatterChart
from settingsKostrGRES import list_scn, dict_repairs, dict_cell_chart, list_scn_name_AO
from R_modules.export_in_excel.chart import ChartExcelOtherSheet
from openpyxl.utils import get_column_letter
from settings_alar import dict_output_results

print('***************************************************************************')
print('***************** Запуск заполнения <Сравнение ЭМПП> **********************')
print('***************************************************************************')

primary_cell = 10

wb = load_workbook(filename=r'L:\SER\Охрименко\03. RastrWin3\02.Dynamic_KosrGRES\Результаты\result93v23.xlsx')
dir_fileSaveExcel = r'L:\SER\Охрименко\03. RastrWin3\02.Dynamic_KosrGRES\Результаты\result93v24.xlsx'
# ws_sheet = wb['Sheet']
# ws_sheet.title = 'Сравнение ЭМПП'
ws_result = wb['Сравнение ЭМПП']
len_list_scn = len(list_scn)
add_col_res = 0
i_dict_cell_result = 0
j_repairs_title_ch = 1
for i_list_scn in range(0, len_list_scn):
    ws_act_ = wb[dict_repairs[j_repairs_title_ch][3]]
    ch_obj_3 = ScatterChart()
    chart_3 = ChartExcelOtherSheet(work_sheet=ws_result, work_book=wb,
                                   chart_obj=ch_obj_3,
                                   chart_title=f"АО:{ws_act_[f'{get_column_letter(2 + add_col_res)}{9}'].value} {dict_output_results[1][3]}",
                                   x_axis_title='R ген', y_axis_title='X ген', switch_command_line=True,
                                   width_chart=30,
                                   height_chart=20)

    ch_obj_4 = ScatterChart()
    chart_4 = ChartExcelOtherSheet(work_book=wb, work_sheet=ws_result,
                                   chart_obj=ch_obj_4,
                                   chart_title=f"АО:{ws_act_[f'{get_column_letter(2 + add_col_res)}{9}'].value} {dict_output_results[2][3]}",
                                   x_axis_title='R ген', y_axis_title='X ген', switch_command_line=True,
                                   width_chart=30,
                                   height_chart=20)

    for j_repairs in dict_repairs:
        if j_repairs != 'Сравнение ЭМПП':
            ws_act_ = wb[dict_repairs[j_repairs][3]]
            chart_3.add_line_chart(min_col_x=2 + add_col_res, min_row_x=1 + primary_cell, max_row_x=ws_act_.max_row,
                                   min_col_y=3 + add_col_res, min_row_y=primary_cell, max_row_y=ws_act_.max_row,
                                   work_sheet_other=dict_repairs[j_repairs][3], title_from_data_ch=False,
                                   title_ch=dict_repairs[j_repairs][3],
                                   width_line_pt=1.5)

        if j_repairs != 'Сравнение ЭМПП':
            ws_act_ = wb[dict_repairs[j_repairs][3]]
            chart_4.add_line_chart(min_col_x=4 + add_col_res, min_row_x=1 + primary_cell, max_row_x=ws_act_.max_row,
                                   min_col_y=5 + add_col_res, min_row_y=primary_cell, max_row_y=ws_act_.max_row,
                                   work_sheet_other=dict_repairs[j_repairs][3], title_from_data_ch=False,
                                   title_ch=dict_repairs[j_repairs][3],
                                   width_line_pt=1.5)
    if i_list_scn > 0:
        chart_3.print_chart(cell_print_chart=f'{dict_cell_chart[i_dict_cell_result + 1][4]}')
        chart_4.print_chart(cell_print_chart=f'{dict_cell_chart[i_dict_cell_result + 1][5]}')
    else:
        chart_3.print_chart(cell_print_chart=f'{dict_cell_chart[i_dict_cell_result + 1][4]}')
        chart_4.print_chart(cell_print_chart=f'{dict_cell_chart[i_dict_cell_result + 1][5]}')

    i_dict_cell_result = i_dict_cell_result + 1
    add_col_res = add_col_res + 6

wb.save(filename=f'{dir_fileSaveExcel}')
