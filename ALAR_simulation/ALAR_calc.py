# -*- coding: utf-8 -*-
import win32com.client
from RastrWinLib.loading.load import LoadRUSTab
from RastrWinLib.loading.shablon import shablon_file_dynamic as sh_rst, shablon_file_scenario as sh_scn, \
    shablon_file_automation as sh_auto
from RastrWinLib.export_in_excel.export_data_rustab import ExportDataRUSTab
from settingsKostrGRES import dir_file_rst, dir_file_scn, list_manual, list_scn, dict_repairs, dict_output_results, \
    dict_cell_chart, dir_fileSaveExcel, list_scn_name_AO
from RastrWinLib.calculation.dyn_rgm_ekv_calc import Dynamic, SteadyState
from openpyxl import Workbook
from RastrWinLib.excel.chart import ChartExcelOtherSheet
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart
from RastrWinLib.getting.get import GettingParameterInstance, GettingParameterAttribute
from RastrWinLib.switch.vetv import SwitchLine

wb = Workbook()
Rastr = win32com.client.Dispatch('Astra.Rastr')

# загрузка файлов для расчта
load_file_rst = LoadRUSTab(rastr_win=Rastr, shablon=sh_rst, switch_command_line=True)
load_file_scn = LoadRUSTab(rastr_win=Rastr, shablon=sh_scn, switch_command_line=True)
load_file_auto = LoadRUSTab(rastr_win=Rastr, shablon=sh_auto, switch_command_line=True)
primary_cell = 10
for i in dict_repairs:
    print(f'===================================================================================')
    print(f'{i}.Режим №{i} => {dict_repairs[i][3]}')
    print(f'===================================================================================')
    ws = wb.create_sheet(title=dict_repairs[i][3])
    add_cols = 0
    i_dict_cellCh = 0
    i_list_scn_name = 0

    ch_obj_1 = ScatterChart()
    chart_1 = ChartExcelOtherSheet(work_book=wb, work_sheet=ws,
                                   chart_obj=ch_obj_1,
                                   chart_title=f'{dict_output_results[1][3]} {dict_repairs[i][3]}',
                                   x_axis_title='R, Ом', y_axis_title='X, Ом', switch_command_line=True,
                                   width_chart=25,
                                   height_chart=17)

    ch_obj_2 = ScatterChart()
    chart_2 = ChartExcelOtherSheet(work_book=wb, work_sheet=ws,
                                   chart_obj=ch_obj_2,
                                   chart_title=f'{dict_output_results[2][3]} {dict_repairs[i][3]}',
                                   x_axis_title='R, Ом', y_axis_title='X, Ом', switch_command_line=True,
                                   width_chart=25,
                                   height_chart=17)

    for j_scn in list_scn:
        # Загрузка файла сценария и rst
        load_file_auto.load(file_path="")
        load_file_rst.load(file_path=f'{dir_file_rst}')
        load_file_scn.load(file_path=f'{dir_file_scn}\\{j_scn}')

        # Отключение ветви
        vetv_ = SwitchLine(rastr_win=Rastr, table='vetv', switch_command_line=True)
        if dict_repairs[i][0] != '0' and dict_repairs[i][1] != '0':
            vetv_.off(ip=int(dict_repairs[i][0]),
                      iq=int(dict_repairs[i][1]),
                      np=int(dict_repairs[i][2]))
        # Расчет режима
        regime_ = SteadyState(rastr_win=Rastr, switch_command_line=True)
        regime_.rgm()

        # Расчет динамики
        dynamic_ = Dynamic(rastr_win=Rastr, calc_time=2.0, snap_max_count=1, switch_command_line=True)
        dynamic_.change_snap_max_count()
        dynamic_.change_calc_time()
        dynamic_.run()

        get_table_vetv = GettingParameterAttribute(rastr_win=Rastr, table='vetv')
        # начало zrb + j*zxb
        # конец  zre + j*zxe
        # Выгружаем параметры сопротивления генератора
        data_vetv = ExportDataRUSTab(rastr_win=Rastr, table='vetv', switch_command_line=True)
        zre_gen_1 = data_vetv.get_array(column='zre',
                                        key=f'(ip={dict_output_results[1][0]} & '
                                            f'iq={dict_output_results[1][1]} &'
                                            f'np={dict_output_results[1][2]})'
                                            f'|'
                                            f'(ip={dict_output_results[1][1]} & '
                                            f'iq={dict_output_results[1][0]} &'
                                            f' np={dict_output_results[1][2]})')

        for index_zrb, (par_zrb, t_zrb) in enumerate(zre_gen_1):
            ws[f'{get_column_letter(1 + add_cols)}{index_zrb + 1 + primary_cell}'] = t_zrb
            ws[f'{get_column_letter(2 + add_cols)}{index_zrb + 1 + primary_cell}'] = par_zrb

        zxe_gen_1 = data_vetv.get_array(column='zxe',
                                        key=f'(ip={dict_output_results[1][0]} &'
                                            f'iq={dict_output_results[1][1]} &'
                                            f'np={dict_output_results[1][2]})'
                                            f'|'
                                            f'(ip={dict_output_results[1][1]} &'
                                            f'iq={dict_output_results[1][0]} &'
                                            f'np={dict_output_results[1][2]})')

        for index_zxb, (par_zxb, t_zxb) in enumerate(zxe_gen_1):
            ws[f'{get_column_letter(3 + add_cols)}{index_zxb + 1 + primary_cell}'] = par_zxb

        zre_gen_2 = data_vetv.get_array(column='zre',
                                        key=f'(ip={dict_output_results[2][0]} & '
                                            f'iq={dict_output_results[2][1]} & '
                                            f'np={dict_output_results[2][2]})'
                                            f'|'
                                            f'(ip={dict_output_results[2][1]} & '
                                            f'iq={dict_output_results[2][0]} & '
                                            f'np={dict_output_results[2][2]})')

        for index_zre, (par_zre, t_zre) in enumerate(zre_gen_2):
            ws[f'{get_column_letter(4 + add_cols)}{index_zre + 1 + primary_cell}'] = par_zre

        zxe_gen_2 = data_vetv.get_array(column='zxe',
                                        key=f'(ip={dict_output_results[2][0]} &'
                                            f' iq={dict_output_results[2][1]} & '
                                            f'np={dict_output_results[2][2]})'
                                            f'|'
                                            f'(ip={dict_output_results[2][1]} & '
                                            f'iq={dict_output_results[2][0]} & '
                                            f'np={dict_output_results[2][2]})')

        for index_zxe, (par_zxe, t_zxe) in enumerate(zxe_gen_2):
            ws[f'{get_column_letter(5 + add_cols)}{index_zxe + 1 + primary_cell}'] = par_zxe

        for name_manual in range(len(list_manual)):
            ws[f'{get_column_letter(1 + add_cols)}{name_manual + 1}'] = list_manual[name_manual]

        attrib_DFWAutoActionScn = GettingParameterAttribute(rastr_win=Rastr, table='DFWAutoActionScn')
        id_DFWAutoActionScn = attrib_DFWAutoActionScn.getting(column='ObjectKey', key='Id=1')
        instance_DFWAutoActionScn = GettingParameterInstance(rastr_win=Rastr, tables='DFWAutoActionScn', key='Type=3')
        key_obj_DFW = instance_DFWAutoActionScn.get(column='ObjectKey')
        arr_key_obj_DFW = key_obj_DFW.split(',')
        dict_keyObjDFW = dict(ip=arr_key_obj_DFW[0], iq=arr_key_obj_DFW[1], np=arr_key_obj_DFW[2])

        attrib_vetv = GettingParameterAttribute(rastr_win=Rastr, table='vetv')
        name_vetv_AO = attrib_vetv.getting(column='name',
                                           key=f'(ip={dict_keyObjDFW.get("ip")}&'
                                               f'iq={dict_keyObjDFW.get("iq")}&'
                                               f'np={dict_keyObjDFW.get("np")})'
                                               f'|'
                                               f'(ip={dict_keyObjDFW.get("iq")}&'
                                               f'iq={dict_keyObjDFW.get("ip")}&'
                                               f'np={dict_keyObjDFW.get("np")})')

        attrib_u_nom_short_circuit_point = GettingParameterAttribute(rastr_win=Rastr, table='node')
        u_nom_short_circuit_point = attrib_u_nom_short_circuit_point.getting(column='uhom',
                                                                             key=f'ny={id_DFWAutoActionScn}')
        off_name_vetv = get_table_vetv.getting(column='name',
                                               key=f'(ip={dict_repairs[i][0]} & '
                                                   f'iq={dict_repairs[i][1]} & '
                                                   f'np={dict_repairs[i][2]})'
                                                   f'|'
                                                   f'(ip={dict_repairs[i][1]} &'
                                                   f' iq={dict_repairs[i][0]} &'
                                                   f' np={dict_repairs[i][2]})')

        ws[f'{get_column_letter(2 + add_cols)}{1}'] = dict_repairs[i][3]
        ws[f'{get_column_letter(2 + add_cols)}{2}'] = name_vetv_AO
        ws[f'{get_column_letter(2 + add_cols)}{3}'] = u_nom_short_circuit_point
        ws[f'{get_column_letter(2 + add_cols)}{4}'] = dict_repairs[i][0]
        ws[f'{get_column_letter(2 + add_cols)}{5}'] = dict_repairs[i][1]
        ws[f'{get_column_letter(2 + add_cols)}{6}'] = dict_repairs[i][2]
        ws[f'{get_column_letter(2 + add_cols)}{7}'] = off_name_vetv
        ws[f'{get_column_letter(2 + add_cols)}{8}'] = j_scn
        ws[f'{get_column_letter(2 + add_cols)}{9}'] = f'{list_scn_name_AO[i_list_scn_name]}'
        ws[f'{get_column_letter(2 + add_cols)}{10}'] = ''

        ws[f'{get_column_letter(3 + add_cols)}{1}'] = f'Файл сценария: {dict_repairs[i][3]}'
        ws[f'{get_column_letter(1 + add_cols)}{primary_cell}'] = 't, c'
        ws[f'{get_column_letter(2 + add_cols)}{primary_cell}'] = f'R_{dict_output_results[1][3]}, Ом'
        ws[f'{get_column_letter(3 + add_cols)}{primary_cell}'] = f'X_{dict_output_results[1][3]}, Ом'
        ws[f'{get_column_letter(4 + add_cols)}{primary_cell}'] = f'R_{dict_output_results[2][3]}, Ом'
        ws[f'{get_column_letter(5 + add_cols)}{primary_cell}'] = f'X_{dict_output_results[2][3]}, Ом'

        # Строим графики
        chart_1.add_line_chart(min_col_x=2 + add_cols, min_row_x=1 + primary_cell, max_row_x=ws.max_row,
                               min_col_y=3 + add_cols, min_row_y=primary_cell, max_row_y=ws.max_row,
                               title_ch=ws[f'{get_column_letter(2 + add_cols)}{9}'].value,
                               title_from_data_ch=False)

        chart_2.add_line_chart(min_col_x=4 + add_cols, min_row_x=1 + primary_cell, max_row_x=ws.max_row,
                               min_col_y=5 + add_cols, min_row_y=primary_cell, max_row_y=ws.max_row,
                               title_ch=ws[f'{get_column_letter(2 + add_cols)}{9}'].value,
                               title_from_data_ch=False)

        add_cols = add_cols + 6
        i_list_scn_name = i_list_scn_name + 1

    chart_1.print_chart(cell_print_chart='G11')
    chart_2.print_chart(cell_print_chart='W11')

    wb.save(filename=f'{dir_fileSaveExcel}\\result_KostGRES.xlsx')

print('***************************************************************************')
print('***************** Запуск заполнения <Сравнение ЭМПП> **********************')
print('***************************************************************************')
ws_sheet = wb['Sheet']
ws_sheet.title = 'Сравнение ЭМПП'
ws_result = wb['Сравнение ЭМПП']
len_list_scn = len(list_scn)
add_col_res = 0
i_dict_cell_result = 0
j_repairs_title_ch = 1
for i_list_scn in range(0, len_list_scn):
    ws_act_ = wb[dict_repairs[j_repairs_title_ch][3]]

    ch_obj_3 = ScatterChart()
    chart_3 = ChartExcelOtherSheet(work_sheet=ws_result, work_book=wb,
                                   chart_obj=ch_obj_3,
                                   chart_title=f"АО:{ws_act_[f'{get_column_letter(2 + add_col_res)}{9}'].value} {dict_output_results[1][3]}",
                                   x_axis_title='R ген', y_axis_title='X ген', switch_command_line=True,
                                   width_chart=30,
                                   height_chart=20)

    ch_obj_4 = ScatterChart()
    chart_4 = ChartExcelOtherSheet(work_book=wb, work_sheet=ws_result,
                                   chart_obj=ch_obj_4,
                                   chart_title=f"АО:{ws_act_[f'{get_column_letter(2 + add_col_res)}{9}'].value} {dict_output_results[2][3]}",
                                   x_axis_title='R ген', y_axis_title='X ген', switch_command_line=True,
                                   width_chart=30,
                                   height_chart=20)

    for j_repairs in dict_repairs:
        if j_repairs != 'Сравнение ЭМПП':
            ws_act_ = wb[dict_repairs[j_repairs][3]]
            chart_3.add_line_chart(min_col_x=2 + add_col_res, min_row_x=1 + primary_cell, max_row_x=ws_act_.max_row,
                                   min_col_y=3 + add_col_res, min_row_y=primary_cell, max_row_y=ws_act_.max_row,
                                   work_sheet_other=dict_repairs[j_repairs][3], title_from_data_ch=False,
                                   title_ch=dict_repairs[j_repairs][3],
                                   width_line_pt=1.5)

        if j_repairs != 'Сравнение ЭМПП':
            ws_act_ = wb[dict_repairs[j_repairs][3]]
            chart_4.add_line_chart(min_col_x=4 + add_col_res, min_row_x=1 + primary_cell, max_row_x=ws_act_.max_row,
                                   min_col_y=5 + add_col_res, min_row_y=primary_cell, max_row_y=ws_act_.max_row,
                                   work_sheet_other=dict_repairs[j_repairs][3], title_from_data_ch=False,
                                   title_ch=dict_repairs[j_repairs][3],
                                   width_line_pt=1.5)
    if i_list_scn > 0:
        chart_3.print_chart(cell_print_chart=f'{dict_cell_chart[i_dict_cell_result + 1][4]}')
        chart_4.print_chart(cell_print_chart=f'{dict_cell_chart[i_dict_cell_result + 1][5]}')
    else:
        chart_3.print_chart(cell_print_chart=f'{dict_cell_chart[i_dict_cell_result + 1][4]}')
        chart_4.print_chart(cell_print_chart=f'{dict_cell_chart[i_dict_cell_result + 1][5]}')

    i_dict_cell_result = i_dict_cell_result + 1
    add_col_res = add_col_res + 6

wb.save(filename=f'{dir_fileSaveExcel}\\result_KostGRES.xlsx')
