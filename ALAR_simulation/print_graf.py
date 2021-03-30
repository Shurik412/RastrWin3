from R_modules.export_in_excel.chart import ChartExcelOtherSheet
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart
from openpyxl.utils import get_column_letter

excel_file = r'L:\SER\Охрименко\03. RastrWin3\02.Dynamic_KosrGRES\Результаты\result93v22.xlsx'

wb = load_workbook(filename=excel_file)
ws = wb.active

num_sheet = wb.sheetnames
name_sheet = 'Сравнение ЭМПП'

name_scn = ['1.scn', '2.scn', '3.scn', '4.scn', '5.scn', '6.scn', '7.scn', '8.scn',
            '9.scn', '10.scn', '11.scn', '12.scn', '13.scn', '14.scn', '15.scn']
primary_cell = 10
for i in num_sheet:
    if i != name_sheet:
        print(i)
        add_col = 0
        ws_new = wb[i]
        ch = ScatterChart()
        chart_one = ChartExcelOtherSheet(work_book=wb, work_sheet=ws_new,
                                         chart_obj=ch,
                                         chart_title=i,
                                         width_chart=40,
                                         height_chart=20,
                                         x_axis_title='X, Ом',
                                         y_axis_title='R, Ом',
                                         switch_command_line=True)
        for j in name_scn:
            chart_one.add_line_chart(min_col_x=4 + add_col, min_row_x=1 + primary_cell, max_row_x=ws_new.max_row,
                                     min_col_y=5 + add_col, min_row_y=primary_cell, max_row_y=ws_new.max_row,
                                     title_ch=ws[f'{get_column_letter(2 + add_col)}{8}'].value,
                                     title_from_data_ch=False)

            add_col = add_col + 6

        chart_one.print_chart(cell_print_chart='CQ10')

wb.save(filename= r'L:\SER\Охрименко\03. RastrWin3\02.Dynamic_KosrGRES\Результаты\result93v23.xlsx')
