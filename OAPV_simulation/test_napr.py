from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(
    filename=r'L:\SER\Охрименко\03. RastrWin3\test_OAPV_macro\ВЛ 500 кВ Рязанская ГРЭС – Липецкая Западная.xlsx')
ws = wb['Раздел 1.1']

max_row = ws.max_row

for i in range(5, max_row):
    cell = ws[f'{get_column_letter(2)}{i}'].value
    ws[f'{get_column_letter(2)}{i}'] = cell * (-1)
    cell2 = ws[f'{get_column_letter(2)}{i}'].value
    print(f'cell_1 = {cell}, cell_2 = {cell2}')

wb.save(filename=r'L:\SER\Охрименко\03. RastrWin3\test_OAPV_macro\ВЛ 500 кВ Рязанская ГРЭС – Липецкая Западная-22.xlsx')