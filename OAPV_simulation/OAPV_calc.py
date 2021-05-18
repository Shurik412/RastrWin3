from icecream import ic
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart

from RastrWinLib.AstraRastr import RASTR
from RastrWinLib.calculation.dynamic import Dynamic
from RastrWinLib.calculation.regim import SteadyState
from RastrWinLib.excel.chart import ChartExcelOtherSheet
from RastrWinLib.export.export_data_rustab import ExportDataRUSTab
from RastrWinLib.loading.load import load_file
from RastrWinLib.loading.shablon import shablon_file_dynamic, shablon_file_scenario
from RastrWinLib.tables.Node.node import Node
from RastrWinLib.tables.Vetv.vetv import Vetv
from create_file_scn import CreateActionsSCN

file_excel = r'L:\SER\Охрименко\03. RastrWin3\18\ВЛ 500 кВ Борино-Воронежская.xlsx'

list_coordinates_of_graphs = [('B2', 'S2', 'AJ2', 'BB2', 'Раздел 1.1'),
                              ('B36', 'S36', 'AJ36', 'BB36', 'Раздел 1.2'),
                              ('B70', 'S70', 'AJ70', 'BB70', 'Раздел 2.1'),
                              ('B104', 'S104', 'AJ104', 'BB104', 'Раздел 2.2')]
start_row = 5
wb = load_workbook(filename=file_excel, data_only=True)
ws_settings = wb['Settings Macro']
ws_scn = wb['Сценарий']
ws_load = wb['Нагрузочные режимы']

dir_name_scn_one = rf'{ws_settings["B7"].value}\{ws_settings["B8"].value}'
dir_name_scn_two = rf'{ws_settings["B7"].value}\{ws_settings["B9"].value}'

# Формирует файлы сценарий.
Rastr = RASTR

scn_one = CreateActionsSCN(rastr_win=Rastr,
                           dir_name_file_excel=file_excel,
                           name_list_excel='Сценарий',
                           switch_command_line=False)

scn_one.create(start=14, finish=32)
scn_one.create_log(start=4, finish=8, switch_command_line=False)
scn_one.save_scn(dir_file_name_save_scn=rf'{ws_settings["B7"].value}\{ws_settings["B8"].value}')

scn_two = CreateActionsSCN(rastr_win=Rastr,
                           dir_name_file_excel=file_excel,
                           name_list_excel='Сценарий',
                           switch_command_line=False)

scn_two.create(start=35, finish=53)
scn_two.create_log(start=4, finish=8, switch_command_line=False)
scn_two.save_scn(dir_file_name_save_scn=rf'{ws_settings["B7"].value}\{ws_settings["B9"].value}')

file_rst_regim_one = fr'{ws_settings["B1"].value}\{ws_settings["B2"].value}'
file_rst_regim_two = fr'{ws_settings["B4"].value}\{ws_settings["B5"].value}'

# ***** начало, конец и номер параллельности ВЛ(ЛЭП) *****
ip = ws_scn["X25"].value
iq = ws_scn["Z25"].value
np = ws_scn["AB25"].value

# ***** Узлы ПС1 и ПС2 *****
node1 = ws_scn["T25"].value
node2 = ws_scn["AE25"].value

# ***** Время расчета *****
t_ras = float(ws_scn["N20"].value)

# ***** < multiplication by -1 > *****
direction_regim_one_ps_one = float(ws_load['I4'].value)
direction_regim_one_ps_two = float(ws_load['I5'].value)
direction_regim_two_ps_one = float(ws_load['I6'].value)
direction_regim_two_ps_two = float(ws_load['I7'].value)

title_1_1 = wb['Раздел 1.1']['A1'].value
p_ps_one_1_1 = wb['Раздел 1.1']['B3'].value
q_ps_one_1_1 = wb['Раздел 1.1']['C3'].value
p_ps_two_1_1 = wb['Раздел 1.1']['D3'].value
q_ps_two_1_1 = wb['Раздел 1.1']['E3'].value
u_ps_one_1_1 = wb['Раздел 1.1']['F3'].value
ag_u_ps_one_1_1 = wb['Раздел 1.1']['G3'].value
u_ps_two_1_1 = wb['Раздел 1.1']['H3'].value
ag_u_ps_two_1_1 = wb['Раздел 1.1']['I3'].value
d_ag_1_1 = wb['Раздел 1.1']['J3'].value

title_1_2 = wb['Раздел 1.2']['A1'].value
p_ps_one_1_2 = wb['Раздел 1.2']['B3'].value
q_ps_one_1_2 = wb['Раздел 1.2']['C3'].value
p_ps_two_1_2 = wb['Раздел 1.2']['D3'].value
q_ps_two_1_2 = wb['Раздел 1.2']['E3'].value
u_ps_one_1_2 = wb['Раздел 1.2']['F3'].value
ag_u_ps_one_1_2 = wb['Раздел 1.2']['G3'].value
u_ps_two_1_2 = wb['Раздел 1.2']['H3'].value
ag_u_ps_two_1_2 = wb['Раздел 1.2']['I3'].value
d_ag_1_2 = wb['Раздел 1.2']['J3'].value

title_2_1 = wb['Раздел 2.1']['A1'].value
p_ps_one_2_1 = wb['Раздел 2.1']['B3'].value
q_ps_one_2_1 = wb['Раздел 2.1']['C3'].value
p_ps_two_2_1 = wb['Раздел 2.1']['D3'].value
q_ps_two_2_1 = wb['Раздел 2.1']['E3'].value
u_ps_one_2_1 = wb['Раздел 2.1']['F3'].value
ag_u_ps_one_2_1 = wb['Раздел 2.1']['G3'].value
u_ps_two_2_1 = wb['Раздел 2.1']['H3'].value
ag_u_ps_two_2_1 = wb['Раздел 2.1']['I3'].value
d_ag_2_1 = wb['Раздел 2.1']['J3'].value

title_2_2 = wb['Раздел 2.2']['A1'].value
p_ps_one_2_2 = wb['Раздел 2.2']['B3'].value
q_ps_one_2_2 = wb['Раздел 2.2']['C3'].value
p_ps_two_2_2 = wb['Раздел 2.2']['D3'].value
q_ps_two_2_2 = wb['Раздел 2.2']['E3'].value
u_ps_one_2_2 = wb['Раздел 2.2']['F3'].value
ag_u_ps_one_2_2 = wb['Раздел 2.2']['G3'].value
u_ps_two_2_2 = wb['Раздел 2.2']['H3'].value
ag_u_ps_two_2_2 = wb['Раздел 2.2']['I3'].value
d_ag_2_2 = wb['Раздел 2.2']['J3'].value

dict_name_chart = {'Раздел 1.1': (
    title_1_1, p_ps_one_1_1, q_ps_one_1_1, p_ps_two_1_1, q_ps_one_1_1, u_ps_one_1_1, ag_u_ps_one_1_1, u_ps_two_1_1,
    ag_u_ps_two_1_1, d_ag_1_1),
    'Раздел 1.2': (
        title_1_2, p_ps_one_1_2, q_ps_one_1_2, p_ps_two_1_2, q_ps_one_1_2, u_ps_one_1_2, ag_u_ps_one_1_2, u_ps_two_1_2,
        ag_u_ps_two_1_2, d_ag_1_2),
    'Раздел 2.1': (
        title_2_1, p_ps_one_2_1, q_ps_one_2_1, p_ps_two_2_1, q_ps_one_2_1, u_ps_one_2_1, ag_u_ps_one_2_1, u_ps_two_2_1,
        ag_u_ps_two_2_1, d_ag_2_1),
    'Раздел 2.2': (
        title_2_2, p_ps_one_2_2, q_ps_one_2_2, p_ps_two_2_2, q_ps_one_2_2, u_ps_one_2_2, ag_u_ps_one_2_2, u_ps_two_2_2,
        ag_u_ps_two_2_2, d_ag_2_2)}

wb.close()

wb_w = load_workbook(filename=file_excel)
ws_1 = wb_w['Раздел 1.1']
ws_2 = wb_w['Раздел 1.2']
ws_3 = wb_w['Раздел 2.1']
ws_4 = wb_w['Раздел 2.2']

# 1. ********************* Режим 1 сценарий 1 ******************
load_file(rastr_win=Rastr, file_path=file_rst_regim_one, shabl=shablon_file_dynamic, switch_command_line=True)
load_file(rastr_win=Rastr, file_path=dir_name_scn_one, shabl=shablon_file_scenario, switch_command_line=True)
load_file(rastr_win=Rastr, switch_command_line=True)

dyn_obj = Dynamic(rastr_win=Rastr, calc_time=t_ras, snap_max_count=1, switch_command_line=True)
rgm_obj = SteadyState(rastr_win=Rastr, par="", switch_command_line=True)

rgm_obj.rgm()
dyn_obj.change_calc_time()
dyn_obj.change_snap_max_count()
dyn_obj.run()

pq_data_r1_scn1 = ExportDataRUSTab(rastr_win=Rastr, table=Vetv.table, switch_command_line=True)
pl_ip_r1_scn1 = pq_data_r1_scn1.get_array(column=Vetv.pl_ip,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')
ql_ip_r1_scn1 = pq_data_r1_scn1.get_array(column=Vetv.ql_ip,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')

for index, (par_pl_ip, par_t) in enumerate(pl_ip_r1_scn1):
    ws_1[f'A{index + start_row}'] = par_t
    ws_1[f'B{index + start_row}'] = par_pl_ip

for index, (ql_ip, t) in enumerate(ql_ip_r1_scn1):
    ws_1[f'C{index + start_row}'] = ql_ip

pl_iq_r1_scn1 = pq_data_r1_scn1.get_array(column=Vetv.pl_iq,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')
ql_iq_r1_scn1 = pq_data_r1_scn1.get_array(column=Vetv.ql_iq,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')

for index, (pl_iq, t) in enumerate(pl_iq_r1_scn1):
    ws_1[f'D{index + start_row}'] = pl_iq

for index, (ql_iq, t) in enumerate(ql_iq_r1_scn1):
    ws_1[f'E{index + start_row}'] = ql_iq

u_angleU_r1_scn1_subs1 = ExportDataRUSTab(rastr_win=Rastr, table=Node.table, switch_command_line=True)
u_r1_scn1_node1 = u_angleU_r1_scn1_subs1.get_array(column=Node.vras, key=f'{Node.ny}={node1}')
angleU_r1_scn1_node1 = u_angleU_r1_scn1_subs1.get_array(column=Node.delta, key=f'{Node.ny}={node1}')

for index, (u_var_node1, t) in enumerate(u_r1_scn1_node1):
    ws_1[f'F{index + start_row}'] = u_var_node1

for index, (angleU_node1, t) in enumerate(angleU_r1_scn1_node1):
    ws_1[f'G{index + start_row}'] = angleU_node1

u_r1_scn1_node2 = u_angleU_r1_scn1_subs1.get_array(column=Node.vras, key=f'{Node.ny}={node2}')
angleU_r1_scn1_node2 = u_angleU_r1_scn1_subs1.get_array(column=Node.delta, key=f'{Node.ny}={node2}')

for index, (u_var_node2, t) in enumerate(u_r1_scn1_node2):
    ws_1[f'H{index + start_row}'] = u_var_node2

for index, (angleU_node2, t) in enumerate(angleU_r1_scn1_node2):
    ws_1[f'I{index + start_row}'] = angleU_node2

for i in range(start_row, ws_1.max_row + 1):
    ws_1[f'J{i}'] = f'=I{i}-G{i}'

ic(ws_1.max_row - 2)
ic(ws_1.max_row)
for i in range(start_row, ws_1.max_row - 2):
    ic(i)
    p_1_1_1 = ws_1[f'B{i}'].value
    j = float(p_1_1_1) * direction_regim_one_ps_one
    ws_1[f'B{i}'] = j

    q_1_1_1 = ws_1[f'C{i}'].value
    j = float(q_1_1_1) * direction_regim_one_ps_one
    ws_1[f'C{i}'] = j

    p_1_1_2 = ws_1[f'D{i}'].value
    j = float(p_1_1_2) * direction_regim_one_ps_two
    ws_1[f'D{i}'] = j

    q_1_1_2 = ws_1[f'E{i}'].value
    j = float(q_1_1_2) * direction_regim_one_ps_two
    ws_1[f'E{i}'] = j

wb_w.save(filename=file_excel)

# 2. ********************* Режим 1 сценарий 2 ******************
load_file(rastr_win=Rastr, file_path=file_rst_regim_one, shabl=shablon_file_dynamic, switch_command_line=True)
load_file(rastr_win=Rastr, file_path=dir_name_scn_two, shabl=shablon_file_scenario, switch_command_line=True)
load_file(rastr_win=Rastr, switch_command_line=True)

rgm_obj.rgm()
dyn_obj.change_calc_time()
dyn_obj.change_snap_max_count()
dyn_obj.run()

pq_data_r1_scn2 = ExportDataRUSTab(rastr_win=Rastr, table=Vetv.table, switch_command_line=True)
pl_ip_r1_scn2 = pq_data_r1_scn2.get_array(column=Vetv.pl_ip,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')
ql_ip_r1_scn2 = pq_data_r1_scn2.get_array(column=Vetv.ql_ip,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')

for index, (pl_ip, t) in enumerate(pl_ip_r1_scn2):
    ws_2[f'A{index + start_row}'] = t
    ws_2[f'B{index + start_row}'] = pl_ip

for index, (ql_ip, t) in enumerate(ql_ip_r1_scn2):
    ws_2[f'C{index + start_row}'] = ql_ip

pl_iq_r1_scn2 = pq_data_r1_scn2.get_array(column=Vetv.pl_iq,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')
ql_iq_r1_scn2 = pq_data_r1_scn2.get_array(column=Vetv.ql_iq,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')

for index, (pl_iq, t) in enumerate(pl_iq_r1_scn2):
    ws_2[f'D{index + start_row}'] = pl_iq

for index, (ql_iq, t) in enumerate(ql_iq_r1_scn2):
    ws_2[f'E{index + start_row}'] = ql_iq

u_angleU_r1_scn2 = ExportDataRUSTab(rastr_win=Rastr, table=Node.table, switch_command_line=True)
u_r1_scn2_node1 = u_angleU_r1_scn2.get_array(column=Node.vras, key=f'({Node.ny}={node1})')
angleU_r1_scn2_node1 = u_angleU_r1_scn2.get_array(column=Node.delta, key=f'({Node.ny}={node1})')

for index, (u_var_node1, t) in enumerate(u_r1_scn2_node1):
    ws_2[f'F{index + start_row}'] = u_var_node1

for index, (angleU_node1, t) in enumerate(angleU_r1_scn2_node1):
    ws_2[f'G{index + start_row}'] = angleU_node1

u_r1_scn2_node2 = u_angleU_r1_scn2.get_array(column=Node.vras, key=f'({Node.ny}={node2})')
angleU_r1_scn2_node2 = u_angleU_r1_scn2.get_array(column=Node.delta, key=f'({Node.ny}={node2})')

for index, (u_var_node2, t) in enumerate(u_r1_scn2_node2):
    ws_2[f'H{index + start_row}'] = u_var_node2

for index, (angleU_node2, t) in enumerate(angleU_r1_scn2_node2):
    ws_2[f'I{index + start_row}'] = angleU_node2

for i in range(start_row, ws_2.max_row + 1):
    ws_2[f'J{i}'] = f'=I{i}-G{i}'

for i in range(start_row, ws_2.max_row + 1):
    p_1_2_1 = ws_2[f'B{i}'].value
    j = p_1_2_1 * direction_regim_one_ps_one
    ws_2[f'B{i}'] = j

    q_1_2_1 = ws_2[f'C{i}'].value
    j = q_1_2_1 * direction_regim_one_ps_one
    ws_2[f'C{i}'] = j

    p_1_2_2 = ws_2[f'D{i}'].value
    j = p_1_2_2 * direction_regim_one_ps_two
    ws_2[f'D{i}'] = j

    q_1_2_2 = ws_2[f'E{i}'].value
    j = q_1_2_2 * direction_regim_one_ps_two
    ws_2[f'E{i}'] = j

wb_w.save(filename=file_excel)

# 3. ********************* Режим 2 сценарий 1 ******************
load_file(rastr_win=Rastr, file_path=file_rst_regim_two, shabl=shablon_file_dynamic, switch_command_line=True)
load_file(rastr_win=Rastr, file_path=dir_name_scn_one, shabl=shablon_file_scenario, switch_command_line=True)
load_file(rastr_win=Rastr, switch_command_line=True)

rgm_obj.rgm()
dyn_obj.change_calc_time()
dyn_obj.change_snap_max_count()
dyn_obj.run()

pq_data_r2_scn1 = ExportDataRUSTab(rastr_win=Rastr, table=Vetv.table, switch_command_line=True)
pl_ip_r2_scn1 = pq_data_r2_scn1.get_array(column=Vetv.pl_ip,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')
ql_ip_r2_scn1 = pq_data_r2_scn1.get_array(column=Vetv.ql_ip,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')

for index, (pl_ip, t) in enumerate(pl_ip_r2_scn1):
    ws_3[f'A{index + start_row}'] = t
    ws_3[f'B{index + start_row}'] = pl_ip

for index, (ql_ip, t) in enumerate(ql_ip_r2_scn1):
    ws_3[f'C{index + start_row}'] = ql_ip

pl_iq_r2_scn1 = pq_data_r2_scn1.get_array(column=Vetv.pl_iq,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')
ql_iq_r2_scn1 = pq_data_r2_scn1.get_array(column=Vetv.ql_iq,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')

for index, (pl_iq, t) in enumerate(pl_iq_r2_scn1):
    ws_3[f'D{index + start_row}'] = pl_iq

for index, (ql_iq, t) in enumerate(ql_iq_r2_scn1):
    ws_3[f'E{index + start_row}'] = ql_iq

u_angleU_r2_scn1 = ExportDataRUSTab(rastr_win=Rastr, table=Node.table, switch_command_line=True)
u_r2_scn1_node1 = u_angleU_r2_scn1.get_array(column=Node.vras, key=f'({Node.ny}={node1})')
angleU_r2_scn1_node1 = u_angleU_r2_scn1.get_array(column=Node.delta, key=f'({Node.ny}={node1})')

for index, (u_var_node1, t) in enumerate(u_r2_scn1_node1):
    ws_3[f'F{index + start_row}'] = u_var_node1

for index, (angleU_node1, t) in enumerate(angleU_r2_scn1_node1):
    ws_3[f'G{index + start_row}'] = angleU_node1

u_r2_scn1_node2 = u_angleU_r2_scn1.get_array(column=Node.vras, key=f'({Node.ny}={node2})')
angleU_r2_scn1_node2 = u_angleU_r2_scn1.get_array(column=Node.delta, key=f'({Node.ny}={node2})')

for index, (u_var_node2, t) in enumerate(u_r2_scn1_node2):
    ws_3[f'H{index + start_row}'] = u_var_node2

for index, (angleU_node2, t) in enumerate(angleU_r2_scn1_node2):
    ws_3[f'I{index + start_row}'] = angleU_node2

for i in range(start_row, ws_3.max_row + 1):
    ws_3[f'J{i}'] = f'=I{i}-G{i}'

for i in range(start_row, ws_3.max_row + 1):
    p_2_1_1 = ws_3[f'B{i}'].value
    j = p_2_1_1 * direction_regim_one_ps_one
    ws_3[f'B{i}'] = j

    q_2_1_1 = ws_3[f'C{i}'].value
    j = q_2_1_1 * direction_regim_one_ps_one
    ws_3[f'C{i}'] = j

    p_2_1_2 = ws_3[f'D{i}'].value
    j = p_2_1_2 * direction_regim_one_ps_two
    ws_3[f'D{i}'] = j

    q_2_1_2 = ws_3[f'E{i}'].value
    j = q_2_1_2 * direction_regim_one_ps_two
    ws_3[f'E{i}'] = j

wb_w.save(filename=file_excel)

# 4. ********************* Режим 2 сценарий 2 ******************
load_file(rastr_win=Rastr, file_path=file_rst_regim_two, shabl=shablon_file_dynamic, switch_command_line=True)
load_file(rastr_win=Rastr, file_path=dir_name_scn_two, shabl=shablon_file_scenario, switch_command_line=True)
load_file(rastr_win=Rastr, switch_command_line=True)

rgm_obj.rgm()
dyn_obj.change_calc_time()
dyn_obj.change_snap_max_count()
dyn_obj.run()

pq_data_r2_scn2 = ExportDataRUSTab(rastr_win=Rastr, table='vetv', switch_command_line=True)
pl_ip_r2_scn2 = pq_data_r2_scn2.get_array(column='pl_ip',
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')
ql_ip_r2_scn2 = pq_data_r2_scn2.get_array(column='ql_ip',
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')

for index, (pl_ip, t) in enumerate(pl_ip_r2_scn2):
    ws_4[f'A{index + start_row}'] = t
    ws_4[f'B{index + start_row}'] = pl_ip

for index, (ql_ip, t) in enumerate(ql_ip_r2_scn2):
    ws_4[f'C{index + start_row}'] = ql_ip

pl_iq_r2_scn2 = pq_data_r2_scn2.get_array(column=Vetv.pl_iq,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')
ql_iq_r2_scn2 = pq_data_r2_scn2.get_array(column=Vetv.ql_iq,
                                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')

for index, (pl_iq, t) in enumerate(pl_iq_r2_scn2):
    ws_4[f'D{index + start_row}'] = pl_iq

for index, (ql_iq, t) in enumerate(ql_iq_r2_scn2):
    ws_4[f'E{index + start_row}'] = ql_iq

u_angleU_r2_scn2 = ExportDataRUSTab(rastr_win=Rastr, table=Node.table, switch_command_line=True)
u_r2_scn2_node1 = u_angleU_r2_scn2.get_array(column=Node.vras, key=f'({Node.ny}={node1})')
angleU_r2_scn2_node1 = u_angleU_r2_scn2.get_array(column=Node.delta, key=f'({Node.ny}={node1})')

for index, (u_var_node1, t) in enumerate(u_r2_scn2_node1):
    ws_4[f'F{index + start_row}'] = u_var_node1

for index, (angleU_node1, t) in enumerate(angleU_r2_scn2_node1):
    ws_4[f'G{index + start_row}'] = angleU_node1

u_r2_scn2_node2 = u_angleU_r2_scn2.get_array(column=Node.vras, key=f'({Node.ny}={node2})')
angleU_r2_scn2_node2 = u_angleU_r2_scn2.get_array(column=Node.delta, key=f'({Node.ny}={node2})')

for index, (u_var_node2, t) in enumerate(u_r2_scn2_node2):
    ws_4[f'H{index + start_row}'] = u_var_node2

for index, (angleU_node2, t) in enumerate(angleU_r2_scn2_node2):
    ws_4[f'I{index + start_row}'] = angleU_node2

for i in range(start_row, ws_4.max_row + 1):
    ws_4[f'J{i}'] = f'=I{i}-G{i}'

for i in range(start_row, ws_4.max_row + 1):
    p_2_2_1 = ws_4[f'B{i}'].value
    j = p_2_2_1 * direction_regim_one_ps_one
    ws_4[f'B{i}'] = j

    q_2_2_1 = ws_4[f'C{i}'].value
    j = q_2_2_1 * direction_regim_one_ps_one
    ws_4[f'C{i}'] = j

    p_2_2_2 = ws_4[f'D{i}'].value
    j = p_2_2_2 * direction_regim_one_ps_two
    ws_4[f'D{i}'] = j

    q_2_2_2 = ws_4[f'E{i}'].value
    j = q_2_2_2 * direction_regim_one_ps_two
    ws_4[f'E{i}'] = j

wb_w.save(filename=file_excel)

# ***********************************************
# ************ Построение графиков **************
# ***********************************************

ws_graphs = wb_w['Графики']

for index, i in enumerate(list_coordinates_of_graphs):
    ws_ = wb_w[i[4]]
    # ********** Chart 1 *************
    ch1_ScatterChart = ScatterChart()
    chart_1 = ChartExcelOtherSheet(work_book=wb_w, work_sheet=ws_graphs, chart_obj=ch1_ScatterChart,
                                   chart_title=dict_name_chart[i[4]][0],
                                   x_axis_title='t, c',
                                   y_axis_title='P[МВт]/Q[Мвар]',
                                   width_chart=28,
                                   height_chart=17,
                                   switch_command_line=True,
                                   legend_position='b',
                                   x_axis_min=0, x_axis_max=float(t_ras))

    chart_1.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=2, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][1],
                           width_line_pt=2.0)

    chart_1.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=3, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][2],
                           width_line_pt=2.0)

    chart_1.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=4, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][3],
                           width_line_pt=2.0)

    chart_1.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=5, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][4],
                           width_line_pt=2.0)

    chart_1.print_chart(cell_print_chart=list_coordinates_of_graphs[index][0])

    # ********** Chart 2 *************
    ch2_ScatterChart = ScatterChart()
    chart_2 = ChartExcelOtherSheet(work_book=wb_w, work_sheet=ws_graphs, chart_obj=ch2_ScatterChart,
                                   chart_title=dict_name_chart[i[4]][0],
                                   x_axis_title='t, c',
                                   y_axis_title='U [кВ]',
                                   width_chart=28,
                                   height_chart=17,
                                   switch_command_line=True,
                                   legend_position='b',
                                   x_axis_min=0, x_axis_max=t_ras)

    chart_2.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=6, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][5],
                           width_line_pt=2.0)

    chart_2.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=8, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][7],
                           width_line_pt=2.0)

    chart_2.print_chart(cell_print_chart=list_coordinates_of_graphs[index][1])

    # ********** Chart 3 *************
    ch3_ScatterChart = ScatterChart()
    chart_3 = ChartExcelOtherSheet(work_book=wb_w, work_sheet=ws_graphs, chart_obj=ch3_ScatterChart,
                                   chart_title=dict_name_chart[i[4]][0],
                                   x_axis_title='t, c',
                                   y_axis_title='Delta [град]',
                                   width_chart=28,
                                   height_chart=17,
                                   switch_command_line=True,
                                   legend_position='b',
                                   x_axis_min=0, x_axis_max=t_ras)

    chart_3.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=7, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][6],
                           width_line_pt=2.0)

    chart_3.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=9, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][8],
                           width_line_pt=2.0)

    chart_3.print_chart(cell_print_chart=list_coordinates_of_graphs[index][2])

    # ********** Chart 4 *************
    ch4_ScatterChart = ScatterChart()
    chart_4 = ChartExcelOtherSheet(work_book=wb_w, work_sheet=ws_graphs, chart_obj=ch4_ScatterChart,
                                   chart_title=dict_name_chart[i[4]][0],
                                   x_axis_title='t, c',
                                   y_axis_title='d Delta [град]',
                                   width_chart=28,
                                   height_chart=17,
                                   switch_command_line=True,
                                   legend_position='b',
                                   x_axis_min=0, x_axis_max=t_ras)

    chart_4.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=10, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][9],
                           width_line_pt=2.0)

    chart_4.print_chart(cell_print_chart=list_coordinates_of_graphs[index][3])

wb_w.save(filename=file_excel)

print(f'-------------------- The END -------------------------')
