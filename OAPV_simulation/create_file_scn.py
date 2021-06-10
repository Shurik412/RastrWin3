# -*- coding: utf-8 -*-
from openpyxl import load_workbook

from RastrWinLib.loading.load import load_file
from RastrWinLib.loading.save import save_file
from RastrWinLib.loading.shablon import Shabl
from RastrWinLib.tables.Scenario.DFWAutoActionScn import DFWAutoActionScn
from RastrWinLib.tables.Scenario.DFWAutoLogicScn import DFWAutoLogicScn
from RastrWinLib.variables.variable_parametrs import Variable


def create_file_scn(WorkSheet):
    # 1 строка
    WorkSheet['A14'] = 1  # N
    WorkSheet['B14'] = 1  # N группы
    WorkSheet['C14'] = 'Узел Rш'  # Тип
    WorkSheet['D14'] = WorkSheet['N3'].value  # Название
    WorkSheet['E14'] = WorkSheet['N25'].value  # Формула
    WorkSheet['F14'] = ''  # Тип объекта
    WorkSheet['G14'] = ''  # Свойства объекта
    WorkSheet['H14'] = f'ny={WorkSheet["T25"].value}'  # Ключ объекта

    # 2 строка
    WorkSheet['A15'] = 2
    WorkSheet['B15'] = 1
    WorkSheet['C15'] = 'Узел Xш'
    WorkSheet['D15'] = WorkSheet['N3'].value
    WorkSheet['E15'] = WorkSheet['N26'].value
    WorkSheet['F15'] = ''
    WorkSheet['G15'] = ''
    WorkSheet['H15'] = f'ny={WorkSheet["T25"].value}'

    # 3 строка
    WorkSheet['A16'] = 3
    WorkSheet['B16'] = 2
    WorkSheet['C16'] = 'Узел Rш'
    WorkSheet['D16'] = WorkSheet['N4'].value
    WorkSheet['E16'] = WorkSheet['N35'].value
    WorkSheet['F16'] = ''
    WorkSheet['G16'] = ''
    WorkSheet['H16'] = f'ny={WorkSheet["AE25"].value}'

    # 4 строка
    WorkSheet['A17'] = 4
    WorkSheet['B17'] = 2
    WorkSheet['C17'] = 'Узел Xш'
    WorkSheet['D17'] = WorkSheet['N4'].value
    WorkSheet['E17'] = WorkSheet['N36'].value
    WorkSheet['F17'] = ''
    WorkSheet['G17'] = ''
    WorkSheet['H17'] = f'ny={WorkSheet["AE25"].value}'

    # 5 строка
    WorkSheet['A18'] = 5
    WorkSheet['B18'] = 2
    WorkSheet['C18'] = 'Объект'
    WorkSheet['D18'] = WorkSheet['N2'].value
    WorkSheet['E18'] = WorkSheet['N32'].value
    WorkSheet['F18'] = 'vetv'
    WorkSheet['G18'] = 'r'
    WorkSheet['H18'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'

    # 6 строка
    WorkSheet['A19'] = 6
    WorkSheet['B19'] = 2
    WorkSheet['C19'] = 'Объект'
    WorkSheet['D19'] = WorkSheet['N2'].value
    WorkSheet['E19'] = WorkSheet['N33'].value
    WorkSheet['F19'] = 'vetv'
    WorkSheet['G19'] = 'x'
    WorkSheet['H19'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'

    # 7 строка
    WorkSheet['A20'] = 7
    WorkSheet['B20'] = 2
    WorkSheet['C20'] = 'Объект'
    WorkSheet['D20'] = WorkSheet['N2'].value
    WorkSheet['E20'] = 0
    WorkSheet['F20'] = 'vetv'
    WorkSheet['G20'] = 'b'
    WorkSheet['H20'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'

    # 8 строка
    WorkSheet['A21'] = 8
    WorkSheet['B21'] = 2
    WorkSheet['C21'] = 'Узел Rш'
    WorkSheet['D21'] = WorkSheet['N3'].value
    WorkSheet['E21'] = WorkSheet['N29'].value
    WorkSheet['F21'] = ''
    WorkSheet['G21'] = ''
    WorkSheet['H21'] = f'ny={WorkSheet["T25"].value}'

    # 9 строка
    WorkSheet['A22'] = 9
    WorkSheet['B22'] = 2
    WorkSheet['C22'] = 'Узел Xш'
    WorkSheet['D22'] = WorkSheet['N3'].value
    WorkSheet['E22'] = WorkSheet['N30'].value
    WorkSheet['F22'] = ''
    WorkSheet['G22'] = ''
    WorkSheet['H22'] = f'ny={WorkSheet["T25"].value}'

    # 10 строка
    WorkSheet['A23'] = 10
    WorkSheet['B23'] = 3
    WorkSheet['C23'] = 'Узел Gш'
    WorkSheet['D23'] = WorkSheet['N4'].value
    WorkSheet['E23'] = 0
    WorkSheet['F23'] = ''
    WorkSheet['G23'] = ''
    WorkSheet['H23'] = f'ny={WorkSheet["AE25"].value}'

    # 11 строка
    WorkSheet['A24'] = 11
    WorkSheet['B24'] = 3
    WorkSheet['C24'] = 'Узел Bш'
    WorkSheet['D24'] = WorkSheet['N4'].value
    WorkSheet['E24'] = 0
    WorkSheet['F24'] = ''
    WorkSheet['G24'] = ''
    WorkSheet['H24'] = f'ny={WorkSheet["AE25"].value}'

    # 12 строка
    WorkSheet['A25'] = 12
    WorkSheet['B25'] = 3
    WorkSheet['C25'] = 'Узел Gш'
    WorkSheet['D25'] = WorkSheet['N3'].value
    WorkSheet['E25'] = 0
    WorkSheet['F25'] = ''
    WorkSheet['G25'] = ''
    WorkSheet['H25'] = f'ny={WorkSheet["T25"].value}'

    # 13 строка
    WorkSheet['A26'] = 13
    WorkSheet['B26'] = 3
    WorkSheet['C26'] = 'Узел Bш'
    WorkSheet['D26'] = WorkSheet['N3'].value
    WorkSheet['E26'] = 0
    WorkSheet['F26'] = ''
    WorkSheet['G26'] = ''
    WorkSheet['H26'] = f'ny={WorkSheet["T25"].value}'

    # 14 строка
    WorkSheet['A27'] = 14
    WorkSheet['B27'] = 3
    WorkSheet['C27'] = 'Объект'
    WorkSheet['D27'] = WorkSheet['N2'].value
    WorkSheet['E27'] = WorkSheet['N38'].value
    WorkSheet['F27'] = 'vetv'
    WorkSheet['G27'] = 'r'
    WorkSheet['H27'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'

    # 15 строка
    WorkSheet['A28'] = 15
    WorkSheet['B28'] = 3
    WorkSheet['C28'] = 'Объект'
    WorkSheet['D28'] = WorkSheet['N2'].value
    WorkSheet['E28'] = WorkSheet['N39'].value
    WorkSheet['F28'] = 'vetv'
    WorkSheet['G28'] = 'x'
    WorkSheet['H28'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'

    # 16 строка
    WorkSheet['A29'] = 16
    WorkSheet['B29'] = 4
    WorkSheet['C29'] = 'Объект'
    WorkSheet['D29'] = WorkSheet['N2'].value
    WorkSheet['E29'] = WorkSheet['N8'].value
    WorkSheet['F29'] = 'vetv'
    WorkSheet['G29'] = 'r'
    WorkSheet['H29'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'

    # 17 строка
    WorkSheet['A30'] = 17
    WorkSheet['B30'] = 4
    WorkSheet['C30'] = 'Объект'
    WorkSheet['D30'] = WorkSheet['N2'].value
    WorkSheet['E30'] = WorkSheet['N9'].value
    WorkSheet['F30'] = 'vetv'
    WorkSheet['G30'] = 'x'
    WorkSheet['H30'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'

    # 18 строка
    WorkSheet['A31'] = 18
    WorkSheet['B31'] = 4
    WorkSheet['C31'] = 'Объект'
    WorkSheet['D31'] = WorkSheet['N2'].value
    WorkSheet['E31'] = WorkSheet['N10'].value
    WorkSheet['F31'] = 'vetv'
    WorkSheet['G31'] = 'b'
    WorkSheet['H31'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'

    ##############################################
    # 1 строка
    WorkSheet['A35'] = 1  # N
    WorkSheet['B35'] = 1  # N группы
    WorkSheet['C35'] = 'Узел Rш'  # Тип
    WorkSheet['D35'] = WorkSheet['N4'].value  # Название
    WorkSheet['E35'] = WorkSheet['N43'].value  # Формула
    WorkSheet['F35'] = ''  # Тип объекта
    WorkSheet['G35'] = ''  # Свойства объекта
    WorkSheet['H35'] = f'ny={WorkSheet["AE25"].value}'  # Ключ объекта

    # 2 строка
    WorkSheet['A36'] = 2  # N
    WorkSheet['B36'] = 1  # N группы
    WorkSheet['C36'] = 'Узел Xш'  # Тип
    WorkSheet['D36'] = WorkSheet['N4'].value  # Название
    WorkSheet['E36'] = WorkSheet['N44'].value  # Формула
    WorkSheet['F36'] = ''  # Тип объекта
    WorkSheet['G36'] = ''  # Свойства объекта
    WorkSheet['H36'] = f'ny={WorkSheet["AE25"].value}'  # Ключ объекта

    # 3 строка
    WorkSheet['A37'] = 3  # N
    WorkSheet['B37'] = 2  # N группы
    WorkSheet['C37'] = 'Узел Rш'  # Тип
    WorkSheet['D37'] = WorkSheet['N3'].value  # Название
    WorkSheet['E37'] = WorkSheet['N47'].value  # Формула
    WorkSheet['F37'] = ''  # Тип объекта
    WorkSheet['G37'] = ''  # Свойства объекта
    WorkSheet['H37'] = f'ny={WorkSheet["T25"].value}'  # Ключ объекта

    # 4 строка
    WorkSheet['A38'] = 4  # N
    WorkSheet['B38'] = 2  # N группы
    WorkSheet['C38'] = 'Узел Xш'  # Тип
    WorkSheet['D38'] = WorkSheet['N3'].value  # Название
    WorkSheet['E38'] = WorkSheet['N48'].value  # Формула
    WorkSheet['F38'] = ''  # Тип объекта
    WorkSheet['G38'] = ''  # Свойства объекта
    WorkSheet['H38'] = f'ny={WorkSheet["T25"].value}'  # Ключ объекта

    # 5 строка
    WorkSheet['A39'] = 5  # N
    WorkSheet['B39'] = 2  # N группы
    WorkSheet['C39'] = 'Объект'  # Тип
    WorkSheet['D39'] = WorkSheet['N2'].value  # Название
    WorkSheet['E39'] = WorkSheet['N50'].value  # Формула
    WorkSheet['F39'] = 'vetv'  # Тип объекта
    WorkSheet['G39'] = 'r'  # Свойства объекта
    WorkSheet[
        'H39'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'  # Ключ объекта

    # 6 строка
    WorkSheet['A40'] = 6  # N
    WorkSheet['B40'] = 2  # N группы
    WorkSheet['C40'] = 'Объект'  # Тип
    WorkSheet['D40'] = WorkSheet['N2'].value  # Название
    WorkSheet['E40'] = WorkSheet['N51'].value  # Формула
    WorkSheet['F40'] = 'vetv'  # Тип объекта
    WorkSheet['G40'] = 'x'  # Свойства объекта
    WorkSheet[
        'H40'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'  # Ключ объекта

    # 7 строка
    WorkSheet['A41'] = 7  # N
    WorkSheet['B41'] = 2  # N группы
    WorkSheet['C41'] = 'Объект'  # Тип
    WorkSheet['D41'] = WorkSheet['N2'].value  # Название
    WorkSheet['E41'] = 0  # Формула
    WorkSheet['F41'] = 'vetv'  # Тип объекта
    WorkSheet['G41'] = 'b'  # Свойства объекта
    WorkSheet[
        'H41'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'  # Ключ объекта

    # 8 строка
    WorkSheet['A42'] = 8  # N
    WorkSheet['B42'] = 2  # N группы
    WorkSheet['C42'] = 'Узел Rш'  # Тип
    WorkSheet['D42'] = WorkSheet['N4'].value  # Название
    WorkSheet['E42'] = WorkSheet['N53'].value  # Формула
    WorkSheet['F42'] = ''  # Тип объекта
    WorkSheet['G42'] = ''  # Свойства объекта
    WorkSheet['H42'] = f'ny={WorkSheet["AE25"].value}'  # Ключ объекта

    # 9 строка
    WorkSheet['A43'] = 9  # N
    WorkSheet['B43'] = 2  # N группы
    WorkSheet['C43'] = 'Узел Xш'  # Тип
    WorkSheet['D43'] = WorkSheet['N4'].value  # Название
    WorkSheet['E43'] = WorkSheet['N54'].value  # Формула
    WorkSheet['F43'] = ''  # Тип объекта
    WorkSheet['G43'] = ''  # Свойства объекта
    WorkSheet['H43'] = f'ny={WorkSheet["AE25"].value}'  # Ключ объекта

    # 10 строка
    WorkSheet['A44'] = 10  # N
    WorkSheet['B44'] = 3  # N группы
    WorkSheet['C44'] = 'Узел Gш'  # Тип
    WorkSheet['D44'] = WorkSheet['N3'].value  # Название
    WorkSheet['E44'] = 0  # Формула
    WorkSheet['F44'] = ''  # Тип объекта
    WorkSheet['G44'] = ''  # Свойства объекта
    WorkSheet['H44'] = f'ny={WorkSheet["T25"].value}'  # Ключ объекта

    # 11 строка
    WorkSheet['A45'] = 11  # N
    WorkSheet['B45'] = 3  # N группы
    WorkSheet['C45'] = 'Узел Bш'  # Тип
    WorkSheet['D45'] = WorkSheet['N3'].value  # Название
    WorkSheet['E45'] = 0  # Формула
    WorkSheet['F45'] = ''  # Тип объекта
    WorkSheet['G45'] = ''  # Свойства объекта
    WorkSheet['H45'] = f'ny={WorkSheet["T25"].value}'  # Ключ объекта

    # 12 строка
    WorkSheet['A46'] = 12  # N
    WorkSheet['B46'] = 3  # N группы
    WorkSheet['C46'] = 'Узел Gш'  # Тип
    WorkSheet['D46'] = WorkSheet['N4'].value  # Название
    WorkSheet['E46'] = 0  # Формула
    WorkSheet['F46'] = ''  # Тип объекта
    WorkSheet['G46'] = ''  # Свойства объекта
    WorkSheet['H46'] = f'ny={WorkSheet["AE25"].value}'  # Ключ объекта

    # 13 строка
    WorkSheet['A47'] = 13  # N
    WorkSheet['B47'] = 3  # N группы
    WorkSheet['C47'] = 'Узел Gш'  # Тип
    WorkSheet['D47'] = WorkSheet['N4'].value  # Название
    WorkSheet['E47'] = 0  # Формула
    WorkSheet['F47'] = ''  # Тип объекта
    WorkSheet['G47'] = ''  # Свойства объекта
    WorkSheet['H47'] = f'ny={WorkSheet["AE25"].value}'  # Ключ объекта

    # 14 строка
    WorkSheet['A48'] = 14  # N
    WorkSheet['B48'] = 3  # N группы
    WorkSheet['C48'] = 'Объект'  # Тип
    WorkSheet['D48'] = WorkSheet['N2'].value  # Название
    WorkSheet['E48'] = WorkSheet['N94'].value  # Формула
    WorkSheet['F48'] = 'vetv'  # Тип объекта
    WorkSheet['G48'] = 'r'  # Свойства объекта
    WorkSheet[
        'H48'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'  # Ключ объекта

    # 15 строка
    WorkSheet['A49'] = 15  # N
    WorkSheet['B49'] = 3  # N группы
    WorkSheet['C49'] = 'Объект'  # Тип
    WorkSheet['D49'] = WorkSheet['N2'].value  # Название
    WorkSheet['E49'] = WorkSheet['N95'].value  # Формула
    WorkSheet['F49'] = 'vetv'  # Тип объекта
    WorkSheet['G49'] = 'x'  # Свойства объекта
    WorkSheet[
        'H49'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'  # Ключ объекта

    # 16 строка
    WorkSheet['A50'] = 16  # N
    WorkSheet['B50'] = 4  # N группы
    WorkSheet['C50'] = 'Объект'  # Тип
    WorkSheet['D50'] = WorkSheet['N2'].value  # Название
    WorkSheet['E50'] = WorkSheet['N8'].value  # Формула
    WorkSheet['F50'] = 'vetv'  # Тип объекта
    WorkSheet['G50'] = 'r'  # Свойства объекта
    WorkSheet[
        'H50'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'  # Ключ объекта

    # 17 строка
    WorkSheet['A51'] = 17  # N
    WorkSheet['B51'] = 4  # N группы
    WorkSheet['C51'] = 'Объект'  # Тип
    WorkSheet['D51'] = WorkSheet['N2'].value  # Название
    WorkSheet['E51'] = WorkSheet['N9'].value  # Формула
    WorkSheet['F51'] = 'vetv'  # Тип объекта
    WorkSheet['G51'] = 'x'  # Свойства объекта
    WorkSheet[
        'H51'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'  # Ключ объекта

    # 18 строка
    WorkSheet['A52'] = 18  # N
    WorkSheet['B52'] = 4  # N группы
    WorkSheet['C52'] = 'Объект'  # Тип
    WorkSheet['D52'] = WorkSheet['N2'].value  # Название
    WorkSheet['E52'] = WorkSheet['N10'].value  # Формула
    WorkSheet['F52'] = 'vetv'  # Тип объекта
    WorkSheet['G52'] = 'b'  # Свойства объекта
    WorkSheet[
        'H52'] = f'ip={WorkSheet["X25"].value},iq={WorkSheet["Z25"].value},np={WorkSheet["AB25"].value}'  # Ключ объекта

    ################################################
    WorkSheet['G4'] = WorkSheet['N14'].value
    WorkSheet['G5'] = WorkSheet['N15'].value
    WorkSheet['G6'] = WorkSheet['N16'].value
    WorkSheet['G7'] = (WorkSheet['N17'].value) + (WorkSheet['N14'].value)
    print('Сформирован файл Excel в части сценария!')


class CreateActionsSCN(Variable):
    def __init__(self, rastr_win, dir_name_file_excel, name_list_excel, switch_command_line=False):

        self.rastr_win = rastr_win
        load_file(rastr_win=rastr_win, shabl=Shabl.shablon_file_scenario)
        excel_wb = load_workbook(filename=dir_name_file_excel, data_only=True)
        self.ws = excel_wb[name_list_excel]

        Variable.__init__(self, rastr_win=self.rastr_win,
                          switch_command_line=switch_command_line)

    def create(self, table=DFWAutoActionScn.table, start=14, finish=31):
        table_ = self.rastr_win.Tables(table)
        for index in range(start, finish):
            table_.AddRow()
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.State,
                                      row_id=index - start,
                                      value=0)  # Cocт
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.Id,
                                      row_id=index - start,
                                      value=self.ws[f'A{index}'].value)  # N
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.ParentId,
                                      row_id=index - start,
                                      value=self.ws[f'B{index}'].value)  # N группы
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.Type,
                                      row_id=index - start,
                                      value=self.ws[f'C{index}'].value)  # Тип
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.Name,
                                      row_id=index - start,
                                      value=self.ws[f'D{index}'].value)  # Название
            Variable.make_changes_row(self, table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.Formula,
                                      row_id=index - start,
                                      value=self.ws[f'E{index}'].value)  # Формула
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.ObjectClass,
                                      row_id=index - start,
                                      value=self.ws[f'F{index}'].value)  # Тип объекта
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.ObjectProp,
                                      row_id=index - start,
                                      value=self.ws[f'G{index}'].value)  # Свойство объекта
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.ObjectKey,
                                      row_id=index - start,
                                      value=self.ws[f'H{index}'].value)  # Ключ объекта
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.OutputMode,
                                      row_id=index - start,
                                      value=0)  # Режим
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.RunsCount,
                                      row_id=index - start,
                                      value=1)  # N сраб
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.TimeStart,
                                      row_id=index - start,
                                      value=0)  # Время начала
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.DT,
                                      row_id=index - start,
                                      value=0)  # Длительность
            Variable.make_changes_row(self,
                                      table=DFWAutoActionScn.table,
                                      column=DFWAutoActionScn.Tag,
                                      row_id=index - start,
                                      value=0)  # Тэг упрощенного сценария

    def create_log(self, start=4, finish=8, switch_command_line=False):

        Variable.__init__(self,
                          rastr_win=self.rastr_win,
                          switch_command_line=switch_command_line)

        table_ = self.rastr_win.Tables(DFWAutoLogicScn.table)
        for index in range(start, finish):
            table_.AddRow()
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.Id,
                                      row_id=index - start,
                                      value=self.ws[f'A{index}'].value)  # Id

            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.Name,
                                      row_id=index - start,
                                      value=self.ws[f'B{index}'].value)  # Название
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.ParentId,
                                      row_id=index - start,
                                      value=self.ws[f'C{index}'].value)  # Номер модуля
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.Type,
                                      row_id=index - start,
                                      value=self.ws[f'D{index}'].value)  # Тип логики
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.Formula,
                                      row_id=index - start,
                                      value=self.ws[f'E{index}'].value)  # Формула
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.Actions,
                                      row_id=index - start,
                                      value=self.ws[f'F{index}'].value)  # Действия
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.Delay,
                                      row_id=index - start,
                                      value=self.ws[f'G{index}'].value)  # Выдержка времени
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.UnitStarters,
                                      row_id=index - start,
                                      value=self.ws[f'H{index}'].value)  # ПО мод
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.UnitConstants,
                                      row_id=index - start,
                                      value=self.ws[f'I{index}'].value)  # Const мод
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.UnitActions,
                                      row_id=index - start,
                                      value=self.ws[f'J{index}'].value)  # Дейст мод
            Variable.make_changes_row(self,
                                      table=DFWAutoLogicScn.table,
                                      column=DFWAutoLogicScn.OutputMode,
                                      row_id=index - start,
                                      value=self.ws[f'H{index}'].value)  # Режим выхода

    def save_scn(self, dir_file_name_save_scn=None, dir_file=None, name_save_scn=None, switch_command_line=False):

        if dir_file_name_save_scn is not None:
            save_file(rastr_win=self.rastr_win,
                      file_path=dir_file_name_save_scn,
                      shabl=Shabl.shablon_file_scenario,
                      switch_command_line=switch_command_line)

        elif (dir_file is not None) and (name_save_scn is not None):
            save_file(rastr_win=self.rastr_win,
                      file_path=f'{dir}/{name_save_scn}',
                      shabl=Shabl.shablon_file_scenario,
                      switch_command_line=switch_command_line)


if __name__ == '__main__':
    import win32com.client
    from openpyxl import load_workbook

    wb = load_workbook(
        filename='L:\\SER\\Охрименко\\03. RastrWin3\\19\\ВЛ 500 кВ Нововоронежская АЭС – Воронежская.xlsx')
    ws = wb['Сценарий']

    create_file_scn(WorkSheet=ws)

    wb.save(filename='L:\\SER\\Охрименко\\03. RastrWin3\\19\\ВЛ 500 кВ Нововоронежская АЭС – Воронежская.xlsx')

    # from openpyxl import load_workbook
    # from RastrWinLib.loading.shablons_dir import shablon_file_scenario
    # from RastrWinLib.loading.save_file_rastrwin import save_file
    # from RastrWinLib.loading.load_file_rastrwin import load_file
    # from RastrWinLib.loading.shablon import Shabl
    #
    # rastr = win32com.client.Dispatch('Astra.Rastr')
    # load_file(rastr_win=rastr, shabl=Shabl.shablon_file_scenario)
    # filename = r'L:\SER\Охрименко\03. RastrWin3\19\ВЛ 500 кВ Нововоронежская АЭС – Воронежская.xlsx'
    # scn = CreateActionsSCN(rastr_win=rastr, name_list_excel='Сценарий', dir_name_file_excel=filename, )
    # scn.create()
    # scn.create_log()
    # scn.save_scn(dir_file_name_save_scn=r'L:\SER\Охрименко\03. RastrWin3\19\scn1.scn', switch_command_line=True)
