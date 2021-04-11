from RastrWinLib.excel.chart import ChartExcelOtherSheet
from openpyxl.chart import ScatterChart
from openpyxl import load_workbook

file_excel = r'L:\SER\Охрименко\03. RastrWinLib\16\ВЛ 500 кВ Рязанская ГРЭС – Липецкая Западная.xlsx'

list_coordinates_of_graphs = [('B2', 'S2', 'AJ2', 'BB2', 'Раздел 1.1'),
                              ('B36', 'S36', 'AJ36', 'BB36', 'Раздел 1.2'),
                              ('B70', 'S70', 'AJ70', 'BB70', 'Раздел 2.1'),
                              ('B104', 'S104', 'AJ104', 'BB104', 'Раздел 2.2')]
start_row = 5
wb = load_workbook(filename=file_excel, data_only=True)
ws_settings = wb['Settings Macro']
ws_scn = wb['Сценарий']

dir_name_scn_one = rf'{ws_settings["B7"].value}\{ws_settings["B8"].value}'
dir_name_scn_two = rf'{ws_settings["B7"].value}\{ws_settings["B9"].value}'

# ***** начало, конец и номер параллельности ВЛ(ЛЭП)
ip = ws_scn["X25"].value
iq = ws_scn["Z25"].value
np = ws_scn["AB25"].value

# ***** Узлы ПС1 и ПС2
node1 = ws_scn["T25"].value
node2 = ws_scn["AE25"].value

# ***** Время расчета
t_ras = ws_scn["N20"].value

title_1_1 = wb['Раздел 1.1']['D1'].value
p_ps_one_1_1 = wb['Раздел 1.1']['B3'].value
q_ps_one_1_1 = wb['Раздел 1.1']['C3'].value
p_ps_two_1_1 = wb['Раздел 1.1']['D3'].value
q_ps_two_1_1 = wb['Раздел 1.1']['E3'].value
u_ps_one_1_1 = wb['Раздел 1.1']['F3'].value
ag_u_ps_one_1_1 = wb['Раздел 1.1']['G3'].value
u_ps_two_1_1 = wb['Раздел 1.1']['H3'].value
ag_u_ps_two_1_1 = wb['Раздел 1.1']['I3'].value
d_ag_1_1 = wb['Раздел 1.1']['J3'].value

title_1_2 = wb['Раздел 1.2']['D1'].value
p_ps_one_1_2 = wb['Раздел 1.2']['B3'].value
q_ps_one_1_2 = wb['Раздел 1.2']['C3'].value
p_ps_two_1_2 = wb['Раздел 1.2']['D3'].value
q_ps_two_1_2 = wb['Раздел 1.2']['E3'].value
u_ps_one_1_2 = wb['Раздел 1.2']['F3'].value
ag_u_ps_one_1_2 = wb['Раздел 1.2']['G3'].value
u_ps_two_1_2 = wb['Раздел 1.2']['H3'].value
ag_u_ps_two_1_2 = wb['Раздел 1.2']['I3'].value
d_ag_1_2 = wb['Раздел 1.2']['J3'].value

title_2_1 = wb['Раздел 2.1']['D1'].value
p_ps_one_2_1 = wb['Раздел 2.1']['B3'].value
q_ps_one_2_1 = wb['Раздел 2.1']['C3'].value
p_ps_two_2_1 = wb['Раздел 2.1']['D3'].value
q_ps_two_2_1 = wb['Раздел 2.1']['E3'].value
u_ps_one_2_1 = wb['Раздел 2.1']['F3'].value
ag_u_ps_one_2_1 = wb['Раздел 2.1']['G3'].value
u_ps_two_2_1 = wb['Раздел 2.1']['H3'].value
ag_u_ps_two_2_1 = wb['Раздел 2.1']['I3'].value
d_ag_2_1 = wb['Раздел 2.1']['J3'].value

title_2_2 = wb['Раздел 2.2']['D1'].value
p_ps_one_2_2 = wb['Раздел 2.2']['B3'].value
q_ps_one_2_2 = wb['Раздел 2.2']['C3'].value
p_ps_two_2_2 = wb['Раздел 2.2']['D3'].value
q_ps_two_2_2 = wb['Раздел 2.2']['E3'].value
u_ps_one_2_2 = wb['Раздел 2.2']['F3'].value
ag_u_ps_one_2_2 = wb['Раздел 2.2']['G3'].value
u_ps_two_2_2 = wb['Раздел 2.2']['H3'].value
ag_u_ps_two_2_2 = wb['Раздел 2.2']['I3'].value
d_ag_2_2 = wb['Раздел 2.2']['J3'].value

dict_name_chart = {'Раздел 1.1': (
    title_1_1, p_ps_one_1_1, q_ps_one_1_1, p_ps_two_1_1, q_ps_one_1_1, u_ps_one_1_1, ag_u_ps_one_1_1, u_ps_two_1_1,
    ag_u_ps_two_1_1, d_ag_1_1),
    'Раздел 1.2': (
        title_1_2, p_ps_one_1_2, q_ps_one_1_2, p_ps_two_1_2, q_ps_one_1_2, u_ps_one_1_2, ag_u_ps_one_1_2, u_ps_two_1_2,
        ag_u_ps_two_1_2, d_ag_1_2),
    'Раздел 2.1': (
        title_2_1, p_ps_one_2_1, q_ps_one_2_1, p_ps_two_2_1, q_ps_one_2_1, u_ps_one_2_1, ag_u_ps_one_2_1, u_ps_two_2_1,
        ag_u_ps_two_2_1, d_ag_2_1),
    'Раздел 2.2': (
        title_2_2, p_ps_one_2_2, q_ps_one_2_2, p_ps_two_2_2, q_ps_one_2_2, u_ps_one_2_2, ag_u_ps_one_2_2, u_ps_two_2_2,
        ag_u_ps_two_2_2, d_ag_2_2)}
wb.close()

file_excel = r'L:\SER\Охрименко\03. RastrWinLib\16\ВЛ 500 кВ Рязанская ГРЭС – Липецкая Западная.xlsx'
wb_w = load_workbook(filename=file_excel)
ws_graphs = wb_w['Графики']
set_width_line_pt = 1.5
for index, i in enumerate(list_coordinates_of_graphs):
    ws_ = wb_w[i[4]]
    # ********** Chart 1 *************
    ch1_ScatterChart = ScatterChart()
    chart_1 = ChartExcelOtherSheet(work_book=wb_w, work_sheet=ws_graphs, chart_obj=ch1_ScatterChart,
                                   chart_title=dict_name_chart[i[4]][0],
                                   x_axis_title='t, c',
                                   y_axis_title='P[МВт]/Q[Мвар]',
                                   width_chart=28,
                                   height_chart=17,
                                   switch_command_line=True,
                                   legend_position='b',
                                   x_axis_min=0, x_axis_max=float(t_ras))

    chart_1.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=2, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][1],
                           width_line_pt=set_width_line_pt)

    chart_1.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=3, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][2],
                           width_line_pt=set_width_line_pt)

    chart_1.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=4, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][3],
                           width_line_pt=set_width_line_pt)

    chart_1.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=5, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][4],
                           width_line_pt=set_width_line_pt)

    chart_1.print_chart(cell_print_chart=list_coordinates_of_graphs[index][0])

    # ********** Chart 2 *************
    ch2_ScatterChart = ScatterChart()
    chart_2 = ChartExcelOtherSheet(work_book=wb_w, work_sheet=ws_graphs, chart_obj=ch2_ScatterChart,
                                   chart_title=dict_name_chart[i[4]][0],
                                   x_axis_title='t, c',
                                   y_axis_title='U [кВ]',
                                   width_chart=28,
                                   height_chart=17,
                                   switch_command_line=True,
                                   legend_position='b',
                                   x_axis_min=0, x_axis_max=t_ras)

    chart_2.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=6, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][5],
                           width_line_pt=set_width_line_pt)

    chart_2.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=8, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][7],
                           width_line_pt=set_width_line_pt)

    chart_2.print_chart(cell_print_chart=list_coordinates_of_graphs[index][1])

    # ********** Chart 3 *************
    ch3_ScatterChart = ScatterChart()
    chart_3 = ChartExcelOtherSheet(work_book=wb_w, work_sheet=ws_graphs, chart_obj=ch3_ScatterChart,
                                   chart_title=dict_name_chart[i[4]][0],
                                   x_axis_title='t, c',
                                   y_axis_title='Delta [град]',
                                   width_chart=28,
                                   height_chart=17,
                                   switch_command_line=True,
                                   legend_position='b',
                                   x_axis_min=0, x_axis_max=t_ras)

    chart_3.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=7, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][6],
                           width_line_pt=set_width_line_pt)

    chart_3.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=9, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][8],
                           width_line_pt=set_width_line_pt)

    chart_3.print_chart(cell_print_chart=list_coordinates_of_graphs[index][2])

    # ********** Chart 4 *************
    ch4_ScatterChart = ScatterChart()
    chart_4 = ChartExcelOtherSheet(work_book=wb_w, work_sheet=ws_graphs, chart_obj=ch4_ScatterChart,
                                   chart_title=dict_name_chart[i[4]][0],
                                   x_axis_title='t, c',
                                   y_axis_title='d Delta [град]',
                                   width_chart=28,
                                   height_chart=17,
                                   switch_command_line=True,
                                   legend_position='b',
                                   x_axis_min=0, x_axis_max=t_ras)

    chart_4.add_line_chart(work_sheet_other=ws_.title,
                           min_col_x=1, min_row_x=5, max_row_x=ws_.max_row,
                           min_col_y=10, min_row_y=5, max_row_y=ws_.max_row,
                           title_from_data_ch=False,
                           title_ch=dict_name_chart[i[4]][9],
                           width_line_pt=set_width_line_pt)

    chart_4.print_chart(cell_print_chart=list_coordinates_of_graphs[index][3])

wb_w.save(filename=file_excel)

print(f'-------------------- The END -------------------------')
