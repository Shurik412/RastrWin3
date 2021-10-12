# -*- coding: utf-8 -*-
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border

wb_skuns = load_workbook(filename=r'C:\Users\Ohrimenko_AG\Desktop\BARS\ВетвиСкунс210v2.xlsx')
ws_skuns = wb_skuns["Ветви"]

r_col_start = 9
x_col_start = 10
b_col_start = 11
ktr_col_start = 12

redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')

max_row = 1432

for i in range(3, max_row):
    r = ws_skuns[f'{get_column_letter(9)}{i}'].value
    x = ws_skuns[f'{get_column_letter(10)}{i}'].value
    b = ws_skuns[f'{get_column_letter(11)}{i}'].value
    ktr = ws_skuns[f'{get_column_letter(12)}{i}'].value
    h = 11
    for j in range(1, 99):
        r_col = ws_skuns[f'{get_column_letter(9 + h)}{i}'].value
        x_col = ws_skuns[f'{get_column_letter(10 + h)}{i}'].value
        b_col = ws_skuns[f'{get_column_letter(11 + h)}{i}'].value
        ktr_col = ws_skuns[f'{get_column_letter(12 + h)}{i}'].value

        if r is not None and r_col is not None and r != '' and r_col != '':
            try:
                one_r_col = r_col * 1.05
                two_r_col = 0.95 * r_col
                if one_r_col > r > two_r_col:
                    ws_skuns[f'{get_column_letter(9)}{i}'].fill = redFill
                    ws_skuns[f'{get_column_letter(9 + h)}{i}'].fill = redFill
            except TypeError as te:
                print('TypeError: float() argument must be a string or a number, not NoneType')

        if x is not None and x_col is not None and x != '' and x_col != '':
            try:
                one_x_col = x_col * 1.05
                two_x_col = 0.95 * x_col
                if one_x_col > x > two_x_col:
                    ws_skuns[f'{get_column_letter(10)}{i}'].fill = redFill
                    ws_skuns[f'{get_column_letter(10 + h)}{i}'].fill = redFill
            except TypeError as te:
                print('TypeError: float() argument must be a string or a number, not NoneType')

        if b is not None and b_col is not None and b != '' and b_col != '':
            try:
                one_b_col = b_col * 1.05
                two_b_col = 0.95 * b_col
                if one_b_col > b > two_b_col:
                    ws_skuns[f'{get_column_letter(11)}{i}'].fill = redFill
                    ws_skuns[f'{get_column_letter(11 + h)}{i}'].fill = redFill
            except TypeError as te:
                print('TypeError: float() argument must be a string or a number, not NoneType')

        if ktr is not None and ktr_col is not None and ktr != '' and ktr_col != '':
            try:
                one_ktr_col = ktr_col * 1.05
                two_ktr_col = 0.95 * ktr_col
                if one_ktr_col > float(ktr) > two_ktr_col:
                    ws_skuns[f'{get_column_letter(12)}{i}'].fill = redFill
                    ws_skuns[f'{get_column_letter(12 + h)}{i}'].fill = redFill
            except TypeError as te:
                print('TypeError: float() argument must be a string or a number, not NoneType')

        h += 11

wb_skuns.save(filename=r'C:\Users\Ohrimenko_AG\Desktop\BARS\ВетвиСкунс210v2.xlsx')
