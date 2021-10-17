# -*- coding: utf-8 -*-
from RastrWinLib.AstraRastr import RASTR
from RastrWinLib.Load import load_file
from RastrWinLib.Load.shablon import Shabl
from RastrWinLib.Tables.Vetv.vetv import Vetv
from RastrWinLib.ActionsObject.Get import GettingParameter
from RastrWinLib.Tools.timer import timethis
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def delete_letter(str_one, str_two):
    list_simvol = ['-', '–', ':', 'ПС', 'пс', 'кВ', 'кв', 'вл', 'кл', '.', 'II', 'I']
    flag = True
    flag_one = True
    flag_two = True
    while flag:
        for j in list_simvol:
            if j in str_one:
                str_one.remove(j)
                flag_one = True
            else:
                flag_one = False
        for i in list_simvol:
            if i in str_two:
                str_two.remove(i)
                flag_two = True
            else:
                flag_two = False
        if flag_one == False and flag_two == False:
            flag = False


def percentage_of_coincidence_two_str(string_one, string_two, print_results=False):
    string_one = string_one.lower()
    string_two = string_two.lower()
    str_one = string_one.split(" ")
    str_two = string_two.split(" ")
    # print(f'do -> {str_one}')
    # print(f'do -> {str_two}')
    delete_letter(str_one, str_two)
    # print(f'after -> {str_one}')
    # print(f'after -> {str_two}')
    percent = 100
    percent_of_one_string = percent / len(str_one)
    for index, i in enumerate(str_one):
        if i in str_two:
            if print_results:
                print(f'{index + 1}.Percent = {round(percent)}')
        else:
            percent -= percent_of_one_string
            if print_results:
                print(f'{index + 1}.Percent = {round(percent)}')
    return round(percent)


@timethis
def filling_list_of_data(table, column, rastr_win):
    list_ = []
    getting_parameter_obj = GettingParameter(rastr_win=rastr_win)
    count = getting_parameter_obj.get_count_table(table)
    print(f'count={count}')
    print(f"Количество строк в таблице {table}: {count}")
    for i in range(count):
        char = getting_parameter_obj.get_cell_row(table=table,
                                                  column=column,
                                                  row_id=i)
        list_.append(char)
    return list_


def char_param(table, column, rastr_win, row):
    char = getting_parameter_obj.get_cell_row(table=table,
                                              column=column,
                                              row_id=row)
    return char


def add_new_vetv(name, tip, ip, iq, np, r, x, b, ktr, vetv_list):
    vetv_list.append(
        {"name": name,
         "tip": tip,
         "ip": ip,
         "iq": iq,
         "np": np,
         "r": r,
         "x": x,
         "b": b,
         "ktr": ktr
         },
    )


wb = load_workbook(filename=r'C:\Users\Ohrimenko_AG\Desktop\BARS\14_09_21.xlsx')
ws = wb.active
# load - BARS
load_file(rastr_win=RASTR,
          file_path=r'C:\Users\Ohrimenko_AG\Desktop\BARS\120921-17.mpt',
          shabl=Shabl.shablon_file_mega_mpt,
          switch_command_line=True)

vetv_list_bars = []
getting_parameter_obj = GettingParameter(rastr_win=RASTR)
count_bars_vetv = getting_parameter_obj.get_count_table(Vetv.table)
for i in range(count_bars_vetv):
    name = char_param(table=Vetv.table, column=Vetv.name, rastr_win=RASTR, row=i)
    tip = char_param(table=Vetv.table, column=Vetv.tip, rastr_win=RASTR, row=i)
    ip = char_param(table=Vetv.table, column=Vetv.ip, rastr_win=RASTR, row=i)
    iq = char_param(table=Vetv.table, column=Vetv.iq, rastr_win=RASTR, row=i)
    np = char_param(table=Vetv.table, column=Vetv.np, rastr_win=RASTR, row=i)
    r = char_param(table=Vetv.table, column=Vetv.r, rastr_win=RASTR, row=i)
    x = char_param(table=Vetv.table, column=Vetv.x, rastr_win=RASTR, row=i)
    b = char_param(table=Vetv.table, column=Vetv.b, rastr_win=RASTR, row=i)
    ktr = char_param(table=Vetv.table, column=Vetv.ktr, rastr_win=RASTR, row=i)

    add_new_vetv(name, tip, ip, iq, np, r, x, b, ktr, vetv_list=vetv_list_bars)

# load - SMZU
load_file(rastr_win=RASTR,
          file_path=r'C:\Users\Ohrimenko_AG\Desktop\BARS\smzu_mega.mptsmz',
          shabl=Shabl.without_shabl,
          switch_command_line=True)

vetv_list_smzu = []
count_smzu_vetv = getting_parameter_obj.get_count_table(Vetv.table)
for i in range(count_smzu_vetv):
    name = char_param(table=Vetv.table, column=Vetv.name, rastr_win=RASTR, row=i)
    tip = char_param(table=Vetv.table, column=Vetv.tip, rastr_win=RASTR, row=i)
    ip = char_param(table=Vetv.table, column=Vetv.ip, rastr_win=RASTR, row=i)
    iq = char_param(table=Vetv.table, column=Vetv.iq, rastr_win=RASTR, row=i)
    np = char_param(table=Vetv.table, column=Vetv.np, rastr_win=RASTR, row=i)
    r = char_param(table=Vetv.table, column=Vetv.r, rastr_win=RASTR, row=i)
    x = char_param(table=Vetv.table, column=Vetv.x, rastr_win=RASTR, row=i)
    b = char_param(table=Vetv.table, column=Vetv.b, rastr_win=RASTR, row=i)
    ktr = char_param(table=Vetv.table, column=Vetv.ktr, rastr_win=RASTR, row=i)

    add_new_vetv(name, tip, ip, iq, np, r, x, b, ktr, vetv_list=vetv_list_smzu)

for i in range(count_bars_vetv - 1):
    name_bars = vetv_list_bars[i]['name']
    tip_bars = vetv_list_bars[i]['tip']
    ip_bars = vetv_list_bars[i]['ip']
    iq_bars = vetv_list_bars[i]['iq']
    np_bars = vetv_list_bars[i]['np']
    r_bars = vetv_list_bars[i]['r']
    x_bars = vetv_list_bars[i]['x']
    b_bars = vetv_list_bars[i]['b']
    ktr_bars = vetv_list_bars[i]['ktr']
    g = 3
    ws[f'A{i + g}'] = tip_bars
    ws[f'B{i + g}'] = ip_bars
    ws[f'C{i + g}'] = iq_bars
    ws[f'D{i + g}'] = np_bars
    ws[f'F{i + g}'] = name_bars
    ws[f'G{i + g}'] = r_bars
    ws[f'H{i + g}'] = x_bars
    ws[f'I{i + g}'] = b_bars
    ws[f'J{i + g}'] = ktr_bars

    print(f'\n---------------- {name_bars} -----------------')
    h = 0
    for j in range(len(vetv_list_bars)):
        name_smzu = vetv_list_bars[j]['name']
        tip_smzu = vetv_list_bars[j]['tip']
        ip_smzu = vetv_list_bars[j]['ip']
        iq_smzu = vetv_list_bars[j]['iq']
        np_smzu = vetv_list_bars[j]['np']
        r_smzu = vetv_list_bars[j]['r']
        x_smzu = vetv_list_bars[j]['x']
        b_smzu = vetv_list_bars[j]['b']
        ktr_smzu = vetv_list_bars[j]['ktr']

        percent_name = percentage_of_coincidence_two_str(string_one=name_bars,
                                                         string_two=name_smzu,
                                                         print_results=False)

        if r_bars == r_smzu and percent_name > 49:
            if x_bars == x_smzu:
                if b_bars == b_smzu:
                    if ktr_bars == ktr_smzu:
                        if tip_bars == tip_smzu:
                            print(f'r_bars - {r_bars} => r_smzu - {r_smzu}')
                            print(f'x_bars - {x_bars} => x_smzu - {x_smzu}')
                            print(f'b_bars - {b_bars} => b_smzu - {b_smzu}')
                            print(f'ktr_bars - {ktr_bars} => ktr_smzu - {ktr_smzu}')
                            print(f'tip_bars - {tip_bars} => tip_smzu - {tip_smzu}')
                            print(f'name_bars - {name_bars} => name_smzu - {name_smzu}')

                            ws[f'{get_column_letter(11 + h)}{i + g}'] = tip_smzu
                            ws[f'{get_column_letter(12 + h)}{i + g}'] = ip_smzu
                            ws[f'{get_column_letter(13 + h)}{i + g}'] = iq_smzu
                            ws[f'{get_column_letter(14 + h)}{i + g}'] = np_smzu
                            ws[f'{get_column_letter(16 + h)}{i + g}'] = f'{name_smzu} ({percent_name})'
                            ws[f'{get_column_letter(17 + h)}{i + g}'] = r_smzu
                            ws[f'{get_column_letter(18 + h)}{i + g}'] = x_smzu
                            ws[f'{get_column_letter(19 + h)}{i + g}'] = b_smzu
                            ws[f'{get_column_letter(20 + h)}{i + g}'] = ktr_smzu

                            ws[f'{get_column_letter(11 + h)}{2}'] = 'Тип'
                            ws[f'{get_column_letter(12 + h)}{2}'] = 'N_нач'
                            ws[f'{get_column_letter(13 + h)}{2}'] = 'N_кон'
                            ws[f'{get_column_letter(14 + h)}{2}'] = 'N_п'
                            ws[f'{get_column_letter(15 + h)}{2}'] = 'ID Группы'
                            ws[f'{get_column_letter(16 + h)}{2}'] = 'Название'
                            ws[f'{get_column_letter(17 + h)}{2}'] = 'R'
                            ws[f'{get_column_letter(18 + h)}{2}'] = 'X'
                            ws[f'{get_column_letter(19 + h)}{2}'] = 'B'
                            ws[f'{get_column_letter(20 + h)}{2}'] = 'Кт/r'

                            h += 10

wb.save(filename=r'C:\Users\Ohrimenko_AG\Desktop\BARS\14_09_21.xlsx')
