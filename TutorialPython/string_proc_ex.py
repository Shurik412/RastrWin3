# -*- coding: utf-8 -*-
from RastrWinLib.AstraRastr import RASTR
from RastrWinLib.loading.load import load_file
from RastrWinLib.loading.shablon import Shabl
from RastrWinLib.Tables.Vetv.vetv import Vetv
from RastrWinLib.Getting.get import GettingParameter
from RastrWinLib.Tools.timer import timethis
from icecream import ic
from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter, get_column_interval


def delete_letter(str_one, str_two):
    list_simvol = ['-', '–', ':', 'ПС', 'пс', 'кВ', 'кв', 'вл', 'кл', '.', 'II', 'I', '750', '500', '110', '220', '330',
                   'КВЛ',
                   'цепь', 'ср.т.', '1', '2', '3', '4', '5', '6', '7', '8', '9', '500-1', '500-2', '0-Т.', 'Отп.',
                   'отп.',
                   'сек.', 'ПТ', 'пт', '1с', '2с', 'СШ', 'сш', '1,2', '3,4', 'на', '(220/110)', '(500/110)', 'АТ-1',
                   'ат-1', 'АТ-1', 'ат-2', 'АТ-2', 'ат-3', 'АТ-3', 'ат-4', 'АТ-4', '(т)', '(районная)', '(р)',
                   '(330/110)',
                   '№2', '№1', '№3', '№4', '(новая площадка)', '(нов)', 'отпайкой', 'СЭВ', 'cэв', 'I-II сек.',
                   '(ВЛ-109)',
                   '(вл-109)', 'I,', 'i,', 'ii', 'i', 'сек', '№', 'секц', '1сш', '2сш', '4c', '330 _ФИКТ.',
                   '330 _фикт.',
                   '(КлнАЭС)', '(клнаэс)', '0-т.', '1-4 c', '+',
                   ]

    should_continue = True
    should_continue_j = True
    should_continue_i = True
    while should_continue:
        for j in list_simvol:
            if j in str_one:
                str_one.remove(j)
                should_continue_j = True
            else:
                should_continue_j = False
        for i in list_simvol:
            if i in str_two:
                str_two.remove(i)
                should_continue_i = True
            else:
                should_continue_i = False
        if not should_continue_j:
            if not should_continue_i:
                should_continue = False


def percentage_of_coincidence_two_str(string_one, string_two, print_results=False):
    string_one = string_one.lower()
    string_two = string_two.lower()
    str_one = string_one.split(" ")
    str_two = string_two.split(" ")
    delete_letter(str_one, str_two)
    percent = 100
    percent_of_one_string = percent / len(str_one)
    for index, i in enumerate(str_one):
        if i in str_two:
            if print_results:
                print(f'{index + 1}.Percent = {round(percent)}')
        else:
            percent -= percent_of_one_string
            if print_results:
                print(f'{index + 1}.Percent = {round(percent)}')
    return round(percent)


# @timethis
def filling_list_of_data(table, column, rastr_win):
    list_ = []
    getting_parameter_obj = GettingParameter(rastr_win=rastr_win)
    count = getting_parameter_obj.get_count_table(table)
    print(f'count={count}')
    print(f"Количество строк в таблице {table}: {count}")
    for i in range(count):
        char = getting_parameter_obj.get_cell_row(table=table,
                                                  column=column,
                                                  row_id=i)
        list_.append(char)
    return list_


def char_param(table, column, rastr_win, row):
    char = getting_parameter_obj.get_cell_row(table=table,
                                              column=column,
                                              row_id=row)
    return char


def add_new_vetv(name, tip, ip, iq, np, r, x, b, ktr, vetv_list):
    vetv_list.append(
        {"name": name,
         "tip": tip,
         "ip": ip,
         "iq": iq,
         "np": np,
         "r": r,
         "x": x,
         "b": b,
         "ktr": ktr
         },
    )


def create_data_list(list_vetv, work_sheet, index):
    name = str(work_sheet[f'G{index}'].value)
    tip = str(work_sheet[f'F{index}'].value)
    ip = str(work_sheet[f'B{index}'].value)
    iq = str(work_sheet[f'C{index}'].value)
    np = str(work_sheet[f'D{index}'].value)
    r = str(work_sheet[f'I{index}'].value)
    x = str(work_sheet[f'J{index}'].value)
    b = str(work_sheet[f'K{index}'].value)
    ktr = str(work_sheet[f'L{index}'].value)
    add_new_vetv(name=name, tip=tip, ip=ip, iq=iq,
                 np=np, r=r, x=x, b=b, ktr=ktr,
                 vetv_list=list_vetv)


def create_data_smzu(list_vetv, i):
    name = char_param(table=Vetv.table, column=Vetv.name, rastr_win=RASTR, row=i)
    tip = char_param(table=Vetv.table, column=Vetv.tip, rastr_win=RASTR, row=i)
    ip = char_param(table=Vetv.table, column=Vetv.ip, rastr_win=RASTR, row=i)
    iq = char_param(table=Vetv.table, column=Vetv.iq, rastr_win=RASTR, row=i)
    np = char_param(table=Vetv.table, column=Vetv.np, rastr_win=RASTR, row=i)
    r = char_param(table=Vetv.table, column=Vetv.r, rastr_win=RASTR, row=i)
    x = char_param(table=Vetv.table, column=Vetv.x, rastr_win=RASTR, row=i)
    b = char_param(table=Vetv.table, column=Vetv.b, rastr_win=RASTR, row=i)
    ktr = char_param(table=Vetv.table, column=Vetv.ktr, rastr_win=RASTR, row=i)

    if tip == 0:
        tip = 'ЛЭП'
    elif tip == 1:
        tip = 'Тр'
    elif tip == 2:
        tip = 'Выкл'

    add_new_vetv(name=name, tip=tip, ip=ip,
                 iq=iq, np=np, r=r, x=x, b=b,
                 ktr=ktr, vetv_list=list_vetv)


# load - BARS
wb_skuns = load_workbook(filename=r'C:\Users\Ohrimenko_AG\Desktop\BARS\ВетвиСкунс.xlsx')
ws_skuns = wb_skuns["Ветви"]
vetv_list_bars = []
for i in range(3, 1430):
    create_data_list(list_vetv=vetv_list_bars,
                     work_sheet=ws_skuns,
                     index=i)

# load - SMZU
load_file(rastr_win=RASTR,
          file_path=r'C:\Users\Ohrimenko_AG\Desktop\BARS\smzu_mega.mptsmz',
          shabl=Shabl.without_shabl,
          switch_command_line=True)

vetv_list_smzu = []
getting_parameter_obj = GettingParameter(rastr_win=RASTR)
count_smzu_vetv = getting_parameter_obj.get_count_table(Vetv.table)
for i in range(count_smzu_vetv):
    create_data_smzu(list_vetv=vetv_list_smzu,
                     i=i)

count_bars_vetv = len(vetv_list_bars)
for i in range(count_bars_vetv - 3):
    name_bars = vetv_list_bars[i]['name']
    tip_bars = vetv_list_bars[i]['tip']
    ip_bars = vetv_list_bars[i]['ip']
    iq_bars = vetv_list_bars[i]['iq']
    np_bars = vetv_list_bars[i]['np']
    r_bars = vetv_list_bars[i]['r']
    x_bars = vetv_list_bars[i]['x']
    b_bars = vetv_list_bars[i]['b']
    ktr_bars = vetv_list_bars[i]['ktr']
    g = 3
    h = 0
    n = 2
    print(f'\n---------------- {name_bars} -----------------')
    for j in range(len(vetv_list_smzu)):
        name_smzu = vetv_list_smzu[j]['name']
        tip_smzu = vetv_list_smzu[j]['tip']
        ip_smzu = vetv_list_smzu[j]['ip']
        iq_smzu = vetv_list_smzu[j]['iq']
        np_smzu = vetv_list_smzu[j]['np']
        r_smzu = vetv_list_smzu[j]['r']
        x_smzu = vetv_list_smzu[j]['x']
        b_smzu = vetv_list_smzu[j]['b']
        ktr_smzu = vetv_list_smzu[j]['ktr']

        percent_name = percentage_of_coincidence_two_str(string_one=name_bars,
                                                         string_two=name_smzu,
                                                         print_results=False)

        if percent_name > 50 and tip_bars == tip_smzu:
            print(f'tip_bars - {tip_bars} => tip_smzu - {tip_smzu}')
            print(f'name_bars - {name_bars} => name_smzu - {name_smzu}')
            ws_skuns[f'{get_column_letter(14 + h)}{i + g}'] = tip_smzu
            ws_skuns[f'{get_column_letter(15 + h)}{i + g}'] = ip_smzu
            ws_skuns[f'{get_column_letter(16 + h)}{i + g}'] = iq_smzu
            ws_skuns[f'{get_column_letter(17 + h)}{i + g}'] = np_smzu
            ws_skuns[f'{get_column_letter(18 + h)}{i + g}'] = 0
            ws_skuns[f'{get_column_letter(19 + h)}{i + g}'] = f'{name_smzu} ({percent_name})'
            ws_skuns[f'{get_column_letter(20 + h)}{i + g}'] = r_smzu
            ws_skuns[f'{get_column_letter(21 + h)}{i + g}'] = x_smzu
            ws_skuns[f'{get_column_letter(22 + h)}{i + g}'] = b_smzu
            ws_skuns[f'{get_column_letter(23 + h)}{i + g}'] = ktr_smzu

            ws_skuns[f'{get_column_letter(14 + h)}{n}'] = 'Тип'
            ws_skuns[f'{get_column_letter(15 + h)}{n}'] = 'N_нач'
            ws_skuns[f'{get_column_letter(16 + h)}{n}'] = 'N_кон'
            ws_skuns[f'{get_column_letter(17 + h)}{n}'] = 'N_п'
            ws_skuns[f'{get_column_letter(18 + h)}{n}'] = 'ID Группы'
            ws_skuns[f'{get_column_letter(19 + h)}{n}'] = 'Название'
            ws_skuns[f'{get_column_letter(20 + h)}{n}'] = 'R'
            ws_skuns[f'{get_column_letter(21 + h)}{n}'] = 'X'
            ws_skuns[f'{get_column_letter(22 + h)}{n}'] = 'B'
            ws_skuns[f'{get_column_letter(23 + h)}{n}'] = 'Кт/r'

            h += 11

wb_skuns.save(filename=r'C:\Users\Ohrimenko_AG\Desktop\BARS\ВетвиСкунс210.xlsx')

# print(percent_name)
# if r_bars == r_smzu and percent_name > 49:
#     if x_bars == x_smzu:
#         if b_bars == b_smzu:
#             if ktr_bars == ktr_smzu:
#                 if tip_bars == tip_smzu:
#                     print(f'r_bars - {r_bars} => r_smzu - {r_smzu}')
#                     print(f'x_bars - {x_bars} => x_smzu - {x_smzu}')
#                     print(f'b_bars - {b_bars} => b_smzu - {b_smzu}')
#                     print(f'ktr_bars - {ktr_bars} => ktr_smzu - {ktr_smzu}')
#                     print(f'tip_bars - {tip_bars} => tip_smzu - {tip_smzu}')
#                     print(f'name_bars - {name_bars} => name_smzu - {name_smzu}')
#
#                     ws_skuns[f'{get_column_letter(14 + h)}{i + g}'] = tip_smzu
#                     ws_skuns[f'{get_column_letter(15 + h)}{i + g}'] = ip_smzu
#                     ws_skuns[f'{get_column_letter(16 + h)}{i + g}'] = iq_smzu
#                     ws_skuns[f'{get_column_letter(17 + h)}{i + g}'] = np_smzu
#                     ws_skuns[f'{get_column_letter(18 + h)}{i + g}'] = 0
#                     ws_skuns[f'{get_column_letter(19 + h)}{i + g}'] = f'{name_smzu} ({percent_name})'
#                     ws_skuns[f'{get_column_letter(20 + h)}{i + g}'] = r_smzu
#                     ws_skuns[f'{get_column_letter(21 + h)}{i + g}'] = x_smzu
#                     ws_skuns[f'{get_column_letter(22 + h)}{i + g}'] = b_smzu
#                     ws_skuns[f'{get_column_letter(23 + h)}{i + g}'] = ktr_smzu
#
#                     ws_skuns[f'{get_column_letter(14 + h)}{n}'] = 'Тип'
#                     ws_skuns[f'{get_column_letter(15 + h)}{n}'] = 'N_нач'
#                     ws_skuns[f'{get_column_letter(16 + h)}{n}'] = 'N_кон'
#                     ws_skuns[f'{get_column_letter(17 + h)}{n}'] = 'N_п'
#                     ws_skuns[f'{get_column_letter(18 + h)}{n}'] = 'ID Группы'
#                     ws_skuns[f'{get_column_letter(19 + h)}{n}'] = 'Название'
#                     ws_skuns[f'{get_column_letter(20 + h)}{n}'] = 'R'
#                     ws_skuns[f'{get_column_letter(21 + h)}{n}'] = 'X'
#                     ws_skuns[f'{get_column_letter(22 + h)}{n}'] = 'B'
#                     ws_skuns[f'{get_column_letter(23 + h)}{n}'] = 'Кт/r'
#
#                     h += 11
