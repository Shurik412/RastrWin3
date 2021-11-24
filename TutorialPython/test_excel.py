# -*- coding: utf-8 -*-
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook


def set_cell_range(ws, cell_range_one):
    thick = 'thick'  # жирное выделение
    thin = 'thin'  # точнкое выделение
    excretion = thin
    thin_border = Border(
        left=Side(border_style=excretion),
        right=Side(border_style=excretion),
        top=Side(border_style=excretion),
        bottom=Side(border_style=excretion)
    )
    rows = ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10)
    for row in rows:
        for cell in row:
            cell.border = thin_border


import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
fill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

# create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# populate some sample data
worksheet["A1"] = "Fruit"
worksheet["B1"] = "Color"
worksheet["A2"] = "Apple"
worksheet["B2"] = "Red"
worksheet["A3"] = "Banana"
worksheet["B3"] = "Yellow"
worksheet["A4"] = "Coconut"
worksheet["B4"] = "Brown"
worksheet[f'B4'].fill = fill
# define a table style
mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                      showRowStripes=True)
# create a table
table = openpyxl.worksheet.table.Table(ref='A1:B4',
                                       displayName='FruitColors',
                                       tableStyleInfo=mediumStyle)
# add the table to the worksheet
worksheet.add_table(table)

# save the workbook file
workbook.save('fruit.xlsx')