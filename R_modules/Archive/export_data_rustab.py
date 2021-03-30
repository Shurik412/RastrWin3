# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from os.path import basename, expanduser
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.utils import get_column_letter
from datetime import datetime


class CreateExcelWorkBook:
    """Класс ExcelPlot используется для построения графиков в Excel
        с помощью библиотеки OpenPyXl.
        Основное применение для вышрузки результатов расчетов из RastrWin (RUSTab).

        Attributes
        ----------
        par: dest_file_wb_name : str
                полный путь к директрории где будет сохранияться файл
        par: wb_filename : str
                полный путь к рабочей книге Excel (файлу Excel)
        par: work_sheet : str
                название листа для добавления графика
        par: y_axis_title : str
                название оси Y
        par: scaling_x_max (...x_min, ...y_max, ...y_min)
                масштаб по осям
        par: max_col_x : int
                максимальное значение столбца для ячейки массива значений оси X
        par: min_col_x : int
                минимальное значение столбца для ячейки массива значений оси X
        par: max_row_x : int
                максимальное значение строки для ячейки массива значений оси X
        par: min_row_x : int
                минимальное значение строки для ячейки массива значений оси Х
        par: max_col_y : int
                максимальное значение столбца для ячейки массива значений оси Y
        par: min_col_y : int
                минимальное значение столбца для ячейки массива значений оси Y
        par: max_row_y : int
                максимальное значение строки для ячейки массива значений оси Y
        par: min_row_y : int
                минимальное значение строки для ячейки массива значений оси Y
        Methods
        ---------
        chart()
            строит график по заданными данным.

        """

    def __init__(self, dir_file_name=None, dir_new_file_name=None, name_work_sheet=None, switch_command_line=False):
        self.dir_file_name = dir_file_name
        self.dir_new_file_name = dir_new_file_name
        self.switch = switch_command_line
        self.name_work_sheet = str(name_work_sheet)

    def work_book_sheet(self):
        if self.dir_file_name is not None:
            work_book = load_workbook(filename=self.dir_file_name)
            if self.switch is not False:
                print(f'Была загружена книга Excel: {basename(self.dir_file_name)}.')
        else:
            work_book = Workbook()
            if self.switch is not False:
                print(f'Была новая книга Excel.')
        return work_book

    def workbook_save(self):
        work_book = self.work_book_sheet()
        if self.dir_file_name is not None:
            work_book.save(filename=self.dir_file_name)
            file_name_save = f'{basename(self.dir_file_name)}'
        elif self.dir_new_file_name is not None:
            work_book.save(filename=self.dir_new_file_name)
            file_name_save = f'{basename(self.dir_new_file_name)}'
        else:
            work_book.save(filename=expanduser(f'~/Documents/{datetime.now()}.xlsx'))
            file_name_save = f'{datetime.now()}.xlsx'
        return print(f'Файл {file_name_save} сохранен.')


class ChartsExcel:
    def __init__(self, work_sheet, min_col_x, max_col_x, min_row_x, max_row_x,
                 min_col_y, max_col_y, min_row_y, max_row_y,
                 col_chart=5, row_chart=5, title_chart='График', y_axis_title=None, switch_command_line=False):
        self.switch = bool(switch_command_line)
        self.y_axis_title = y_axis_title
        self.scaling_x_min = None
        self.scaling_x_max = None
        self.scaling_y_min = None
        self.scaling_y_max = None
        self.col_chart = col_chart
        self.row_chart = row_chart

        self.title_chart = str(title_chart)

        self.min_col_x = min_col_x
        self.max_col_x = max_col_x
        self.min_row_x = min_row_x
        self.max_row_x = max_row_x

        self.min_col_y = min_col_y
        self.max_col_y = max_col_y
        self.min_row_y = min_row_y
        self.max_row_y = max_row_y

        self.work_sheet = work_sheet

    def chart(self):
        chart_obj = ScatterChart()
        x_value = Reference(worksheet=self.work_sheet,
                            min_col=self.min_col_x, max_col=self.max_col_x,
                            min_row=self.min_row_x, max_row=self.max_row_x)

        y_value = Reference(worksheet=self.work_sheet,
                            min_col=self.min_col_y, max_col=self.max_col_y,
                            min_row=self.min_row_y, max_row=self.max_row_y)

        series_chart = Series(values=y_value, xvalues=x_value, title_from_data=True)
        chart_obj.title = self.title_chart

        chart_obj.x_axis.title = 'Время, с'

        if self.y_axis_title is not None:
            chart_obj.y_axis.title = self.y_axis_title
        else:
            chart_obj.y_axis.title = 'Параметр'

        chart_obj.append(series_chart)

        if self.scaling_x_min is not None:
            chart_obj.x_axis.scaling.min = self.scaling_x_min
        if self.scaling_x_max is not None:
            chart_obj.x_axis.scaling.max = self.scaling_x_max
        if self.scaling_y_min is not None:
            chart_obj.y_axis.scaling.min = self.scaling_y_min
        if self.scaling_y_max is not None:
            chart_obj.y_axis.scaling.max = self.scaling_y_max

        self.work_sheet.add_chart(chart=chart_obj, anchor=get_column_letter(self.col_chart) + str(self.row_chart))

        if self.switch is not False:
            print(f'Построен график "{self.title_chart}";')
            print(f'\t-на листе: "{self.work_sheet}";')


def sheet(work_book, name_sheet=None, new_sheet=None):
    if name_sheet is not None:
        work_sheet = work_book[str(name_sheet)]
    elif new_sheet is not None:
        work_sheet = work_book.create_sheet(str(new_sheet))
    else:
        work_sheet = work_book.active
    return work_sheet


def workbook_save(work_book, dir_file_name=None, dir_new_file_name=None):
    work_book = work_book
    if dir_file_name is not None:
        work_book.save(filename=dir_file_name)
        file_name_save = f'{basename(dir_file_name)}'
    elif dir_new_file_name is not None:
        work_book.save(filename=dir_new_file_name)
        file_name_save = f'{basename(dir_new_file_name)}'
    else:
        work_book.save(filename=expanduser(f'~/Documents/{datetime.now()}.xlsx'))
        file_name_save = f'{datetime.now()}.xlsx'
    return print(f'Файл {file_name_save} сохранен.')


if __name__ == '__main__':
    wb_class = CreateExcelWorkBook(dir_file_name=r'C:\Users\Ohrimenko_AG\Desktop\ex_r.xlsx',
                                   dir_new_file_name=None, name_work_sheet=None, switch_command_line=False)
    wb = wb_class.work_book_sheet()

    # ws = sheet(work_book=wb, name_sheet=None, new_sheet=None)
    ws = wb.active
    print(ws)
    ch_class = ChartsExcel(work_sheet=ws, y_axis_title='Акт.мощ.[МВт]')
    ch_class.chart()
    workbook_save(work_book=wb, dir_file_name=None, dir_new_file_name=r'C:\Users\Ohrimenko_AG\Desktop\ex_r.xlsx')
