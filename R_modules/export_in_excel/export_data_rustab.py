# -*- coding: utf-8 -*-


class ExportDataRUSTab:
    """
    :param rastr_win => RastrWin3 (Astra.Rastr)
    :param table : str => Name table ('Generator')
    :param column : str => Name parametrs ('P', 'Q')
    :param key : str => ('Num=51700088')
    """

    def __init__(self, rastr_win, table, switch_command_line=False):
        self.table_name = table
        self.rastr_win = rastr_win
        self.table = self.rastr_win.Tables(table)
        self.switch_command_line = switch_command_line

    def get_array(self, column, key):
        self.table.SetSel(key)
        row_id = self.table.FindNextSel(-1)
        if self.switch_command_line is not False:
            print(f'Получены результаты расчета: \n'
                  f'\t\t- из таблицы: "{self.table_name}", параметра "{column}" => "{key}"')
        return self.rastr_win.GetChainedGraphSnapshot(self.table.Name, column, row_id, 0)


if __name__ == '__main__':
    import win32com.client
    from R_modules.load_and_save_file.load_file_rastrwin import load_file
    from openpyxl import Workbook
    from R_modules.load_and_save_file.shablons_dir import shablon_file_dynamic as sh_rst, \
        shablon_file_scenario as sh_scn, shablon_file_automation as sh_dfw
    from R_modules.calculation.dyn_rgm_calc import Dynamic
    from R_modules.directory_rastrwin.dir_test_rastr import file_RUSTab_9_rst, file_RUSTab_9_scn
    from openpyxl.utils import get_column_letter
    from R_modules.variables.variable_parametrs import VariableSetSel

    Rastr = win32com.client.Dispatch('Astra.Rastr')
    wb = Workbook()
    ws = wb.active

    load_file(rastr_win=Rastr, file_path=file_RUSTab_9_rst, shablon=sh_rst, switch_command_line=True)
    load_file(rastr_win=Rastr, file_path=file_RUSTab_9_scn, shablon=sh_scn, switch_command_line=True)
    load_file(rastr_win=Rastr, shablon=sh_dfw, switch_command_line=True)

    dyn = Dynamic(rastr_win=Rastr, switch_command_line=True)

    var3 = VariableSetSel(rastr_win=Rastr, table='Generator', column='P', key='Num=2', switch_command_line=True)
    var3.make_changes(value=150)
    dyn.run()

    var = ExportDataRUSTab(rastr_win=Rastr, table="vetv", switch_command_line=True)
    ip = 1347
    iq = 1320
    np = 0
    var_2 = var.get_array(column='zxe',
                          key=f'(ip={ip} & iq={iq} & np={np})|(ip={iq} & iq={ip} & np={np})')

    for index, (p, t) in enumerate(var_2):
        ws[f'{get_column_letter(4)}{index + 2}'] = t
        ws[f'{get_column_letter(5)}{index + 2}'] = p
        # ws.cell(column=1, row=index + 2).value = t
        # ws.cell(column=2, row=index + 2).value = p

    wb.save(filename='C:\\Users\\Ohrimenko_AG\\Desktop\\21\\777.xlsx')
