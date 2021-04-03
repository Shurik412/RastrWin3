# -*- coding: utf-8 -*-
from openpyxl.chart import Reference, Series
from R_modules.export_in_excel.openpyxl_chart_obj import SCATTER_CHART


class ChartExcel:
    def __init__(self, work_sheet, chart_obj=SCATTER_CHART,
                 chart_title='График', x_axis_title='Ось Х', y_axis_title='Ось Y',
                 width_chart=15, height_chart=7.5,
                 switch_command_line=False):
        self.work_sheet = work_sheet
        self.chart_obj = chart_obj
        self.chart_obj.title = chart_title
        self.chart_obj.x_axis.title = x_axis_title
        self.chart_obj.y_axis.title = y_axis_title
        chart_obj.width = float(width_chart)
        chart_obj.height = float(height_chart)
        self.switch_command_line = switch_command_line
        if switch_command_line is not False:
            print(f'Создан объект графика с названием "{chart_title}", на листе {work_sheet}:')

    def add_line_chart(self, min_col_x, min_row_x, max_row_x,
                       min_col_y, min_row_y, max_row_y, title_from_data_ch=True, title_ch=None,
                       width_line_pt=None):

        if title_from_data_ch is not True:
            min_row_y = min_row_y + 1
        x_values = Reference(self.work_sheet, min_col=min_col_x, min_row=min_row_x, max_row=max_row_x,
                             range_string=None)
        y_values = Reference(self.work_sheet, min_col=min_col_y, min_row=min_row_y, max_row=max_row_y,
                             range_string=None)
        series_xy = Series(y_values, x_values, title_from_data=title_from_data_ch, title=title_ch)

        if width_line_pt is not None:
            series_xy.graphicalProperties.line.width = 12700 * width_line_pt  # 12700 width in EMUs

        self.chart_obj.series.append(series_xy)

        if self.switch_command_line is not False:
            print(f'\t- добавлена линия (кривая) на график;')

    def print_chart(self, cell_print_chart='A10'):
        self.work_sheet.add_chart(self.chart_obj, cell_print_chart)
        if self.switch_command_line is not False:
            print(f'\t\t- график выведен на лист {self.work_sheet}, в ячейку {cell_print_chart}.')


class ChartExcelOtherSheet:
    """
    legend_position: позиция легенды на графике
        - r - справа, l - слева, t - сверху , b - снизу.
    """

    def __init__(self, work_sheet, work_book, chart_obj=SCATTER_CHART,
                 chart_title='График', x_axis_title='Ось Х', y_axis_title='Ось Y',
                 width_chart=15, height_chart=7.5,
                 legend_position='r',
                 x_axis_min=None, x_axis_max=None,
                 y_axis_min=None, y_axis_max=None,
                 switch_command_line=False):
        self.work_book = work_book
        self.work_sheet = work_sheet
        self.chart_obj = chart_obj
        self.chart_obj.title = chart_title
        self.chart_obj.x_axis.title = x_axis_title
        self.chart_obj.y_axis.title = y_axis_title
        chart_obj.width = float(width_chart)
        chart_obj.height = float(height_chart)
        self.switch_command_line = switch_command_line
        self.chart_obj.legend.position = str(legend_position)
        self.chart_obj.x_axis.scaling.min = x_axis_min
        self.chart_obj.x_axis.scaling.max = x_axis_max
        self.chart_obj.y_axis.scaling.min = y_axis_min
        self.chart_obj.y_axis.scaling.max = y_axis_max

        if switch_command_line is not False:
            print(f'Создан объект графика с названием "{chart_title}", на листе {work_sheet}:')

    def add_line_chart(self, min_col_x, min_row_x, max_row_x,
                       min_col_y, min_row_y, max_row_y,
                       work_sheet_other=None, title_from_data_ch=True, title_ch=None,
                       width_line_pt=None):
        if work_sheet_other is None:
            work_sheet_other_ = self.work_sheet
        else:
            work_sheet_other_ = self.work_book[work_sheet_other]

        if title_from_data_ch is not True:
            min_row_y = min_row_y + 1

        x_values = Reference(worksheet=work_sheet_other_, min_col=min_col_x, min_row=min_row_x, max_row=max_row_x,
                             range_string=None)
        y_values = Reference(worksheet=work_sheet_other_, min_col=min_col_y, min_row=min_row_y, max_row=max_row_y,
                             range_string=None)
        series_xy = Series(y_values, x_values, title_from_data=title_from_data_ch, title=title_ch)

        # series_xy.graphicalProperties.line.solidFill = "00AAAA"
        # series_xy.graphicalProperties.line.dashStyle = "sysDot"
        if width_line_pt is not None:
            series_xy.graphicalProperties.line.width = 12700 * width_line_pt  # 12700 width in EMUs

        self.chart_obj.series.append(series_xy)

        if self.switch_command_line is not False:
            print(f'\t- добавлена линия (кривая) на график;')

    def print_chart(self, cell_print_chart='A10'):
        self.work_sheet.add_chart(self.chart_obj, cell_print_chart)
        if self.switch_command_line is not False:
            print(f'\t\t- график выведен на лист {self.work_sheet}, в ячейку {cell_print_chart}.')


if __name__ == '__main__':
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import ScatterChart  # , Reference, Series
    from openpyxl.drawing.drawing import Drawing
    from icecream import ic

    # wb = load_workbook(filename='C:\\Users\\Ohrimenko_AG\\Desktop\\21\\78.xlsx')
    wb = Workbook()
    ws = wb.active
    rows = [
        ['Size', 'Batch 1', 'Batch 2'],
        [2, 40, 30],
        [3, 40, 25],
        [4, 50, 30],
        [5, 30, 25],
        [6, 25, 35],
        [7, 20, 40],
    ]

    for row in rows:
        ws.append(row)

    # ch1 = ScatterChart()
    # ch1.width = 100
    # ch1.height = 100
    # x_values = Reference(ws, min_col=1, min_row=1, max_row=7,
    #                      range_string=None)
    # y_values = Reference(ws, min_col=2, min_row=1, max_row=7,
    #                      range_string=None)
    # series_xy = Series(y_values, x_values, title_from_data=True)
    #
    # ws.add_chart(ch1, "B10")
    # wb.save('C:\\Users\\User\\Desktop\\21\\stop.xlsx')

    dict_test = {1: ['G2', 'Q2'],
                 2: ['G17', 'Q17']}

    for i in dict_test:
        ch22 = ScatterChart()
        ch2 = ChartExcelOtherSheet(work_sheet=ws, chart_obj=ch22, work_book=wb, switch_command_line=True)

        ch2.add_line_chart(min_col_x=1, min_row_x=2, max_row_x=32,
                           min_col_y=2, min_row_y=1, max_row_y=32,
                           work_sheet_other=None, title_from_data_ch=False, title_ch='A2',
                           width_line_pt=1)

        ch2.print_chart(cell_print_chart=dict_test[i][0])

        ch33 = ScatterChart()
        ch3 = ChartExcelOtherSheet(work_sheet=ws, chart_obj=ch33, work_book=wb, switch_command_line=True,
                                   height_chart=10, width_chart=5)

        ch3.add_line_chart(min_col_x=1, min_row_x=2, max_row_x=32,
                           min_col_y=2, min_row_y=1, max_row_y=32,
                           work_sheet_other=None, title_from_data_ch=False, title_ch='A1',
                           width_line_pt=2)

        ch3.print_chart(cell_print_chart=dict_test[i][1])

    wb.save(filename='C:\\Users\\User\\Desktop\\21\\stop.xlsx')
