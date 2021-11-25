# -*- coding: utf-8 -*-
import os

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from win32com.client import Dispatch, WithEvents


class RastrEvents:
    """
    Метод Onprot - выводит сообщения написанные: rastr.Printp("Сообщение из Printp")\n
    Метод OnLog
    """

    def OnLog(self, code, level, id, name, index, description, formName):
        if code == 2:
            print('[Error]', description)
        elif code == 3:
            print('[Warning]', description)
        elif code == 4:
            print('[Lightbulb]', description)
        elif code == 5:
            print('[Info]', description)
        else:
            print([code, description])

    def Onprot(self, message):
        print(message)


RASTR = Dispatch('Astra.Rastr')
WithEvents(RASTR, RastrEvents)

SHABLON_RG2 = r'C:\Program Files\RastrWin3\RastrWin3\SHABLON\режим.rg2'
SHABLON_RST = r'C:\Program Files\RastrWin3\RastrWin3\SHABLON\динамика.rst'

FILE_PATH_RASTR = r'Зимний минимум 2027 минус 31.rst'


def dir_name(file: str):
    path_file = os.path.dirname(file)
    file_basename = os.path.basename(file)
    file_name, file_extension = os.path.splitext(file_basename)
    return path_file, file_name, file_extension


def corr_dname_vetv_model(FILE_PATH_RASTR: str, SHABLON: str):
    path_file, file_name, file_extension = dir_name(file=FILE_PATH_RASTR)
    RASTR.Load(1, FILE_PATH_RASTR, SHABLON)
    vetv = RASTR.Tables("vetv")
    for index_vetv in range(0, vetv.Count):
        dname_old = vetv.Cols("dname").Z(index_vetv)
        dname_old_split_ = dname_old.split()
        for index, i in enumerate(dname_old_split_):
            if "-" in i:
                i_ = i.split("-")
                _i = " - ".join(i_)
                dname_old_split_[index] = _i
                dname_new = " ".join(dname_old_split_)
        dname_new = " ".join(dname_old_split_)
        vetv.Cols("dname").SetZ(index_vetv, dname_new)
    # file_save = path_file + name_file + "" + file_extension
    file_save = FILE_PATH_RASTR
    RASTR.Save(rf'{file_save}', SHABLON)


def simile(dir_file_list: list, sh_list: list) -> None:
    RASTR1 = Dispatch('Astra.Rastr')
    WithEvents(RASTR1, RastrEvents)
    RASTR2 = Dispatch('Astra.Rastr')
    WithEvents(RASTR2, RastrEvents)
    RASTR1.Load(1, dir_file_list[0], sh_list[0])
    RASTR2.Load(1, dir_file_list[1], sh_list[1])
    vetv_1 = RASTR1.Tables("vetv")
    vetv_2 = RASTR2.Tables("vetv")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ветви'
    START_ROW_LOAD = 2
    START_ROW_HEAD = 1
    ws[f'A{START_ROW_HEAD}'] = "Тип"
    ws[f'B{START_ROW_HEAD}'] = "N_нач"
    ws[f'C{START_ROW_HEAD}'] = "N_кон"
    ws[f'D{START_ROW_HEAD}'] = "N_п"
    ws[f'E{START_ROW_HEAD}'] = "Название"
    ws[f'F{START_ROW_HEAD}'] = "Дисп.назв."
    ws[f'G{START_ROW_HEAD}'] = "R,Ом"
    ws[f'H{START_ROW_HEAD}'] = "X,ОМ"
    ws[f'I{START_ROW_HEAD}'] = "B,мкСм"
    ws[f'J{START_ROW_HEAD}'] = "Кт/р"
    ws[f'K{START_ROW_HEAD}'] = "I_доп_обор_ДДТН"
    ws[f'L{START_ROW_HEAD}'] = "I_доп_обор_АДТН"
    ws[f'M{START_ROW_HEAD}'] = "Iдоп_25_ДДТН"
    ws[f'N{START_ROW_HEAD}'] = "Iдоп_25_АДТН"

    for i in range(0, vetv_1.Count):
        ws[f'A{i + START_ROW_LOAD}'] = vetv_1.Cols("tip").Z(i)
        ws[f'B{i + START_ROW_LOAD}'] = vetv_1.Cols("ip").Z(i)
        ws[f'C{i + START_ROW_LOAD}'] = vetv_1.Cols("iq").Z(i)
        ws[f'D{i + START_ROW_LOAD}'] = vetv_1.Cols("np").Z(i)
        ws[f'E{i + START_ROW_LOAD}'] = vetv_1.Cols("name").Z(i)
        ws[f'F{i + START_ROW_LOAD}'] = vetv_1.Cols("dname").Z(i)

        ws[f'G{i + START_ROW_LOAD}'] = vetv_1.Cols("r").Z(i)
        ws[f'H{i + START_ROW_LOAD}'] = vetv_1.Cols("x").Z(i)
        ws[f'I{i + START_ROW_LOAD}'] = vetv_1.Cols("b").Z(i)

        ws[f'J{i + START_ROW_LOAD}'] = vetv_1.Cols("ktr").Z(i)

        ws[f'K{i + START_ROW_LOAD}'] = vetv_1.Cols("i_dop_ob").Z(i)
        ws[f'L{i + START_ROW_LOAD}'] = vetv_1.Cols("i_dop_ob_av").Z(i)

        ws[f'M{i + START_ROW_LOAD}'] = vetv_1.Cols("i_dop").Z(i)
        ws[f'N{i + START_ROW_LOAD}'] = vetv_1.Cols("i_dop_av").Z(i)

    wb.save('ветви.xlsx')


def simile2(dir_file_list: list, sh_list: list):
    RASTR1 = Dispatch('Astra.Rastr')
    WithEvents(RASTR1, RastrEvents)
    RASTR2 = Dispatch('Astra.Rastr')
    WithEvents(RASTR2, RastrEvents)
    RASTR1.Load(1, dir_file_list[0], sh_list[0])
    RASTR2.Load(1, dir_file_list[1], sh_list[1])
    vetv_1 = RASTR1.Tables("vetv")
    vetv_2 = RASTR2.Tables("vetv")
    list1 = []
    list2 = []
    for i in range(0, vetv_1.Count):
        val1 = str(vetv_1.Cols('dname').Z(i))
        list1.append(val1)

    for j in range(0, vetv_2.Count):
        val2 = str(vetv_2.Cols('dname').Z(j))
        list2.append(val2)

    dict_cols_rastr_excel = {
        "tip": "A", "ip": "B", "iq": "C", "np": "D", "name": "E", "dname": "F",
        "r": "G", "x": "H", "b": "I", "ktr": "J", "i_dop_ob": "K", "i_dop_ob_av": "L",
        "i_dop": "M", "i_dop_av": "N"
    }

    wb = load_workbook(filename='ветви.xlsx')
    ws = wb['Ветви']
    fill = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')
    for index, value in enumerate(list1):
        if value != '' or value != '0' or value != ' ':
            if value in list2:
                row_ = list2.index(value)
                for key in dict_cols_rastr_excel.keys():
                    value_old = ws[f"{dict_cols_rastr_excel[key]}{index + 2}"].value
                    value_rastr = vetv_2.Cols(key).Z(row_)
                    if (key != "tip") and (key != "ip") and (key != "iq") and (key != "np") and (key != "name") and (key != "dname"):
                        if value_old != value_rastr:
                            ws[f"{dict_cols_rastr_excel[key]}{index + 2}"].fill = fill
                    value_new = f'{value_old} ({value_rastr})'
                    ws[f"{dict_cols_rastr_excel[key]}{index + 2}"] = value_new

    wb.save(filename='ветви2.xlsx')


dir_file_list = ['Зимний минимум 2027 минус 31.rst', 'БРМ лето 2021 МАКСИМУМ_7.rg2']
sh_list = [SHABLON_RST, SHABLON_RG2]

# for index, name_file in enumerate(dir_file_list):
#     corr_dname_vetv_model(name_file, sh_list[index])

# simile(dir_file_list, sh_list)
simile2(dir_file_list, sh_list)
