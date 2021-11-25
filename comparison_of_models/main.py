# -*- coding: utf-8 -*-
import time
from os import listdir, path

from openpyxl import load_workbook, Workbook
from openpyxl import worksheet
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from Generator import Generator, param_gen, par_gen_x
from Exciter import Exciter, param_exciter, param_exciter_x
from module import CellExcelDyn, RASTR, generator

path_file_name_excel = r'L:\SER\Okhrimenko\10.Проекты\29 (Костромская ГРЭС)\Расчетный период А\Дин_набор ДРМ х64 (08.11.21)_.xlsm'
# path_file_name_rst = r'L:\SER\Okhrimenko\10.Проекты\29 (Костромская ГРЭС)\Расчетный период А\Зимний максимум 2025 минус 31.rst'
SHABLON = r'C:\Program Files\RastrWin3\RastrWin3\SHABLON\динамика.rst'
START_SELL = 3


def changing_number_of_semicolons(number, digits=0):
    try:
        return f"{number:.{digits}f}"
    except TypeError:
        return number


def file_rst():
    list_file = []
    list_filename = []
    for _file in listdir():
        filename, file_ext = path.splitext(_file)
        if file_ext == '.rst':
            list_filename.append(filename)
            list_file.append(_file)
    return list_file, list_filename


def percentage_agreement(arg_1, arg_2):
    try:
        result = (100 - ((100 * float(arg_1)) / float(arg_2)))
    except ZeroDivisionError:
        result = (100 - ((100 * float(arg_1)) / float(0.000001)))
    except TypeError:
        result = 100
    return abs(result)


def unloading_rst_in_excel_gen(file_name_rst: str, file_name: str):
    RASTR.Load(1, file_name_rst, SHABLON)
    _wb = Workbook()
    _ws = _wb.active
    _ws.title = 'Генераторы'
    table_ = RASTR.Tables(Generator.table)
    for index, i_ in enumerate(param_gen):
        _ws[f'{get_column_letter(index + 1)}{1}'].value = i_

    for row in range(0, table_.Count - 1):
        formula_2 = (table_.Cols(Generator.Pnom).Z(row)) / (
                ((table_.Cols(Generator.Ugnom).Z(row)) ** 2) * (table_.Cols(Generator.cosFi).Z(row)))
        for index, j in enumerate(param_gen):
            if j in par_gen_x:
                _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = \
                    changing_number_of_semicolons(((table_.Cols(j).Z(row)) * formula_2), digits=3)
            else:
                if j == "ModelType":
                    if table_.Cols(j).Z(row) == 3:
                        _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = 'Ур.движения'
                    elif table_.Cols(j).Z(row) == 7:
                        _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = '3к-Парк'
                    elif table_.Cols(j).Z(row) == 0:
                        _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = 'АвтоВыбор'
                    elif table_.Cols(j).Z(row) == 2:
                        _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = 'ШБМ'
                    else:
                        _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = table_.Cols(j).Z(row)
                elif j == "Tj=Mj/Pnom":
                    _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = \
                        changing_number_of_semicolons(
                            (table_.Cols(Generator.Mj).Z(row) / table_.Cols(Generator.Pnom).Z(row)),
                            digits=3)
                else:
                    _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = table_.Cols(j).Z(row)
    _wb.save(filename=f'{file_name}.xlsx')


def data_comparison(file_name: str, file_dyn: str):
    wb = load_workbook(filename=file_dyn, data_only=True, keep_vba=False)
    ws = wb['Дин Набор']
    wb_ = load_workbook(filename=f'{file_name}.xlsx', read_only=False, keep_vba=False)
    ws_ = wb_["Генераторы"]
    fill = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')
    counter = 0
    percent = 1
    for i in range(2, ws_.max_row):
        name_ = ws_[f'{get_column_letter(2)}{i}'].value
        try:
            name_list_gen = generator[name_]
            print(name_list_gen)
            counter += 1
            print(f'{counter}. Название: {name_} - список:{name_list_gen};')
            if name_list_gen != '-' or name_list_gen != 'Выведен из эксплуатации':
                model_type_ = ws_[f'{get_column_letter(3)}{i}'].value
                p_nom_ = ws_[f'{get_column_letter(4)}{i}'].value
                ugnom_ = ws_[f'{get_column_letter(5)}{i}'].value
                cosfi_ = ws_[f'{get_column_letter(6)}{i}'].value
                demp_ = ws_[f'{get_column_letter(7)}{i}'].value
                tj_ = ws_[f'{get_column_letter(9)}{i}'].value
                xd1_ = ws_[f'{get_column_letter(10)}{i}'].value
                xd_ = ws_[f'{get_column_letter(11)}{i}'].value
                xq_ = ws_[f'{get_column_letter(12)}{i}'].value
                xd2_ = ws_[f'{get_column_letter(13)}{i}'].value
                xq2_ = ws_[f'{get_column_letter(14)}{i}'].value
                td01_ = ws_[f'{get_column_letter(15)}{i}'].value
                td02_ = ws_[f'{get_column_letter(16)}{i}'].value
                tq02_ = ws_[f'{get_column_letter(17)}{i}'].value
                xq01_ = ws_[f'{get_column_letter(18)}{i}'].value
                xl_ = ws_[f'{get_column_letter(19)}{i}'].value

                for j in range(3, 2235):
                    name_gen_excel = ws[f'{CellExcelDyn.name}{str(j)}'].value
                    if name_gen_excel == name_list_gen:
                        ws_[f'{get_column_letter(2)}{i}'] = f'{name_} ({name_gen_excel})'
                        ws_[
                            f'{get_column_letter(3)}{i}'] = f'{model_type_} ({ws[CellExcelDyn.model_type + str(j)].value})'
                        if percentage_agreement(ws[CellExcelDyn.p_nom + str(j)].value, p_nom_) > percent:
                            ws_[f'{get_column_letter(4)}{i}'] = f'{p_nom_} ({ws[CellExcelDyn.p_nom + str(j)].value})'
                            ws_[f'{get_column_letter(4)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.u_nom + str(j)].value, ugnom_) > percent:
                            ws_[f'{get_column_letter(5)}{i}'] = f'{ugnom_} ({ws[CellExcelDyn.u_nom + str(j)].value})'
                            ws_[f'{get_column_letter(5)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.cosf + str(j)].value, cosfi_) > percent:
                            ws_[f'{get_column_letter(6)}{i}'] = f'{cosfi_} ({ws[CellExcelDyn.cosf + str(j)].value})'
                            ws_[f'{get_column_letter(6)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.k_demp + str(j)].value, demp_) > percent:
                            ws_[f'{get_column_letter(7)}{i}'] = f'{demp_} ({ws[CellExcelDyn.k_demp + str(j)].value})'
                            ws_[f'{get_column_letter(7)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.Tj_p + str(j)].value, tj_) > percent:
                            ws_[f'{get_column_letter(9)}{i}'].fill = fill
                            ws_[f'{get_column_letter(9)}{i}'] = f'{tj_} ' \
                                                                f'({changing_number_of_semicolons(ws[CellExcelDyn.Tj_p + str(j)].value, digits=3)})'
                        if percentage_agreement(ws[CellExcelDyn.xd1 + str(j)].value, xd1_) > percent:
                            ws_[f'{get_column_letter(10)}{i}'] = f'{xd1_} ' \
                                                                 f'({changing_number_of_semicolons(ws[CellExcelDyn.xd1 + str(j)].value, digits=3)})'
                            ws_[f'{get_column_letter(10)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.xd + str(j)].value, xd_) > percent:
                            ws_[f'{get_column_letter(11)}{i}'] = f'{xd_} ' \
                                                                 f'({changing_number_of_semicolons(ws[CellExcelDyn.xd + str(j)].value, digits=3)})'
                            ws_[f'{get_column_letter(11)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.xq + str(j)].value, xq_) > percent:
                            ws_[
                                f'{get_column_letter(12)}{i}'] = f'{xq_} ({changing_number_of_semicolons(ws[CellExcelDyn.xq + str(j)].value, digits=3)})'
                            ws_[f'{get_column_letter(12)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.xd2 + str(j)].value, xd2_) > percent:
                            ws_[
                                f'{get_column_letter(13)}{i}'] = f'{xd2_} ({changing_number_of_semicolons(ws[CellExcelDyn.xd2 + str(j)].value, digits=3)})'
                            ws_[f'{get_column_letter(13)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.xq2 + str(j)].value, xq2_) > percent:
                            ws_[
                                f'{get_column_letter(14)}{i}'] = f'{xq2_} ({changing_number_of_semicolons(ws[CellExcelDyn.xq2 + str(j)].value, digits=3)})'
                            ws_[f'{get_column_letter(14)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.t1d0 + str(j)].value, td01_) > percent:
                            ws_[
                                f'{get_column_letter(15)}{i}'] = f'{td01_} ({changing_number_of_semicolons(ws[CellExcelDyn.t1d0 + str(j)].value, digits=3)})'
                            ws_[f'{get_column_letter(15)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.t2d0 + str(j)].value, td02_) > percent:
                            ws_[
                                f'{get_column_letter(16)}{i}'] = f'{td02_} ({changing_number_of_semicolons(ws[CellExcelDyn.t2d0 + str(j)].value, digits=3)})'
                            ws_[f'{get_column_letter(16)}{i}'].fill = fill
                        if percentage_agreement(ws[CellExcelDyn.t2q0 + str(j)].value, tq02_) > percent:
                            ws_[
                                f'{get_column_letter(17)}{i}'] = f'{tq02_} ({changing_number_of_semicolons(ws[CellExcelDyn.t2q0 + str(j)].value, digits=3)})'
                            ws_[f'{get_column_letter(17)}{i}'].fill = fill
                        ws_[f'{get_column_letter(18)}{i}'] = f'{xq01_} (0)'
                        if percentage_agreement(ws[CellExcelDyn.xs + str(j)].value, xl_) > percent:
                            ws_[
                                f'{get_column_letter(19)}{i}'] = f'{xl_} ({changing_number_of_semicolons(ws[CellExcelDyn.xs + str(j)].value, digits=3)})'
                            ws_[f'{get_column_letter(19)}{i}'].fill = fill
            elif name_list_gen == 'Выведен из эксплуатации':
                ws_[f'{get_column_letter(2)}{i}'] = f'{name_} (Выведен из эксплуатации)'
            else:
                print(f'{name_} - "-"')
        except NameError:
            print(f' - NameError: {name_} - список: ;')
        except KeyError:
            print(f' - KeyError: {name_} - список: ;')

    # define a table style
    mediumStyle = worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                 showRowStripes=True)
    # create a table
    table = worksheet.table.Table(ref=f'A1:S{ws_.max_row}',
                                  displayName='FruitColors',
                                  tableStyleInfo=mediumStyle)
    # add the table to the worksheet
    ws_.add_table(table)

    wb_.save(filename=f'{file_name}.xlsx')


def unloading_rst_in_excel_exiter(file_name_rst: str, file_name: str):
    RASTR.Load(1, file_name_rst, SHABLON)
    _wb = load_workbook(filename=file_name)
    _wb.create_sheet(title='Возбудители')
    _ws = _wb['Возбудители']
    table_ = RASTR.Tables(Exciter.table)
    for index, i_ in enumerate(param_exciter):
        _ws[f'{get_column_letter(index + 1)}{1}'].value = i_
    for row in range(0, table_.Count - 1):
        for index, j in enumerate(param_exciter):
            if j in param_exciter_x:
                _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = \
                    changing_number_of_semicolons((table_.Cols(j).Z(row)), digits=3)
            else:
                _ws[f'{get_column_letter(index + 1)}{row + 2}'].value = table_.Cols(j).Z(row)
    _wb.save(filename=f'{file_name}.xlsx')


start = time.time()
file_name_rst, file_name = file_rst()
for index, file in enumerate(file_name_rst):
    unloading_rst_in_excel_gen(file_name_rst=file, file_name=file_name[index])
    unloading_rst_in_excel_exiter(file_name_rst=file, file_name=file_name[index])
    data_comparison(file_name=file_name[index], file_dyn='Дин_набор ДРМ х64 (08.11.21).xlsm')
end = time.time()
print(f'Время работы: {changing_number_of_semicolons(number=(end - start), digits=1)} сек.')
