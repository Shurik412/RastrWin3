# -*- coding: utf-8 -*-
from win32com.client import Dispatch, WithEvents
from openpyxl import Workbook

area = []

generator = {
    "ГТУ ТЭЦ ЛУЧ  ГТУ-1": "ГТУ ТЭЦ Луч - ГТУ-1",
    "ГТУ ТЭЦ ЛУЧ  ГТУ-2 ": "ГТУ ТЭЦ Луч - ГТУ-2",
    "БЕЛГОРОДСКАЯ ТЭЦ  ГТ-1 ": "Белгородская ТЭЦ - ГТУ-1",
    "БЕЛГОРОДСКАЯ ТЭЦ  ГТ-2  ": "Белгородская ТЭЦ - ГТУ-2",
    "ГТ ТЭЦ МИЧУРИНСКАЯ ГТУ-1": "Мичуринская ГТ-ТЭЦ - ГТУ-1",
    "ГТ ТЭЦ МИЧУРИНСКАЯ ГТУ-3": "Мичуринская ГТ-ТЭЦ - ГТУ-3",
    "ГТ ТЭЦ МИЧУРИНСКАЯ ГТУ-4": "Мичуринская ГТ-ТЭЦ - ГТУ-4",
    "НовАЭС Блок-6": "Нововоронежская АЭС - 10МКА",
    "ТЭЦ-1 ТГ-9": "Воронежская ТЭЦ-1 - ТГ-9",
    "ТЭЦ-1 ТГ-8": "Воронежская ТЭЦ-1 - ТГ-8",
    "ТЭЦ-1 ТГ-7": "Воронежская ТЭЦ-1 - ТГ-7",
    "ТЭЦ-1 ТГ-6": "Воронежская ТЭЦ-1 - ТГ-6",
    "ТЭЦ-1 ТГ-5": "Воронежская ТЭЦ-1 - ТГ-5",
    "ТЭЦ-1 ТГ-4": "Воронежская ТЭЦ-1 - ТГ-4",
    "ТЭЦ-2 ТГ-2": "Воронежская ТЭЦ-2 - ТГ-2",
    "ТЭЦ-2 ГТУ-2": "Воронежская ТЭЦ-2 - ГТУ-2",
    "ТЭЦ-2 ГТУ-1": "Воронежская ТЭЦ-2 - ГТУ-1",
    "НВАЭС ТГ-14": "Нововоронежская АЭС - ТГ-14",
    "НВАЭС ТГ-13": "Нововоронежская АЭС - ТГ-13",
    "НВАЭС ТГ-12": "Нововоронежская АЭС - ТГ-12",
    "НВАЭС ТГ-11": "Нововоронежская АЭС - ТГ-11",
    "НВАЭС ТГ-10": "Выведен из эксплуатации",
    "НВАЭС ТГ-9": "Выведен из эксплуатации",
    "ТЭЦ-2 ТГ-3": "Воронежская ТЭЦ-2 - ТГ-3",
    "Курская ТЭЦ-1-110: ТГ-3 ": "Курская ТЭЦ-1 - ТГ 3",
    "Курская ТЭЦ-1-110: ТГ-4 ": "Курская ТЭЦ-1 - ТГ 4",
    "Курская ТЭЦ-1-110: ТГ-5 ": "-",
    "Орловская ТЭЦ-110: ТГ-5 ": "Орловская ТЭЦ - ТГ-5",
    "Орловская ТЭЦ-110: ТГ-6 ": "Орловская ТЭЦ - ТГ-6",
    "Орловская ТЭЦ-110: ТГ-7 ": "Орловская ТЭЦ - ТГ-7",
    "Орловская ГТ ТЭЦ-10: ГТУ-1 ": "Орловская ГТ-ТЭЦ - ГТУ-1",
    "Орловская ГТ ТЭЦ-10: ГТУ-2": "Орловская ГТ-ТЭЦ - ГТУ-2",
    "Курская АЭС-330: ТГ-1": "Курская АЭС - ТГ-1",
    "Курская АЭС-330: ТГ-2 ": "Курская АЭС - ТГ-2",
    "Курская АЭС-330: ТГ-3 ": "Курская АЭС - ТГ-3",
    "Курская АЭС-330: ТГ-4 ": "Курская АЭС - ТГ-4",
    "Курская АЭС-750: ТГ-5 ": "Курская АЭС - ТГ-5",
    "Курская АЭС-750: ТГ-6 ": "Курская АЭС - ТГ-6",
    "Курская АЭС-750: ТГ-7 ": "Курская АЭС - ТГ-7",
    "Курская АЭС-750: ТГ-8 ": "Курская АЭС - ТГ-8",
    "Курская ТЭЦ КСЗР-110: ТГ-1 ": "Курская ТЭЦ СЗР - 1ГТ",
    "Курская ТЭЦ КСЗР-110: ТГ-2 ": "Курская ТЭЦ СЗР - 2 ГТ",
    "Курская ТЭЦ КСЗР-110: ПТ-3": "Курская ТЭЦ СЗР - 3 ПТ",
    "Курская ТЭЦ-4-6: ТГ-1 ": "Курская ТЭЦ-4 - Г-1",
    "Ливенская ТЭЦ-6: ТГ 1 ": "Ливенская ТЭЦ - ТГ 1",
    "Ливенская ТЭЦ-6: ТГ2 ": "-",
    "Ливенская ТЭЦ-10: ГТ-3 ": "Ливенская ТЭЦ - ГТУ-3",
    "Смоленская АЭС: 1Г": "Смоленская АЭС - 1Г",
    "Калужская ТЭЦ: ТГ-2": "Калужская ТЭЦ - ТГ-2",
    "Калужская ТЭЦ: ТГ-3": "Калужская ТЭЦ - ТГ-3",
    "Смоленская АЭС: 2Г": "Смоленская АЭС - 2Г",
    "Смоленская АЭС: 3Г": "Смоленская АЭС - 3Г",
    "Смоленская АЭС: 4Г": "Смоленская АЭС - 4Г",
    "Смоленская АЭС: 5Г": "Смоленская АЭС - 5Г",
    "Смоленская АЭС: 6Г": "Смоленская АЭС - 6Г",
    "Смоленская ГРЭС: ТГ-1": "Смоленская ГРЭС - ТГ-1",
    "Смоленская ГРЭС: ТГ-2": "Смоленская ГРЭС - ТГ-2",
    "Смоленская ГРЭС: ТГ-3": "Смоленская ГРЭС - ТГ-3",
    "Смоленская ТЭЦ 2: ТГ-1": "Смоленская ТЭЦ-2 - ТГ-1",
    "Смоленская ТЭЦ 2: ТГ-2": "Смоленская ТЭЦ-2 - ТГ-2",
    "Смоленская ТЭЦ 2: ТГ-3": "Смоленская ТЭЦ-2 - ТГ-3",
    "Дорогобужская ТЭЦ: ТГ-1": "Дорогобужская ТЭЦ - ТГ-1",
    "Дорогобужская ТЭЦ: ТГ-2": "-",
    "Дорогобужская ТЭЦ: ТГ-4": "Дорогобужская ТЭЦ - Г-4",
    "Дорогобужская ТЭЦ: Г-1": "Дорогобужская ТЭЦ - Г-1",
    "Дорогобужская ТЭЦ: Г-2": "Дорогобужская ТЭЦ - Г-2",
    "Клинцовская ТЭЦ: ТГ N3": "-",
    "Клинцовская ТЭЦ: ТГ N4": "Клинцовская ТЭЦ: ГПА4",
    "Калужская ТЭЦ: ТГ-4": "Калужская ТЭЦ - ТГ-4",
    "Обнинская ТЭЦ-1: Г1": "Обнинская ТЭЦ-1 - Г-1",
    "ТЭЦ Тулачермет: ТГ-5": "ТЭЦ Тулачермет - ТГ-5",
    "ТЭЦ Тулачермет: ТГ-4": "ТЭЦ Тулачермет - ТГ-4",
    "ТЭЦ Тулачермет: ТГ-3": "ТЭЦ Тулачермет - ТГ-3",
    "ТЭЦ Тулачермет: ТГ-2": "ТЭЦ Тулачермет - ТГ-2",
    "НГРЭС-ПТУ-64": "Новомосковская ГРЭС - ГПТУ №9",
    "НГРЭС-ГТУ-124": "Новомосковская ГРЭС - ГГТУ №8",
    "ЕТЭЦ ТГ-7": "Ефремовская ТЭЦ - ТГ-7",
    "ЕТЭЦ ТГ-6": "Ефремовская ТЭЦ - ТГ-6",
    "ЕТЭЦ ТГ-5": "Ефремовская ТЭЦ - ТГ-5",
    "ЕТЭЦ ТГ-4": "-",
    "ПТЭЦ ТГ-5": "Первомайская ТЭЦ - ТГ-5",
    "ПТЭЦ ТГ-4": "Первомайская ТЭЦ - ТГ-4",
    "ПТЭЦ ТГ-3": "Первомайская ТЭЦ - ТГ-3",
    "ПТЭЦ ТГ-2": "Первомайская ТЭЦ - ТГ-2",
    "ПТЭЦ ТГ-1": "Первомайская ТЭЦ - ТГ-1",
    "Липецкая ТЭЦ ТГ-1": "Липецкая ТЭЦ-2 - ТГ-1",
    "Заводская 2 С Г-1": "-",
    "НГРЭС ТГ-7": "Новомосковская ГРЭС - ТГ-7",
    "НГРЭС ТГ-4": "Новомосковская ГРЭС - ТГ-4",
    "НГРЭС бл.1": "-",
    "ЩГРЭС бл.12": "Щекинская ГРЭС - ТГ-12",
    "ЩГРЭС бл.11": "Щекинская ГРЭС - ТГ-11",
    "ЧГРЭС: бл. 9": "Черепетская ГРЭС - Г-9",
    "ЧГРЭС: бл. 8": "Черепетская ГРЭС - Г-8",
    "Аксеново 2 С Г-1": "-",
    "Липецкая ТЭЦ ТГ-4 ": "Липецкая ТЭЦ-2 - ТГ-4",
    "Липецкая ТЭЦ ТГ-3": "Липецкая ТЭЦ-2 - ТГ-3",
    "Липецкая ТЭЦ ТГ-2 ": "Липецкая ТЭЦ-2 - ТГ-2",
    "Рязанская ГРЭС Блок 2": "Рязанская ГРЭС - ТГ-2",
    "Рязанская ГРЭС Блок 5": "Рязанская ГРЭС - ТГ-5",
    "Рязанская ГРЭС Блок 6": "Рязанская ГРЭС - ТГ-6",
    "Рязанская ГРЭС Блок 1 ТГ-1 ГРЭС-24": "Рязанская ГРЭС - ТГ-1",
    "Рязанская ГРЭС Блок 1 ТГ-2 ГРЭС-24": "Рязанская ГРЭС - ТГ-2",
    "Дягилевская ТЭЦ ТГ-3": "Дягилевская ТЭЦ - ТГ-3",
    "Дягилевская ТЭЦ ТГ-4": "Дягилевская ТЭЦ - ТГ-4",
    "Ново-Рязанская ТЭЦ ТГ-1": "Ново-Рязанская ТЭЦ - ТГ-1",
    "Ново-Рязанская ТЭЦ ТГ-2": "Ново-Рязанская ТЭЦ - ТГ-2",
    "Ново-Рязанская ТЭЦ ТГ-3": "Ново-Рязанская ТЭЦ - ТГ-3",
    "Ново-Рязанская ТЭЦ ТГ-5": "Ново-Рязанская ТЭЦ - ТГ-5",
    "Ново-Рязанская ТЭЦ ТГ-6": "Ново-Рязанская ТЭЦ - ТГ-6",
    "Ново-Рязанская ТЭЦ ТГ-7": "Ново-Рязанская ТЭЦ - ТГ-7",
    "Ново-Рязанская ТЭЦ ТГ-8": "Ново-Рязанская ТЭЦ - ТГ-8",
    "Ново-Рязанская ТЭЦ ТГ-9": "Ново-Рязанская ТЭЦ - ТГ-9",
    "Сасовская ГТ-ТЭЦ ТГ-1": "Сасовская ГТ-ТЭЦ - ТГ-1",
    "Сасовская ГТ-ТЭЦ ТГ-2": "Сасовская ГТ-ТЭЦ - ТГ-2",
    "Касимовская ГТ-ТЭЦ ТГ-1": "Касимовская ГТ-ТЭЦ - ТГ-1",
    "Касимовская ГТ-ТЭЦ ТГ-2": "Касимовская ГТ-ТЭЦ - ТГ-2",
    "ГЭС-1:Блок 27": "ГЭС-1 - ТГ-27",
    "ГЭС-1:Блок 28": "ГЭС-1 - ТГ-28",
    "ГЭС-1:Блок 29": "ГЭС-1 - ТГ-29",
    "ГЭС-1:Блок 30": "ГЭС-1 - ТГ-30",
    "ГЭС-1:Блок 31": "ГЭС-1 - ТГ-31",
    "ЗАГАЭС:Блок 1": "Загорская ГАЭС - А-1",
    "ЗАГАЭС:Блок 2": "Загорская ГАЭС - А-2",
    "ЗАГАЭС:Блок 3": "Загорская ГАЭС - А-3",
    "ЗАГАЭС:Блок 4": "Загорская ГАЭС - А-4",
    "ЗАГАЭС:Блок 5": "Загорская ГАЭС - А-5",
    "ЗАГАЭС:Блок 6": "Загорская ГАЭС - А-6",
    "ПАВЛОВО:ТЭЦ-30:ГТУ 1": "ТЭЦ-30 - ГТУ-1",
    "ПАВЛОВО:ТЭЦ-30:ГТУ 2": "ТЭЦ-30 - ГТУ-2",
    "ЗаГАЭС:дв. реж А-3": "Загорская ГАЭС - А-3",
    "ЗаГАЭС:дв. реж А-4": "Загорская ГАЭС - А-4",
    "ЗаГАЭС:дв. реж А-5": "Загорская ГАЭС - А-5",
    "ЗаГАЭС:дв. реж А-6": "Загорская ГАЭС - А-6",
    "ТЭЦ-16: ТГ-8А": "ТЭЦ-16 - ТГ-8А",
    "ТЭЦ-16: ТГ-8Б": "ТЭЦ-16 - ТГ-8Б",
    "ТЭС МЕЖДУНАРОДНАЯ:ГТУ 1": "ТЭС Международная - Г-1",
    "ТЭС МЕЖДУНАРОДНАЯ:ГТУ 2": "ТЭС Международная - Г-2",
    "ТЭС МЕЖДУНАРОДНАЯ:ГТУ 4": "ТЭС Международная - Г-4",
    "ТЭС МЕЖДУНАРОДНАЯ:ГТУ 5": "ТЭС Международная - Г-5",
    "ТЭС МЕЖДУНАРОДНАЯ:ПТУ 3": "ТЭС Международная - Г-3",
    "ТЭС МЕЖДУНАРОДНАЯ:ПТУ 6": "ТЭС Международная - Г-6",
    "ТЭЦ 507 (Энергоцентр г.Клин): Ямуга": "-",
    "ТЭЦ Воскресенск ПО Минудобрения: Азотная": "-",
    "ТЭЦ МКГЗ: Коксогаз": "-",
    "ТЭЦ-11:ТГ 10": "ТЭЦ-11 - ТГ-10",
    "ТЭЦ-11:ТГ 7": "ТЭЦ-11 - ТГ-7",
    "ТЭЦ-11:ТГ 8": "ТЭЦ-11 - ТГ-8",
    "ТЭЦ-11:ТГ 9": "ТЭЦ-11 - ТГ-9",
    "ТЭЦ-12:ТГ 5": "ТЭЦ-12 - ТГ-5",
    "ТЭЦ-12:ТГ 6": "ТЭЦ-12 - ТГ-6",
    "ТЭЦ-12:ТГ 7": "ТЭЦ-12 - ТГ-7",
    "ТЭЦ-12:ТГ 8": "ТЭЦ-12 - ТГ-8",
    "ТЭЦ-12:ТГ9": "ТЭЦ-12 - ТГ-9",
    "ТЭЦ-16:ТГ 5": "ТЭЦ-16 - ТГ-5",
    "ТЭЦ-16:ТГ 6": "ТЭЦ-16 - ТГ-6",
    "ТЭЦ-16:ТГ 7": "ТЭЦ-16 - ТГ-7",
    "Тамбовская ГТ-ТЭЦ ГТУ-1": "Тамбовская",
    "Тамбовская ТЭЦ ТГ-6": "Тамбовская ТЭЦ - ТГ-6",
    "Тамбовская ТЭЦ ТГ-5": "Тамбовская ТЭЦ - ТГ-5",
    "Тамбовская ТЭЦ ТГ-8": "Тамбовская ТЭЦ - ТГ-8",
    "ТЭЦ-20:ТГ 10": "ТЭЦ-20 - ТГ-10",
    "Котовская ТЭЦ-2 ТГ-4": "-",
    "ТЭЦ-22: ТГ-9": "-",
    "ТЭЦ-20:ТГ 4": "-",
    "Г-2 ТЭС ЦБК": "-",
    "ТЭЦ-20:ТГ 6": "ТЭЦ-20 - ТГ-6",
    "ТЭЦ-20:ТГ 7": "ТЭЦ-20 - ТГ-7",
    "ТЭЦ-20:ТГ 8": "ТЭЦ-20 - ТГ-8",
    "ТЭЦ-20:ТГ 9": "ТЭЦ-20 - ТГ-9",
    "ТЭЦ-21: ГТ 11Б": "ТЭЦ-21 - ТГ-11Б",
    "ТЭЦ-21:Блок 8": "ТЭЦ-21 - ТГ-9",
    "ТЭЦ-21:Блок 9": "ТЭЦ-21 - ТГ-9",
    "ТЭЦ-21:ГТ 11В": "ТЭЦ-21 - ТГ-11В",
    "ТЭЦ-21:ПТ 11А": "ТЭЦ-21 - ТГ-11А",
    "ТЭЦ-21:ТГ 1": "ТЭЦ-21 - ТГ-1",
    "ТЭЦ-21:ТГ 2": "ТЭЦ-21 - ТГ-2",
    "ТЭЦ-21:ТГ 3": "ТЭЦ-21 - ТГ-3",
    "ТЭЦ-21:ТГ 4": "ТЭЦ-21 - ТГ-4",
    "ТЭЦ-21:ТГ 5": "ТЭЦ-21 - ТГ-5",
    "ТЭЦ-21:ТГ 6": "ТЭЦ-21 - ТГ-6",
    "ТЭЦ-21:ТГ 7": "ТЭЦ-21 - ТГ-7",
    "ТЭЦ-21:ТГ10": "ТЭЦ-21 - ТГ-10",
    "ТЭЦ-22:Блок 10": "ТЭЦ-22 - ТГ-10",
    "ТЭЦ-22:Блок 11": "ТЭЦ-22 - ТГ-11",
    "ТЭЦ-22:ТГ 1": "ТЭЦ-22 - ТГ-1",
    "ТЭЦ-22:ТГ 2": "ТЭЦ-22 - ТГ-2",
    "ТЭЦ-22:ТГ 3": "ТЭЦ-22 - ТГ-3",
    "ТЭЦ-22:ТГ 4": "ТЭЦ-22 - ТГ-4",
    "ТЭЦ-22:ТГ 5": "ТЭЦ-22 - ТГ-5",
    "ТЭЦ-22:ТГ 6": "ТЭЦ-22 - ТГ-6",
    "ТЭЦ-22:ТГ 7": "ТЭЦ-22 - ТГ-7",
    "ТЭЦ-22:ТГ 8": "ТЭЦ-22 - ТГ-8",
    "ТЭЦ-23:Блок 5": "ТЭЦ-23 - ТГ-5",
    "ТЭЦ-23:Блок 6": "ТЭЦ-23 - ТГ-6",
    "ТЭЦ-23:Блок 7": "ТЭЦ-23 - ТГ-7",
    "ТЭЦ-23:Блок 8": "ТЭЦ-23 - ТГ-8",
    "ТЭЦ-23:ТГ 1": "ТЭЦ-23 - ТГ-1",
    "ТЭЦ-23:ТГ 2": "ТЭЦ-23 - ТГ-2",
    "ТЭЦ-23:ТГ 3": "ТЭЦ-23 - ТГ-3",
    "ТЭЦ-23:ТГ 4": "ТЭЦ-23 - ТГ-4",
    "ТЭЦ-25:Блок 3": "ТЭЦ-25 - ТГ-3",
    "ТЭЦ-25:Блок 4": "ТЭЦ-25 - ТГ-4",
    "ТЭЦ-25:Блок 5": "ТЭЦ-25 - ТГ-5",
    "ТЭЦ-25:Блок 6": "ТЭЦ-25 - ТГ-6",
    "ТЭЦ-25:Блок 7": "ТЭЦ-25 - ТГ-7",
    "ТЭЦ-25:ТГ 1": "ТЭЦ-25 - ТГ-1",
    "ТЭЦ-25:ТГ 2": "ТЭЦ-25 - ТГ-2",
    "ТЭЦ-26: ГТ 8А": "ТЭЦ-26 - Г-8А",
    "ТЭЦ-26: ПТ 8Б": "ТЭЦ-26 - Г-8Б",
    "ТЭЦ-26:Блок 1": "ТЭЦ-26 - Г-1",
    "ТЭЦ-26:Блок 2": "ТЭЦ-26 - Г-2",
    "ТЭЦ-26:Блок 3": "ТЭЦ-26 - Г-3",
    "ТЭЦ-26:Блок 4": "ТЭЦ-26 - Г-4",
    "ТЭЦ-26:Блок 5": "ТЭЦ-26 - Г-5",
    "ТЭЦ-26:Блок 6": "ТЭЦ-26 - Г-6",
    "ТЭЦ-26:Блок 7": "ТЭЦ-26 - Г-7",
    "ТЭЦ-27:Блок 1": "ТЭЦ-27 - Г-1",
    "ТЭЦ-27:Блок 2": "ТЭЦ-27 - Г-2",
    "ТЭЦ-27:ГТ 31": "ТЭЦ-27 - Г-31",
    "ТЭЦ-27:ГТ 32": "ТЭЦ-27 - Г-32",
    "ТЭЦ-27:ГТ 41": "ТЭЦ-27 - Г-41",
    "ТЭЦ-27:ГТ 42": "ТЭЦ-27 - Г-42",
    "ТЭЦ-27:ПТ 33": "ТЭЦ-27 - Г-33",
    "ТЭЦ-27:ПТ 43": "ТЭЦ-27 - Г-43",
    "ТЭЦ-29: ГТУ-1": "ТЭЦ-29 - Г-1",
    "Солнечногорск 2 С Г-1": "-",
    "Свистягино 2 С Г-1": "-",
    "ТЭЦ-8:ТГ 10": "ТЭЦ-8 - ТГ-10",
    "ТЭЦ-8:ТГ 11": "ТЭЦ-8 - ТГ-11",
    "ТЭЦ-8:ТГ 6": "ТЭЦ-8 - ТГ-6",
    "ТЭЦ-8:ТГ 7": "ТЭЦ-8 - ТГ-7",
    "ТЭЦ-8:ТГ 8": "ТЭЦ-8 - ТГ-8",
    "ТЭЦ-8:ТГ 9": "ТЭЦ-8 - ТГ-9",
    "ТЭЦ-9:ТГ 4": "ТЭЦ-9 - ТГ-4",
    "ТЭЦ-9:ТГ 5": "ТЭЦ-9 - ТГ-5",
    "ТЭЦ-9:ТГ 7": "ТЭЦ-9 - ТГ-7",
    "ГЭС Иваньково Г-1 10кВ": "Иваньковская ГЭС - Г-1",
    "ГЭС Иваньково Г-2 10кВ ": "Иваньковская ГЭС - Г-2",
    "ТЭЦ-9: ТГ-1": "ТЭЦ-9 - ТГ-1",
    "ЗаГАЭС:дв. реж А-1": "Загорская ГАЭС - А-1",
    "ЗаГАЭС:дв. реж А-2": "Загорская ГАЭС - А-2",
    "ТЭЦ-12: ТГ-1Б": "ТЭЦ-12 - ТГ-1Б",
    "ТЭЦ-12: ТГ-1А": "ТЭЦ-12 - ТГ-1А",
    "ТЭЦ-20: ТГ-11Б": "ТЭЦ-20 - ТГ-11Б",
    "ТЭЦ-20: ТГ-11А": "ТЭЦ-20 - ТГ-11А",
    "Тамбовская ТЭЦ ТГ-7": "Тамбовская ТЭЦ - ТГ-7",
    "Елецкая ТЭЦ ТГ-4 ": "Елецкая ТЭЦ - ТГ-4",
    "Данковская ТЭЦ ТГ-2 ": "Данковская ТЭЦ - ТГ-2",
    "Данковская ТЭЦ ТГ-1 ": "Данковская ТЭЦ - ТГ-1",
    "Елецкая ТЭЦ ТГ-5 ПГУ ": "-",
    "Елецкая ТЭЦ ГТ-2 ПГУ ": "-",
    "Елецкая ТЭЦ ГТ-1 ПГУ ": "-",
    "Липецкая ТЭЦ ТГ-5": "Липецкая ТЭЦ-2 - ТГ-5",
    "ТЭЦ НЛМК ТГ-8 ": "ТЭЦ НЛМК - ТГ-8",
    "ГТРС НЛМК ГУБТ-2": "ГТРС - ГУБТ-2",
    "ГТРС НЛМК ГУБТ-1": "ГТРС - ГУБТ-1",
    "Тамбовская ГТ-ТЭЦ ГТУ-2": "Тамбовская ГТ-ТЭЦ - Г2",
    "Шатурская ГРЭС:Блок 1": "Шатурская ГРЭС - ТГ-1",
    "Шатурская ГРЭС:Блок 2": "Шатурская ГРЭС - ТГ-2",
    "Шатурская ГРЭС:Блок 3": "Шатурская ГРЭС - ТГ-3",
    "Шатурская ГРЭС:Блок 4": "Шатурская ГРЭС - ТГ-4",
    "Шатурская ГРЭС:Блок 5": "Шатурская ГРЭС - ТГ-5",
    "Шатурская ГРЭС:Блок 6": "Шатурская ГРЭС - ТГ-6",
    "Шатурская ГРЭС:Блок 7": "Шатурская ГРЭС - ТГ-7",
    "ГТЭС КОЛОМЕНСКОЕ:Блок 1": "ГТЭС Коломенское - ГТУ-1",
    "ГТЭС КОЛОМЕНСКОЕ:Блок 2": "ГТЭС Коломенское - ГТУ-2",
    "ГТЭС КОЛОМЕНСКОЕ:Блок 3": "ГТЭС Коломенское - ГТУ-3",
    "ГТЭС Постниково: ГТУ-1": "ГТЭС Постниково - ТГ-1",
    "ГТЭС Постниково:ГТУ2": "ГТЭС Постниково - ТГ-2",
    "ГТЭС Терешково: Г-1": "ГТЭС Терешково - ГТУ-1",
    "ГТЭС Терешково: Г-2": "ГТЭС Терешково - ГТУ-2",
    "ГТЭС Терешково: Г-3": "ГТЭС Терешково - ГТУ-3",
    "ГТЭС Терешково:Г-4": "ГТЭС Терешково - ПТ-4",
    "ГТЭЦ Щелково: 1": "Щёлковская ГТ-ТЭЦ - ГТУ-1",
    "ГТЭЦ Щелково: 2": "Щёлковская ГТ-ТЭЦ - ГТУ-2",
    "ГЭС Сходня Г-1 10кВ: Тушино": "Сходненская ГЭС - Г-1",
    "ГЭС Сходня Г-2 10кВ: Тушино": "Сходненская ГЭС - Г-2",
    "ГЭС-1:Блок 26": "ГЭС-1 - ТГ-26",
    "Мини ТЭС Курьяново": "Мини ТЭС Курьяново - Экв. ген. Мини ТЭС Курьяново",
    "ТЭЦ МЭИ": "ТЭЦ МЭИ - экв. ген. Г-1,2",
    "ГЭС Карамышевская": "ГЭС Карамышевская - ГЭС Карамышевская Г-1",
    "Конаковская ГРЭС_Блок 1": "Конаковская ГРЭС - ТГ-1",
    "Конаковская ГРЭС_Блок 2": "Конаковская ГРЭС - ТГ-2",
    "Конаковская ГРЭС_Блок 3": "Конаковская ГРЭС - ТГ-3",
    "Конаковская ГРЭС_Блок 4": "Конаковская ГРЭС - ТГ-4",
    "Конаковская ГРЭС_Блок 5": "Конаковская ГРЭС - ТГ-5",
    "Конаковская ГРЭС_Блок 6": "Конаковская ГРЭС - ТГ-6",
    "Конаковская ГРЭС_Блок 7": "Конаковская ГРЭС - ТГ-7",
    "Конаковская ГРЭС_Блок 8": "Конаковская ГРЭС - ТГ-8",
    "Калининская АЭС_Блок 1": "Калининская АЭС - 1GT (Г-1)",
    "Калининская АЭС_Блок 2": "Калининская АЭС - 2GT (Г-2)",
    "Калининская АЭС_Блок 3": "Калининская АЭС - 3GT (Г-3)",
    "Калининская АЭС_Блок 4": "Калининская АЭС - 4GT (Г-4)",
    "Тверская ТЭЦ-3_ТГ-1": "Тверская ТЭЦ-3 - ТГ-1",
    "Тверская ТЭЦ-3_ТГ-2": "Тверская ТЭЦ-3 - ТГ-2",
    "Тверская ТЭЦ-4_ТГ-1": "Тверская ТЭЦ-4 - ТГ-1",
    "Тверская ТЭЦ-4_ТГ-3": "Тверская ТЭЦ-4 - ТГ-3",
    "Тверская ТЭЦ-4_ТГ-4": "Тверская ТЭЦ-4 - ТГ-4",
    "Тверская ТЭЦ-4_ТГ-5": "Тверская ТЭЦ-4 - ТГ-5",
    "Тверская ТЭЦ-4_ТГ-7": "Тверская ТЭЦ-4 - ТГ-7",
    "Тверская ТЭЦ-1_ТГ-4": "Тверская ТЭЦ-1 - ТГ-4",
    "Новотверецкая ГЭС_ГГ": "Новотверецкая ГЭС - ГГ-2",
    "Вышневолоцкая ТЭЦ_Г-1": "-",
    "КС-16 Юбилейного ЛПУМГ ТГ-3": "ПС 110 кВ Погорелово - Г-3 ЭСН КС16 Юбилейного ЛПУМГ",
    "КС-16 Юбилейного ЛПУМГ ТГ-2": "ПС 110 кВ Погорелово - Г-2 ЭСН КС16 Юбилейного ЛПУМГ",
    "КС-16 Юбилейного ЛПУМГ ТГ-1": "ПС 110 кВ Погорелово - Г-1 ЭСН КС16 Юбилейного ЛПУМГ",
    '"ОАО ""Северсталь"" УЭС ТСЦ "': 'ОАО "Северсталь" УЭС ТСЦ ',
    "КС-15 Нюксенского ЛПУМГ ТГ-3": "ПС 110 кВ Погорелово - Г-3 ЭСН КС16 Юбилейного ЛПУМГ",
    "КС-15 Нюксенского ЛПУМГ ТГ-2": "ПС 110 кВ Погорелово - Г-2 ЭСН КС16 Юбилейного ЛПУМГ",
    "КС-15 Нюксенского ЛПУМГ ТГ-1": "ПС 110 кВ Погорелово - Г-1 ЭСН КС16 Юбилейного ЛПУМГ",
    "ПГУ-420 ЧГРЭС": "Череповецкая ГРЭС - ТГ 4 Блок 4",
    "Вологодская ТЭЦ Г-2 ПГУ-110": "Вологодская ТЭЦ - Г-2 ПГУ",
    "Вологодская ТЭЦ Г-1 ПГУ-110": "Вологодская ТЭЦ - Г-1 ПГУ",
    '"ГТЭС ""ФосАгро-Череповец"" ТГ№1"': "ГТЭС ФосАгро-Череповец - ТГ-1",
    "Вологодская ТЭЦ ТГ-3": "Вологодская ТЭЦ - ТГ-3",
    "Вологодская ТЭЦ ТГ-2": "Вологодская ТЭЦ - ТГ-2",
    "Вологодская ТЭЦ ТГ-1": "Вологодская ТЭЦ - ТГ-1",
    "Шекснинская ГЭС 1Г, 2Г, 3Г, 4Г": "-",
    '"ОАО ""Северсталь"" ГУБТ-25"': "ПС 220 кВ ГПП-6 - ГУБТ-25",
    '"ТЭЦ ПВС ""Северсталь"" ТГ-7"': "ТЭЦ-ПВС - ТГ-7",
    '"ТЭЦ ПВС ""Северсталь"" ТГ-6"': "ТЭЦ-ПВС - ТГ-6",
    '"ТЭЦ ПВС ""Северсталь"" ТГ-5"': "ТЭЦ-ПВС - ТГ-5",
    '"ТЭЦ ПВС ""Северсталь"" ТГ-4"': "ТЭЦ-ПВС - ТГ-4",
    '"ТЭЦ ПВС ""Северсталь"" ТГ-3"': "ТЭЦ-ПВС - ТГ-3",
    '"ТЭЦ ПВС ""Северсталь"" ТГ-2"': "ТЭЦ-ПВС - ТГ-2",
    '"ТЭЦ ПВС ""Северсталь"" ТГ-1"': "ТЭЦ-ПВС - ТГ-1",
    "Красавинская ГТ ТЭЦ ГТ-2": "Красавинская ГТ ТЭЦ - ГТ-2",
    "Красавинская ГТ ТЭЦ ГТ-1": "Красавинская ГТ ТЭЦ - ГТ-1",
    "Красавинская ГТ ТЭЦ ПТ-4": "Красавинская ГТ ТЭЦ - ПТ-4",
    "Красавинская ГТ ТЭЦ ГТ-3": "Красавинская ГТ ТЭЦ - ГТ-3",
    '"ТЭЦ ""ФосАгро-Череповец"" ТГ 7"': "ТЭЦ ФосАгро – Череповец - ТГ-7",
    '"ТЭЦ ""ФосАгро-Череповец"" ТГ 6"': "ТЭЦ ФосАгро – Череповец - ТГ-6",
    '"ТЭЦ ""ФосАгро-Череповец"" ТГ 5"': "ТЭЦ ФосАгро – Череповец - ТГ-5",
    '"ТЭЦ ""ФосАгро-Череповец"" ТГ 4"': "ТЭЦ ФосАгро – Череповец - ТГ-4",
    '"ТЭЦ ""ФосАгро-Череповец"" ТГ 3"': "ТЭЦ ФосАгро – Череповец - ТГ-3",
    '"ТЭЦ ""ФосАгро-Череповец"" ТГ 2"': "ТЭЦ ФосАгро – Череповец - ТГ-2",
    '"ТЭЦ ""ФосАгро-Череповец"" ТГ 1"': "ТЭЦ ФосАгро – Череповец - ТГ-1",
    "Череповецкая ГРЭС блок 3": "Выведен из эксплуатации",
    "Череповецкая ГРЭС блок 2": "Выведен из эксплуатации",
    "Череповецкая ГРЭС блок 1": "Выведен из эксплуатации",
    '"ТЭЦ ЭВС-2 ""Северсталь"" ТГ-2"': "ТЭЦ-ЭВС-2 - ТГ-2",
    '"ТЭЦ ЭВС-2 ""Северсталь"" ТГ-1"': "ТЭЦ-ЭВС-2 - ТГ-1",
    "ТЭЦ-1 ТГ3": "-",
    "ТЭЦ-1 ТГ4": "Воронежская ТЭЦ-1 - ТГ-4",
    "ТЭЦ-1 ТГ6": "Воронежская ТЭЦ-1 - ТГ-6",
    "ТЭЦ-1 ТГ7": "Воронежская ТЭЦ-1 - ТГ-7",
    "ТЭЦ-2 ТГ2": "Воронежская ТЭЦ-2 - ТГ-2",
    "ТЭЦ-2 ТГ4": "Ярославская ТЭЦ-2 - ТГ4",
    "ТЭЦ-2 ТГ5": "Ярославская ТЭЦ-2 - ТГ5",
    "ТЭЦ-2 ТГ6": "Ярославская ТЭЦ-2 - ТГ6",
    "ТЭЦ-3 ТГ1": "Ярославская ТЭЦ-3 - ТГ-1",
    "ТЭЦ-3 ТГ2": "Ярославская ТЭЦ-3 - ТГ-2",
    "ТЭЦ-3 ТГ4": "Ярославская ТЭЦ-3 - ТГ-4",
    "ТЭЦ-3 ТГ5": "Ярославская ТЭЦ-3 - ТГ-5",
    "РыбГЭС 5Г": "Рыбинская ГЭС - 5Г",
    "РыбГЭС 6Г": "Рыбинская ГЭС - 6Г",
    "РыбГЭС 1Г": "Рыбинская ГЭС - 1Г",
    "РыбГЭС 2Г": "Рыбинская ГЭС - 2Г",
    "РыбГЭС 3Г": "Рыбинская ГЭС - 3Г",
    "РыбГЭС 4Г ": "Рыбинская ГЭС - 4Г",
    "УглГЭС 1Г": "Угличская ГЭС - Г1Г",
    "УглГЭС 2Г": "Угличская ГЭС - Г2Г",
    "Костромская ГРЭС Блок 1": "Костромская ГРЭС - ТГ-1",
    "Костромская ГРЭС Блок 2": "Костромская ГРЭС - ТГ-2",
    "Костромская ГРЭС Блок 3": "Костромская ГРЭС - ТГ-3",
    "Костромская ГРЭС Блок 4": "Костромская ГРЭС - ТГ-4",
    "Костромская ГРЭС Блок 5": "Костромская ГРЭС - ТГ-5",
    "Костромская ГРЭС Блок 6": "Костромская ГРЭС - ТГ-6",
    "Костромская ГРЭС Блок 7": "Костромская ГРЭС - ТГ-7",
    "Костромская ГРЭС Блок 8": "Костромская ГРЭС - ТГ-8",
    "Костромская ГРЭС Блок 9": "Костромская ГРЭС - ТГ-9",
    "Костромская ТЭЦ-2 ТГ-1": "Костромская ТЭЦ-2 - ТГ-1",
    "Костромская ТЭЦ-2 ТГ-2": "Костромская ТЭЦ-2 - ТГ-2",
    "Костромская ТЭЦ-1 ТГ-2": "Костромская ТЭЦ-1 - ТГ-2",
    "Костромская ТЭЦ-1 ТГ-5": "Костромская ТЭЦ-1 - ТГ-5",
    "Костромская ТЭЦ-1 ТГ-4": "-",
    "Костромская ТЭЦ-1 ТГ-6": "Костромская ТЭЦ-1 - ТГ-6",
    "Шарьинская ТЭЦ ТГ-1": "МУП «Шарьинская ТЭЦ» - ТГ-1",
    "Шарьинская ТЭЦ ТГ-2": "МУП «Шарьинская ТЭЦ» - ТГ-2",
    "Шарьинская ТЭЦ ТГ-3": "МУП «Шарьинская ТЭЦ» - ТГ-3",
    "Ивановская ТЭЦ-3 ТГ-1": "Ивановская ТЭЦ-3 - ТГ-1",
    "Ивановская ТЭЦ-3 ТГ-2": "Ивановская ТЭЦ-3 - ТГ-2",
    "Ивановская ТЭЦ-3 ТГ-3": "Ивановская ТЭЦ-3 - ТГ-3",
    "Ивановская ТЭЦ-3 ТГ-4": "Ивановская ТЭЦ-3 - ТГ-4",
    "Ивановская ТЭЦ-2 ТГ-1": "Ивановская ТЭЦ-2 - ТГ-1",
    "Ивановская ТЭЦ-2 ТГ-2": "Ивановская ТЭЦ-2 - ТГ-2",
    "Ивановская ТЭЦ-2 ТГ-3": "Ивановская ТЭЦ-2 - ТГ-3",
    "Ивановская ТЭЦ-2 ТГ-4": "Ивановская ТЭЦ-2 - ТГ-4",
    "Г-1 ТЭС ЦБК": "-",
    "Ивановские ПГУ Г-3 (Блок 2)": "Ивановские ПГУ - ГТ-21",
    "Ивановские ПГУ Г-4 (Блок 2)": "Ивановские ПГУ - ГТ-22",
    "Ивановские ПГУ ГП-2 (Блок 2)": "Ивановские ПГУ - ПТ-20",
    "Кольчугинской ТЭЦ ТГ2": "-",
    "Кольчугинской ТЭЦ ТГ1": "-",
    "Владимирской ТЭЦ-2 ТГ1": "Владимирская ТЭЦ-2 - ТГ1",
    "Владимирской ТЭЦ-2 ТГ7": "Владимирская ТЭЦ-2 - ТГ7",
    "Владимирской ТЭЦ-2 ТГ4": "Владимирская ТЭЦ-2 - ТГ4",
    "Владимирской ТЭЦ-2 ТГ3": "Владимирская ТЭЦ-2 - ТГ3",
    "Владимирской ТЭЦ-2 ТГ6": "Владимирская ТЭЦ-2 - ТГ6",
    "Владимирской ТЭЦ-2 ТГ5": "Владимирская ТЭЦ-2 - ТГ5",
    "ТЭЦ НЛМК ТГ-1 ": "ТЭЦ НЛМК - ТГ-1",
    "УТЭЦ НЛМК ТГ-1 ": "УТЭЦ НЛМК - ТГ1",
    "УТЭЦ НЛМК ТГ-2 ": "УТЭЦ НЛМК - ТГ2",
    "УТЭЦ НЛМК ТГ-3 ": "УТЭЦ НЛМК - ТГ3",
    "ТЭЦ НЛМК ТГ-3 ": "ТЭЦ НЛМК - ТГ-3",
    "ТЭЦ НЛМК ТГ-4 ": "ТЭЦ НЛМК - ТГ-4",
    "ТЭЦ НЛМК ТГ-2 ": "ТЭЦ НЛМК - ТГ-2",
    "ТЭЦ НЛМК ТГ-5 ": "ТЭЦ НЛМК - ТГ-5",
    "ТЭЦ НЛМК ТГ-6 ": "ТЭЦ НЛМК - ТГ-6",
    "ТЭЦ НЛМК ТГ-7 ": "ТЭЦ НЛМК - ТГ-7",
    "ЯрТЭС ТГ 1": "Ярославская ТЭС - ТГ 1",
    "ЯрТЭС ТГ 2": "Ярославская ТЭС - ТГ 2",
    "ЯрТЭС ТГ 3 ": "Ярославская ТЭС - ТГ 3",
    "Тутаевская ПГУ Б-2": "-",
    "Тутаевская ПГУ Б-1": "-",
    "СаровТЭЦ тг9": "Саровская ТЭЦ - ТГ-8",
    "СаровТЭЦ тг8": "Саровская ТЭЦ - ТГ-9",
    "НгТЭЦг2": "Новогорьковская ТЭЦ - ГТУ-2",
    "НгТЭЦг1": "Новогорьковская ТЭЦ - ГТУ-1",
    "ТГ-1-3 ТЭЦ-1 з-д им. Свердлова": "-",
    "СаровТЭЦ тг7": "Саровская ТЭЦ - ТГ-7",
    "СаровТЭЦ тг6": "Саровская ТЭЦ - ТГ-6",
    "ДзТЭЦ ГТУ3": "Дзержинская ТЭЦ - ГТУ-3",
    "НгТЭЦг8": "Новогорьковская ТЭЦ - ТГ-8",
    "НгТЭЦг6": "Новогорьковская ТЭЦ - ТГ-6",
    "ДзерТЭЦг5": "Дзержинская ТЭЦ - ТГ-5",
    "ДзерТЭЦг4": "Дзержинская ТЭЦ - ТГ-4",
    "ДзерТЭЦг2": "Дзержинская ТЭЦ - ТГ-2",
    "ДзерТЭЦг1": "Дзержинская ТЭЦ - ТГ-1",
    "СормТЭЦг4": "Сормовская ТЭЦ - ТГ-4",
    "СормТЭЦг3": "Сормовская ТЭЦ - ТГ-3",
    "СормТЭЦг2": "Сормовская ТЭЦ - ТГ-2",
    "СормТЭЦг1": "Сормовская ТЭЦ - ТГ-1",
    "НиГРЭС110г3": "Нижегородская ГРЭС - ТГ-3",
    "НиГРЭС110г2": "Нижегородская ГРЭС - ТГ-2",
    "НиГЭСтг8": "Нижегородская ГЭС - ГГ-8",
    "НиГЭСтг7": "Нижегородская ГЭС - ГГ-7",
    "НиГЭСтг6": "Нижегородская ГЭС - ГГ-6",
    "НиГЭСтг5": "Нижегородская ГЭС - ГГ-5",
    "НиГЭСтг4": "Нижегородская ГЭС - ГГ-4",
    "НиГЭСтг3": "Нижегородская ГЭС - ГГ-3",
    "НиГЭСтг2": "Нижегородская ГЭС - ГГ-2",
    "НиГЭСтг1": "Нижегородская ГЭС - ГГ-1",
    "ДзеТЭЦ-6 ": "Дзержинская ТЭЦ - ТГ-6",
    "Авт.ТЭЦ  тг8": "Автозаводская ТЭЦ - ТГ-8",
    "Авт.ТЭЦ  тг9": "Автозаводская ТЭЦ - ТГ-9",
    "Авт.ТЭЦ  тг7": "Автозаводская ТЭЦ - ТГ-7",
    "Авт.ТЭЦ  тг11": "Автозаводская ТЭЦ - ТГ-11",
    "Авт.ТЭЦ  тг10": "Автозаводская ТЭЦ - ТГ-10",
    "Авт.ТЭЦ тг12": "Автозаводская ТЭЦ - ТГ-12",
    "ВГЭС-500": "-",
    "Ростовская  ": "-",
    "КармГРЭС1": "-",
    "II СШ 500кВ_Ростовская АЭС": "-",
    "Шахты  ": "-",
    "I СШ 500кВ_Ростовская АЭС": "-",
    "Бугульма": "-",
    "IV СШ 500кВ_Ростовская АЭС": "-",
    "Ленинградская": "-",
    "III СШ 500кВ_Ростовская АЭС": "-",
    "Балаковская АЭС": "-",
    "СарГЭС220  ": "-",
    "ЖГЭС500  ": "-",
    "ЗайГРЭС500  ": "-",
    "Красноармейская": "-",
    "ВоткГЭС 1": "-",
    "ЧеГЭС500  ": "-",
    "Радуга 500": "-",
    "Ростовская АЭС_6": "-",
    "Куйбыш.500  ": "-",
    "Азот500  ": "-",
    "Киндери500  ": "-",
    "Донбасс": "-",
    "Щелоков": "-",
    "1 СШ Южная-500": "-",
    "Ростовская АЭС_17": "-",
    "Ростовская АЭС_2": "-",
    "Помары500  ": "-",
    "1 СШ Трубная-500": "-",
    "Саратовская  ": "-",
    "2СШ Трубная-500": "-",
    "Пенза2-220  ": "-",
    "Вятка   1": "-",
    "НКГЭС500  ": "-",
    "Ключики220  ": "-",
    "2 СШ Волга-500": "-",
    "1 СШ Волга-500": "-",
    "Белорусская  ": "-",
    "Пенза1-220  ": "-",
    "Окуловская": "-",
    "Фроловская I СШ": "-",
    "Курдюм220  ": "-",
    "Вешкайма  ": "-",
    "Витебск   ": "-",
    "отп ЛОСЕВО-330": "-",
    "Удмурт2 1": "-",
    "Змиевская ТЭС-330": "-",
    "ЛОСЕВО-330": "-",
    "Балаш. 2с Iсш": "-",
    "Балаш. 2с IIсш": "-",
    "Дягилевская ТЭЦ-10 ПГУ-1 ТГ-5": "Дягилевская ТЭЦ - ТГ-5",
    "Дягилевская ТЭЦ-10 ПГУ-1 ТГ-6": "Дягилевская ТЭЦ - ТГ-6",
    "Аткарск-220  ": "-",
    "Кузнецк220  ": "-",
    "Победа5": "-",
    "Дягилевская ТЭЦ-10 ПГУ-1 ТГ-7": "Дягилевская ТЭЦ - ТГ-7",
    "Искра": "-",
    "Окул_Тяг": "-",
    "Венера-110 ": "-",
    "ГУБКИН-110": "-",
    "Лиозно  ": "-",
    "КосаяГора-35    ": "-",
    "КТЗ            ": "-",
    "Сердобск220  ": "-",
    "ТЭЦ-3-35 ": "-",
    "Гидрообор 2 секц. 110": "-",
    "Гидрообор 1 секц. 110": "-",
    "Kузьминская": "-",
    "Ртищево220  ": "-",
    "Звезд110-1сш    ": "-",
    "Большевик 6кв": "-",
    "Заячерец": "-",
    "Леб 1 секц. 110": "-",
    "Леб 2 секц. 110": "-",
    "Кондрово: 1 с             ": "-",
    "Мичур 1 СШ 110": "-",
    "Мичур 2 СШ 110": "-",
    "Вельс110": "-",
    "KTЭЦ-2 2 сш 110": "-",
    "АТ-3 Прав 35 Стар": "-",
    "Заов.110": "-",
    "Хопер": "-",
    "Балаш. п.т. ВЛ Балашовская-Фроловская": "-",
    "ПС 500 кВ Балашовская": "-",
    "Кричев330   ": "-",
    "Брянская: СК-2           ": "-",
    "ПС 220 кВ Новобрянская: СК-2           ": "-",
    "Брянская: СК-1           ": "-",
    "УШР Покров ": "-",
    "БЕСКУДНИКОВО-20 АСК-2 ": "ПС 500 кВ Бескудниково - АСК-2",
    "БЕСКУДНИКОВО-20 АСК-1 ": "ПС 500 кВ Бескудниково - АСК-1",
    "Темников": "-",
    "Ельники": "-",
    "Юрино": "-",
    "Мелковка": "-",
    "Теньгушево": "-",
    "Ядрин": "-",
    "Ичалки": "-",
    "Т.Стан": "-",
    "Вад": "-",
    "Балаш. п.т. ВЛ 500 Балашовская-Волга": "-",
    "Рузаевка220": "-",
    "Саранск220": "-",
    "Кон. 220": "-",
    "Балашовская-220": "-",
    "Ивановские ПГУ: ГТ-11": "-",
    "Ивановские ПГУ: ГТ-12": "-",
    "Ивановские ПГУ: ПТ-10": "-",
}


class CellExcelDyn:
    name = 'N'
    model_type = 'W'
    area = 'A'
    p_nom = 'Z'
    u_nom = 'AB'
    cosf = 'AC'
    k_demp = 'AD'
    Tj_p = 'AK'
    xd1 = 'BA'
    xd = 'BB'
    xq = 'BC'
    xd2 = 'BD'
    xq2 = 'BE'
    xs = 'BF'
    t1d0 = 'BG'
    t2d0 = 'BH'
    t2q0 = 'BI'
    td1 = 'BJ'
    td2 = 'BK'


class RastrEvents:
    """
    Метод Onprot - выводит сообщения написанные: rastr.Printp("Сообщение из Printp")\n
    Метод OnLog
    """

    def OnLog(self, code, level, id, name, index, description, formName):
        if code == 2:
            print('[Error]', description)
        elif code == 3:
            print('[Warning]', description)
        elif code == 4:
            print('[Lightbulb]', description)
        elif code == 5:
            print('[Info]', description)
        else:
            print([code, description])

    def Onprot(self, message):
        print(message)


RASTR_1 = Dispatch('Astra.Rastr')
WithEvents(RASTR_1, RastrEvents)

RASTR_2 = Dispatch('Astra.Rastr')
WithEvents(RASTR_2, RastrEvents)


def move():
    rastr_win1 = RASTR_1
    rastr_win2 = RASTR_2
    # wb = Workbook()
    # ws = wb.active
    SHABLON_RST = r'C:\Program Files\RastrWin3\RastrWin3\SHABLON\динамика.rst'
    SHABLON_RG2 = r'C:\Program Files\RastrWin3\RastrWin3\SHABLON\режим.rg2'
    counter = 0
    rastr_win1.Load(1, 'БРМ лето 2021 МАКСИМУМ_7.rg2', SHABLON_RST)
    rastr_win2.Load(1, 'Зимний минимум 2027 минус 31.rst', SHABLON_RG2)
    tables_vetv_2_rst = rastr_win2.Tables("vetv")
    tables_vetv_1_rg2 = rastr_win1.Tables("vetv")

    for i in range(0, tables_vetv_2_rst.Count - 1):
        name_2 = tables_vetv_2_rst.Cols('dname').Z(i)
        name_2_tip = tables_vetv_2_rst.Cols('tip').Z(i)
        name_2_sp = name_2.split(" ")
        len_name_2_sp = len(name_2_sp)
        for j in range(0, tables_vetv_1_rg2.Count - 1):
            name_1 = tables_vetv_1_rg2.Cols('dname').Z(j)
            name_1_tip = tables_vetv_1_rg2.Cols('tip').Z(j)
            name_1_sp = name_1.split(" ")
            litter = 0
            for name in name_2_sp:
                if name in name_1_sp:
                    litter += 1
            if (((100 / len_name_2_sp) * litter) > 90) and (name_1_tip == name_2_tip):
                counter += 1
                print(f'{counter}. {name_2} -> {name_1}')
                # ws[f'A{counter + 1}'] = int(counter)
                # ws[f'B{counter + 1}'] = str(name_2)
                # ws[f'C{counter + 1}'] = str(name_1)

    # wb.save(filename="ветви.xlsx")


move()

