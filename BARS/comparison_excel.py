# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import units
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

wb = load_workbook(filename=r'L:\SER\Okhrimenko\08. БАРС\сравнение\Поиск Узлы.xlsx')

ws_1 = wb['Узлы 30']  # не рабочий режим
ws_2 = wb['Ветви 30']  # не рабочий режим
ws_3 = wb['Узлы 31']  # рабочий режим
ws_4 = wb['Ветви 31']  # рабочий режим

fill = GradientFill(stop=("00FFFF00", "00FFFF00"))

for i in range(2, 4631):
    ws1_col_Type_C = ws_1[f'C{i}'].value
    ws1_col_ny_D = ws_1[f'D{i}'].value
    ws1_col_Name_E = ws_1[f'E{i}'].value
    ws1_col_u_hom_F = ws_1[f'F{i}'].value
    ws1_col_Nsxn_G = ws_1[f'G{i}'].value
    ws1_col_Nrayon_H = ws_1[f'H{i}'].value
    ws1_col_P_n_I = ws_1[f'I{i}'].value
    ws1_col_Q_n_J = ws_1[f'J{i}'].value
    ws1_col_Pg_K = ws_1[f'K{i}'].value
    ws1_col_Qg_L = ws_1[f'L{i}'].value
    ws1_col_Vzd_M = ws_1[f'M{i}'].value
    ws1_col_Q_min_N = ws_1[f'N{i}'].value
    ws1_col_Q_max_O = ws_1[f'O{i}'].value
    ws1_col_B_sh_P = ws_1[f'P{i}'].value
    ws1_col_vras_Q = ws_1[f'Q{i}'].value
    ws1_col_Delta_R = ws_1[f'R{i}'].value

    ws3_col_Type_C = ws_3[f'C{i}'].value
    ws3_col_ny_D = ws_3[f'D{i}'].value
    ws3_col_Name_E = ws_3[f'E{i}'].value
    ws3_col_u_hom_F = ws_3[f'F{i}'].value
    ws3_col_Nsxn_G = ws_3[f'G{i}'].value
    ws3_col_Nrayon_H = ws_3[f'H{i}'].value
    ws3_col_P_n_I = ws_3[f'I{i}'].value
    ws3_col_Q_n_J = ws_3[f'J{i}'].value
    ws3_col_Pg_K = ws_3[f'K{i}'].value
    ws3_col_Qg_L = ws_3[f'L{i}'].value
    ws3_col_Vzd_M = ws_3[f'M{i}'].value
    ws3_col_Q_min_N = ws_3[f'N{i}'].value
    ws3_col_Q_max_O = ws_3[f'O{i}'].value
    ws3_col_B_sh_P = ws_3[f'P{i}'].value
    ws3_col_vras_Q = ws_3[f'Q{i}'].value
    ws3_col_Delta_R = ws_3[f'R{i}'].value

    if ws1_col_Type_C != ws3_col_Type_C:
        ws_1[f'C{i}'].fill = fill
        com_col_Type_C = Comment(f"{ws3_col_Type_C}", "Author")
        ws_1[f'C{i}'].comment = com_col_Type_C

    if ws1_col_ny_D != ws3_col_ny_D:
        ws1_col_ny_D.fill = fill
        com_col_ny_D = Comment(f"{ws3_col_ny_D}", "Author")
        ws_1[f'D{i}'].comment = com_col_ny_D

    if ws1_col_Name_E != ws3_col_Name_E:
        ws_1[f'E{i}'].fill = fill
        com_col_Name_E = Comment(f"{ws3_col_Name_E}", "Author")
        ws_1[f'E{i}'].comment = com_col_Name_E

    if ws1_col_u_hom_F != ws3_col_u_hom_F:
        ws_1[f'F{i}'].fill = fill
        com_col_u_hom_F = Comment(f"{ws3_col_u_hom_F}", "Author")
        ws_1[f'F{i}'].comment = com_col_u_hom_F

    if ws1_col_Nsxn_G != ws3_col_Nsxn_G:
        ws_1[f'G{i}'].fill = fill
        com_col_Nsxn_G = Comment(f"{ws3_col_Nsxn_G}", "Author")
        ws_1[f'G{i}'].comment = com_col_Nsxn_G

    if ws1_col_Nrayon_H != ws3_col_Nrayon_H:
        ws_1[f'H{i}'].fill = fill
        com_col_Nrayon_H = Comment(f"{ws3_col_Nrayon_H}", "Author")
        ws_1[f'H{i}'].comment = com_col_Nrayon_H

    if ws1_col_P_n_I != ws3_col_P_n_I:
        ws_1[f'I{i}'].fill = fill
        com_col_P_n_I = Comment(f"{ws3_col_P_n_I}", "Author")
        ws_1[f'I{i}'].comment = com_col_P_n_I

    if ws1_col_Q_n_J != ws3_col_Q_n_J:
        ws_1[f'J{i}'].fill = fill
        com_col_Q_n_J = Comment(f"{ws3_col_Q_n_J}", "Author")
        ws_1[f'J{i}'].comment = com_col_Q_n_J

    if ws1_col_Pg_K != ws3_col_Pg_K:
        ws_1[f'K{i}'].fill = fill
        com_col_Pg_K = Comment(f"{ws3_col_Pg_K}", "Author")
        ws_1[f'K{i}'].comment = com_col_Pg_K

    if ws1_col_Qg_L != ws3_col_Qg_L:
        ws_1[f'L{i}'].fill = fill
        com_col_Qg_L = Comment(f"{ws3_col_Qg_L}", "Author")
        ws_1[f'L{i}'].comment = com_col_Qg_L

    if ws1_col_Vzd_M != ws3_col_Vzd_M:
        ws_1[f'M{i}'].fill = fill
        com_col_Vzd_M = Comment(f"{ws3_col_Vzd_M}", "Author")
        ws_1[f'M{i}'].comment = com_col_Vzd_M

    if ws1_col_Q_min_N != ws3_col_Q_min_N:
        ws_1[f'N{i}'].fill = fill
        com_col_Q_min_N = Comment(f"{ws3_col_Q_min_N}", "Author")
        ws_1[f'N{i}'].comment = com_col_Q_min_N

    if ws1_col_Q_max_O != ws3_col_Q_max_O:
        ws_1[f'O{i}'].fill = fill
        com_col_Q_max_O = Comment(f"{ws3_col_Q_max_O}", "Author")
        ws_1[f'O{i}'].comment = com_col_Q_max_O

    if ws1_col_B_sh_P != ws3_col_B_sh_P:
        ws_1[f'P{i}'].fill = fill
        com_col_B_sh_P = Comment(f"{ws3_col_B_sh_P}", "Author")
        ws_1[f'P{i}'].comment = com_col_B_sh_P

    if ws1_col_vras_Q != ws3_col_vras_Q:
        ws_1[f'Q{i}'].fill = fill
        com_col_vras_Q = Comment(f"{ws3_col_vras_Q}", "Author")
        ws_1[f'Q{i}'].comment = com_col_vras_Q

    if ws1_col_Delta_R != ws3_col_Delta_R:
        ws_1[f'R{i}'].fill = fill
        com_col_Delta_R = Comment(f"{ws3_col_Delta_R}", "Author")
        ws_1[f'R{i}'].comment = com_col_Delta_R

for i in range(2, 6520):
    ws2_col_Type_C = ws_2[f'C{i}'].value
    ws2_col_N_nach_D = ws_2[f'D{i}'].value
    ws2_col_N_kon_E = ws_2[f'E{i}'].value
    ws2_col_N_par_F = ws_2[f'F{i}'].value
    ws2_col_ID_G = ws_2[f'G{i}'].value
    ws2_col_Name_H = ws_2[f'H{i}'].value
    ws2_col_R_I = ws_2[f'I{i}'].value
    ws2_col_X_J = ws_2[f'J{i}'].value
    ws2_col_B_K = ws_2[f'K{i}'].value
    ws2_col_Ktr_L = ws_2[f'L{i}'].value
    ws2_col_P_nach_O = ws_2[f'O{i}'].value
    ws2_col_Q_nach_P = ws_2[f'P{i}'].value
    ws2_col_Na_Q = ws_2[f'Q{i}'].value
    ws2_col_Imax_R = ws_2[f'R{i}'].value
    ws2_col_Sl_nach_S = ws_2[f'S{i}'].value
    ws2_col_Sl_kon_T = ws_2[f'T{i}'].value

    ws4_col_Type_C = ws_4[f'C{i}'].value
    ws4_col_N_nach_D = ws_4[f'D{i}'].value
    ws4_col_N_kon_E = ws_4[f'E{i}'].value
    ws4_col_N_par_F = ws_4[f'F{i}'].value
    ws4_col_ID_G = ws_4[f'G{i}'].value
    ws4_col_Name_H = ws_4[f'H{i}'].value
    ws4_col_R_I = ws_4[f'I{i}'].value
    ws4_col_X_J = ws_4[f'J{i}'].value
    ws4_col_B_K = ws_4[f'K{i}'].value
    ws4_col_Ktr_L = ws_4[f'L{i}'].value
    ws4_col_P_nach_O = ws_4[f'O{i}'].value
    ws4_col_Q_nach_P = ws_4[f'P{i}'].value
    ws4_col_Na_Q = ws_4[f'Q{i}'].value
    ws4_col_Imax_R = ws_4[f'R{i}'].value
    ws4_col_Sl_nach_S = ws_4[f'S{i}'].value
    ws4_col_Sl_kon_T = ws_4[f'T{i}'].value

    if ws2_col_Type_C != ws4_col_Type_C:
        ws_2[f'C{i}'].fill = fill
        com_col_Type_C = Comment(f"{ws4_col_Type_C}", "Author")
        ws_2[f'C{i}'].comment = com_col_Type_C

    if ws2_col_N_nach_D != ws4_col_N_nach_D:
        ws2_col_N_nach_D.fill = fill
        com_col_N_nach_D = Comment(f"{str(ws4_col_N_nach_D)}", "Author")
        ws_2[f'D{i}'].comment = com_col_N_nach_D

    if ws2_col_N_kon_E != ws4_col_N_kon_E:
        ws2[f'E{i}'].fill = fill
        com_col_N_kon_E = Comment(f"{ws4_col_N_kon_E}", "Author")
        ws2[f'E{i}'].comment = com_col_N_kon_E

    if ws2_col_N_par_F != ws4_col_N_par_F:
        ws_2[f'F{i}'].fill = fill
        com_col_N_par_F = Comment(f"{ws4_col_N_par_F}", "Author")
        ws_2[f'F{i}'].comment = com_col_N_par_F

    if ws2_col_ID_G != ws4_col_ID_G:
        ws_2[f'G{i}'].fill = fill
        com_col_ID_G = Comment(f"{ws4_col_ID_G}", "Author")
        ws_2[f'G{i}'].comment = com_col_ID_G

    if ws2_col_Name_H != ws4_col_Name_H:
        ws_2[f'H{i}'].fill = fill
        com_col_Name_H = Comment(f"{ws4_col_Name_H}", "Author")
        ws_2[f'H{i}'].comment = com_col_Name_H

    if ws2_col_R_I != ws4_col_R_I:
        ws_2[f'I{i}'].fill = fill
        com_col_R_I = Comment(f"{ws4_col_R_I}", "Author")
        ws_2[f'I{i}'].comment = com_col_R_I

    if ws2_col_X_J != ws4_col_X_J:
        ws_2[f'J{i}'].fill = fill
        com_col_X_J = Comment(f"{ws4_col_X_J}", "Author")
        ws_2[f'J{i}'].comment = com_col_X_J

    if ws2_col_B_K != ws4_col_B_K:
        ws_2[f'K{i}'].fill = fill
        com_col_B_K = Comment(f"{ws4_col_B_K}", "Author")
        ws_2[f'K{i}'].comment = com_col_B_K

    if ws2_col_Ktr_L != ws4_col_Ktr_L:
        ws_2[f'L{i}'].fill = fill
        com_col_Ktr_L = Comment(f"{ws4_col_Ktr_L}", "Author")
        ws_2[f'L{i}'].comment = com_col_Ktr_L

    if ws2_col_P_nach_O != ws4_col_P_nach_O:
        ws_2[f'O{i}'].fill = fill
        com_col_P_nach_O = Comment(f"{ws4_col_P_nach_O}", "Author")
        ws_2[f'O{i}'].comment = com_col_P_nach_O

    if ws2_col_Q_nach_P != ws4_col_Q_nach_P:
        ws_2[f'P{i}'].fill = fill
        com_col_Q_nach_P = Comment(f"{ws4_col_Q_nach_P}", "Author")
        ws_2[f'P{i}'].comment = com_col_Q_nach_P

    if ws2_col_Na_Q != ws4_col_Na_Q:
        ws_2[f'Q{i}'].fill = fill
        com_col_Na_Q = Comment(f"{ws4_col_Na_Q}", "Author")
        ws_2[f'Q{i}'].comment = com_col_Na_Q

    if ws2_col_Imax_R != ws4_col_Imax_R:
        ws_2[f'R{i}'].fill = fill
        com_col_Imax_R = Comment(f"{ws4_col_Imax_R}", "Author")
        ws_2[f'R{i}'].comment = com_col_Imax_R

    if ws2_col_Sl_nach_S != ws4_col_Sl_nach_S:
        ws_2[f'S{i}'].fill = fill
        com_col_Sl_nach_S = Comment(f"{ws4_col_Sl_nach_S}", "Author")
        ws_2[f'S{i}'].comment = com_col_Sl_nach_S

    if ws2_col_Sl_kon_T != ws4_col_Sl_kon_T:
        ws_2[f'T{i}'].fill = fill
        com_col_Sl_kon_T = Comment(f"{ws4_col_Sl_kon_T}", "Author")
        ws_2[f'T{i}'].comment = com_col_Sl_kon_T

wb.save(filename=r'L:\SER\Okhrimenko\08. БАРС\сравнение\Поиск Узлы_2.xlsx')
