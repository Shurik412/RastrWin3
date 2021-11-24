# -*- coding: utf-8 -*-
import csv
import xml.etree.ElementTree as et
from re import findall

import pandas as pd
from icecream import ic
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


name_area_dict = {
    "ti":
        {
            "I1889": "Генерация: ОЭС Центра",
            "I1153": "Потребление: ОЭС Центра",
            "I1494": "Генерация: Владимирское РДУ",
            "I1222": "Потребление: Владимирское РДУ",
            "I1105": "Генерация: Вологодское РДУ",
            "I1107": "Потребление: Вологодское РДУ",
            "I1125": "Генерация: Воронежское РДУ",
            "I1133": "Потребление: Воронежское РДУ",
            "I5576": "Генерация: Костромское РДУ",
            "I5577": "Потребление: Костромское РДУ",
            "I110": "Генерация: ЭС Ивановской обл.",
            "I779": "Потребление: ЭС Ивановской обл.",
            "I280": "Генерация: ЭС Костромской обл.",
            "I3936": "Потребление: ЭС Костромской обл.",
            "I1029": "Генерация: Курское РДУ",
            "I1069": "Потребление: Курское РДУ",
            "I1067": "Генерация: ЭС Белгородской обл.",
            "I1073": "Потребление: ЭС Белгородской обл.",
            "I1079": "Генерация: ЭС Курской обл.",
            "I1080": "Потребление: ЭС Курской обл.",
            "I1086": "Генерация: ЭС Орловской обл.",
            "I1087": "Потребление: ЭС Орловской обл.",
            "I5579": "Генерация: Липецкое РДУ",
            "I5580": "Потребление: Липецкое РДУ",
            "I2603": "Генерация: ЭС Липецкой обл.",
            "I2604": "Потребление: ЭС Липецкой обл.",
            "I126": "Генерация: ЭС Тамбовской обл.",
            "I3229": "Потребление: ЭС Тамбовской обл.",
            "I3641": "Генерация: Московское РДУ",
            "I1299": "Потребление: Московское РДУ",
            "I966": "Генерация: Рязанское РДУ",
            "I1167": "Потребление: Рязанское РДУ",
            "I3580": "Генерация: Смоленское РДУ",
            "I3583": "Потребление: Смоленское РДУ",
            "I112": "Генерация: ЭС Брянской обл.",
            "I757": "Потребление: ЭС Брянской обл.",
            "I3204": "Генерация: ЭС Калужской обл.",
            "I3205": "Потребление: ЭС Калужской обл.",
            "I98": "Генерация: ЭС Смоленской обл.",
            "I380": "Потребление: ЭС Смоленской обл.",
            "I1118": "Генерация: Тверское РДУ",
            "I1150": "Потребление: Тверское РДУ",
            "I1260": "Генерация: Тульское РДУ",
            "I1163": "Потребление: Тульское РДУ",
            "I1052": "Генерация: Ярославское РДУ",
            "I1053": "Потребление: Ярославское РДУ"
        },
    "Excel":
        {
            "Генерация: ОЭС Центра": "Ген_ОЭС_Центра",
            "Потребление: ОЭС Центра": "Пот_ОЭС_Центра",
            "Генерация: Владимирское РДУ": "Ген_Владимир_РДУ",
            "Потребление: Владимирское РДУ": "Пот_Владимир_РДУ",
            "Генерация: Вологодское РДУ": "Ген_Вологод_РДУ",
            "Потребление: Вологодское РДУ": "Пот_Вологод_РДУ",
            "Генерация: Воронежское РДУ": "Ген_Воронеж_РДУ",
            "Потребление: Воронежское РДУ": "Пот_Воронеж_РДУ",
            "Генерация: Костромское РДУ": "Ген_Костром_РДУ",
            "Потребление: Костромское РДУ": "Пот_Костром_РДУ",
            "Генерация: ЭС Ивановской обл.": "Ген_Иванов_обл",
            "Потребление: ЭС Ивановской обл.": "Пот_Иванов_обл",
            "Генерация: ЭС Костромской обл.": "Ген_ЭС_Костр_обл",
            "Потребление: ЭС Костромской обл.": "Пот_ЭС_Костр_обл",
            "Генерация: Курское РДУ": "Ген_Курское_РДУ",
            "Потребление: Курское РДУ": "Пот_Курское_РДУ",
            "Генерация: ЭС Белгородской обл.": "Ген_Белгород_обл",
            "Потребление: ЭС Белгородской обл.": "Пот_Белгород_обл",
            "Генерация: ЭС Курской обл.": "Ген_Курской_обл",
            "Потребление: ЭС Курской обл.": "Пот_Курской_обл",
            "Генерация: ЭС Орловской обл.": "Ген_Орловс_обл",
            "Потребление: ЭС Орловской обл.": "Пот_Орловс_обл",
            "Генерация: Липецкое РДУ": "Ген_Липецкое_РДУ",
            "Потребление: Липецкое РДУ": "Пот_Липецкое_РДУ",
            "Генерация: ЭС Липецкой обл.": "Ген_Липецкой_обл",
            "Потребление: ЭС Липецкой обл.": "Пот_Липецкой_обл",
            "Генерация: ЭС Тамбовской обл.": "Ген_Тамбовск_обл",
            "Потребление: ЭС Тамбовской обл.": "Пот_Тамбовск_обл",
            "Генерация: Московское РДУ": "Ген_Мос_РДУ",
            "Потребление: Московское РДУ": "Пот_Мос_РДУ",
            "Потребление: Рязанское РДУ": "Пот_Рязан_РДУ",
            "Генерация: Смоленское РДУ": "Ген_Смолен_РДУ",
            "Потребление: Смоленское РДУ": "Пот_Смолен_РДУ",
            "Генерация: ЭС Брянской обл.": "Ген_Брянской_обл",
            "Потребление: ЭС Брянской обл.": "Пот_Брянской_обл",
            "Генерация: ЭС Калужской обл.": "Ген_Калужск_обл",
            "Потребление: ЭС Калужской обл.": "Пот_Калужск_обл",
            "Генерация: ЭС Смоленской обл.": "Ген_Смолен_обл",
            "Потребление: ЭС Смоленской обл.": "Пот_Смолен_обл",
            "Генерация: Тверское РДУ": "Ген_Тверское_РДУ",
            "Потребление: Тверское РДУ": "Пот_Тверское_РДУ",
            "Генерация: Тульское РДУ": "Ген_Тульское_РДУ",
            "Потребление: Тульское РДУ": "Пот_Тульское_РДУ",
            "Генерация: Ярославское РДУ": "Ген_Ярослав_РДУ",
            "Потребление: Ярославское РДУ": "Пот_Ярослав_РДУ"
        },
    "Vapor":
        {
            "Ген_ОЭС_Центра": "Пот_ОЭС_Центра",
            "Ген_Владимир_РДУ": "Пот_Владимир_РДУ",
            "Ген_Вологод_РДУ": "Пот_Вологод_РДУ",
            "Ген_Воронеж_РДУ": "Пот_Воронеж_РДУ",
            "Ген_Костром_РДУ": "Пот_Костром_РДУ",
            "Ген_Иванов_обл": "Пот_Иванов_обл",
            "Ген_ЭС_Костр_обл": "Пот_ЭС_Костр_обл",
            "Ген_Курское_РДУ": "Пот_Курское_РДУ",
            "Ген_Белгород_обл": "Пот_Белгород_обл",
            "Ген_Курской_обл": "Пот_Курской_обл",
            "Ген_Орловс_обл": "Пот_Орловс_обл",
            "Ген_Липецкое_РДУ": "Пот_Липецкое_РДУ",
            "Ген_Липецкой_обл": "Пот_Липецкой_обл",
            "Ген_Тамбовск_обл": "Пот_Тамбовск_обл",
            "Ген_Мос_РДУ": "Пот_Мос_РДУ",
            "Ген_Смолен_РДУ": "Пот_Смолен_РДУ",
            "Ген_Брянской_обл": "Пот_Брянской_обл",
            "Ген_Калужск_обл": "Пот_Калужск_обл",
            "Ген_Смолен_обл": "Пот_Смолен_обл",
            "Ген_Тверское_РДУ": "Пот_Тверское_РДУ",
            "Ген_Тульское_РДУ": "Пот_Тульское_РДУ",
            "Ген_Ярослав_РДУ": "Пот_Ярослав_РДУ"
        }
}


def create_dict_out_xml(dict_name, file_name_xml: str = "file.xml"):
    xtree = et.parse(file_name_xml)
    xroot = xtree.getroot()
    df_cols = ["Name", "Time", "Value"]
    rows = []
    for index, node in enumerate(xroot):
        s_name = node.find("Name").text if node is not None else None
        s_time = node.find("Time").text if node is not None else None
        s_value = node.find("Value").text if node is not None else None
        rows.append({
            "Name": s_name,
            "Time": s_time,
            "Value": s_value
        })
    out_df = pd.DataFrame(rows, columns=df_cols)
    pattern = r'\d{2}.\d{2}.\d{4} \d{2}:\d{2}:\d{2}'
    for key in dict_name:
        out_df.Name = out_df.Name.replace(key, dict_name[key])
    for val in out_df.Time:
        rename_format_time = findall(pattern, val)
        out_df.Time = out_df.Time.replace(val, rename_format_time[0])
    return out_df


def save_csv(frame, file_name_csv: str) -> None:
    frame.to_csv(file_name_csv, sep=';', encoding='1251')


def save_xlsx(file_csv: str, file_name_save: str = 'file.xlsx') -> None:
    wb = Workbook()
    ws = wb.active
    with open(file_csv) as f:
        reader = csv.reader(f, delimiter=';')
        for row in reader:
            ws.append(row)
    wb.save(file_name_save)


def create_xlsx(name_file_load: str = 'file.xlsx',
                name_file2_save: str = 'file2.xlsx',
                name_file3_save: str = 'file3.xlsx') -> None:
    """

    :param name_file_load:
    :param name_file2_save:
    :param name_file3_save:
    :return:
    """
    wb = load_workbook(filename=name_file_load)
    ws = wb.active

    for key in name_area_dict['ti']:
        list_name_ = []
        list_time_ = []
        list_value_ = []
        title_ = ''

        for key_ in name_area_dict['Excel']:
            if key_ == name_area_dict['ti'][key]:
                title_ = name_area_dict['Excel'][key_]

        ws_ = wb.create_sheet(title=str(title_))
        for i in range(1, ws.max_row):
            name_excel = ws[f'{get_column_letter(2)}{i}'].value
            if name_excel == name_area_dict['ti'][key]:
                _name = ws[f'{get_column_letter(2)}{i}'].value
                _time = ws[f'{get_column_letter(3)}{i}'].value
                _value = ws[f'{get_column_letter(4)}{i}'].value
                list_name_.append(_name)
                list_time_.append(_time)
                list_value_.append(_value)

        for i in range(0, 25):
            ws_[f'{get_column_letter(1)}{i + 2}'] = i

        for index, val in enumerate(list_name_):
            ws_[f'{get_column_letter(2)}{index + 2}'] = val

        for index, val in enumerate(list_time_):
            ws_[f'{get_column_letter(3)}{index + 2}'] = val

        for index, val in enumerate(list_value_):
            ws_[f'{get_column_letter(4)}{index + 2}'] = val

    wb.save(name_file2_save)
    wb.close()

    wb = load_workbook(filename=name_file2_save)
    list_name_title_excel = ['Час(точка)', 'Название', 'Время', 'Р, МВт']
    for key in name_area_dict['Vapor']:
        ws_2 = wb[key]
        ws_1 = wb[name_area_dict['Vapor'][key]]
        for i, row in enumerate(ws_1.iter_rows()):
            for j, col in enumerate(row):
                ws_2.cell(row=i + 1, column=j + 6).value = col.value
                ws_2.cell(row=1, column=j + 6).value = list_name_title_excel[j]
                ws_2.cell(row=1, column=j + 1).value = list_name_title_excel[j]
    wb.save(name_file3_save)

    for sheetName in wb.worksheets:
        if sheetName.title not in name_area_dict['Vapor']:
            wb.remove_sheet(sheetName)
    wb.save(name_file3_save)


frame_csv = create_dict_out_xml(dict_name=name_area_dict['ti'], file_name_xml="file.xml")
save_csv(frame=frame_csv, file_name_csv='csv_area2.csv')
save_xlsx(file_csv='csv_area2.csv')
create_xlsx()
