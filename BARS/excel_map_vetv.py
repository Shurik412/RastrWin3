# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb_corr = load_workbook(filename=r'C:\Users\Ohrimenko_AG\Desktop\Корр_Карт_Ветвей\Карта Ветвей_corr.xlsx')
wb_new = load_workbook(filename=r'C:\Users\Ohrimenko_AG\Desktop\Корр_Карт_Ветвей\Карта Ветвей.xlsx')

ws_corr = wb_corr['Лист1']
ws_new = wb_new['Лист1']

for row in range(2, 3382):
    type_bars_new = ws_new[f'A{row}'].value
    ip_bars_new = ws_new[f'B{row}'].value
    iq_bars_new = ws_new[f'C{row}'].value
    np_bars_new = ws_new[f'D{row}'].value
    name_bars_new = ws_new[f'E{row}'].value
    i = 0

    for row_ in range(2, 3045):
        type_corr = ws_corr[f'A{row_}'].value
        ip_corr = ws_corr[f'B{row_}'].value
        iq_corr = ws_corr[f'C{row_}'].value
        np_corr = ws_corr[f'D{row_}'].value
        name_corr = ws_corr[f'E{row_}'].value

        if type_bars_new == type_corr and ip_bars_new == ip_corr and iq_bars_new == iq_corr and np_bars_new == np_corr and name_bars_new == name_corr:
            ip_smz = ws_corr[f'F{row_}'].value
            iq_smz = ws_corr[f'G{row_}'].value
            np_smz = ws_corr[f'H{row_}'].value
            group_smz = ws_corr[f'I{row_}'].value
            name_smz = ws_corr[f'J{row_}'].value

            ws_new[f'{get_column_letter(11 + i)}{row}'] = ip_smz
            ws_new[f'{get_column_letter(12 + i)}{row}'] = iq_smz
            ws_new[f'{get_column_letter(13 + i)}{row}'] = np_smz
            ws_new[f'{get_column_letter(14 + i)}{row}'] = group_smz
            ws_new[f'{get_column_letter(15 + i)}{row}'] = name_smz

            i = i + 6
            print(f'row={row} => row_={row_} --> if (-_-).')

wb_new.save(filename=r'C:\Users\Ohrimenko_AG\Desktop\Корр_Карт_Ветвей\Карта Ветвей_v2.xlsx')
