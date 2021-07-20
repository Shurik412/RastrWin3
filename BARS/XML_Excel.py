# -*- coding: utf-8 -*-
from openpyxl import load_workbook, Workbook

wb = load_workbook(r'C:\Users\Ohrimenko_AG\Desktop\TSTI.xlsx')
ws_one = wb.active

wb_ = Workbook()
ws = wb_.active

ws['A1'].value = 'Number'
ws['B1'].value = 'Name'
ws['C1'].value = 'Type TS/TI'
ws['D1'].value = 'Formula'

index = 2

for i in range(2, 1515):
    number = ws_one[f'H{i}'].value
    name = ws_one[f'E{i}'].value
    type_TS_TI = ws_one[f'G{i}'].value
    formula_ = ws_one[f'D{i}'].value

    if number is not None and number != 0:
        ws[f'A{index}'].value = number
        ws[f'B{index}'].value = name
        ws[f'C{index}'].value = type_TS_TI
        ws[f'D{index}'].value = formula_

        index = index + 1

wb_.save(filename=r'C:\Users\Ohrimenko_AG\Desktop\TS_TI.xlsx')
