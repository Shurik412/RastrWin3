# -*- coding: utf-8 -*-
from openpyxl import load_workbook

NACHALO_2 = 'F'
KONEC_2 = 'G'
PARR_2 = 'H'
GROUP_2 = 'I'
NAME_SMZU = 'J'

file2 = r'L:\SER\Охрименко\08. БАРС\Карта ветвей от РДУ\Московское РДУ\Карта ветвей (МосРДУ).xlsx'
name_ws_file2 = 'Лист1'
file_main = r'L:\SER\Охрименко\08. БАРС\Карта ветвей от РДУ\Карта Ветвей.xlsx'

wb_main = load_workbook(file_main)
wb_file2 = load_workbook(file2)

ws_main = wb_main['Лист1']
ws_file2 = wb_file2[name_ws_file2]

for i in range(2, 3400):
    # ws_main
    nach_2_main = ws_main[f'{NACHALO_2}{i}'].value
    kon_2_main = ws_main[f'{KONEC_2}{i}'].value
    parr_2_main = ws_main[f'{PARR_2}{i}'].value
    group_2_main = ws_main[f'{GROUP_2}{i}'].value
    name_smzu_main = ws_main[f'{NAME_SMZU}{i}'].value

    # ws_file2
    nach_2 = ws_file2[f'{NACHALO_2}{i}'].value
    kon_2 = ws_file2[f'{KONEC_2}{i}'].value
    parr_2 = ws_file2[f'{PARR_2}{i}'].value
    group_2 = ws_file2[f'{GROUP_2}{i}'].value
    name_smzu = ws_file2[f'{NAME_SMZU}{i}'].value

    if nach_2 is not None:
        print(f'{i} - IF: BARS = {nach_2} => SMZU = {nach_2_main}')
        ws_main[f'{NACHALO_2}{i}'] = nach_2
        ws_main[f'{KONEC_2}{i}'] = kon_2
        ws_main[f'{PARR_2}{i}'] = parr_2
        ws_main[f'{GROUP_2}{i}'] = group_2
        ws_main[f'{NAME_SMZU}{i}'] = name_smzu
    else:
        print(f'{i} - ELSE: BARS = {nach_2} => SMZU = {nach_2_main}')

wb_main.save(file_main)