# -*- coding: utf-8 -*-
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_interval, column_index_from_string
from icecream import ic


#####################################################################

def sampling_from_string(string):
    list_t = []
    list_end = []
    list_str = []
    for index, i in enumerate(string):
        if i == 't':
            list_t.append(index)
        if i == ']':
            list_end.append(index + 1)

    for index_list_t, t in enumerate(list_t):
        num_str_list_end = list_end[index_list_t]
        ts_plus_str = string[t:num_str_list_end]
        list_str.append(ts_plus_str)
    return list_str


def create_dop_ts(ws):
    STR_COLS = 'A'
    SWITCH = 0
    INDEX_STR_COLS = column_index_from_string(STR_COLS)
    for row in range(2, ws.max_row):
        cell_ = ws[f'{STR_COLS}{row}'].value
        ic(cell_)
        if cell_ is not None:
            list_cell = sampling_from_string(string=str(cell_))
            for index, val_list_cell in enumerate(list_cell):
                index_ = index + 1
                ic(val_list_cell)
                ws[f'{str(get_column_letter(INDEX_STR_COLS + index_))}{row}'].value = val_list_cell

        if SWITCH == 1:
            if cell_ is not None:
                try:
                    list_cell = list(map(int, re.findall(r"\[([+-]?\d+)\]", cell_)))  # \[(\d+)\]'
                    for index, val_list_cell in enumerate(list_cell):
                        index_ = index + 1
                        ws[f'{str(get_column_letter(INDEX_STR_COLS + index_))}{row}'].value = val_list_cell
                except TypeError:
                    print('Не соответствует типу!')


def create_list_ts(ws):
    my_list = []
    STR_COLS_TWO = 'A'
    for row in range(2, ws.max_row):
        cell_ = ws[f'{STR_COLS_TWO}{row}'].value
        if cell_ is not None and cell_ != 0 and cell_ != 1 and cell_ != '':
            col_index = column_index_from_string(STR_COLS_TWO)
            for col_iter in range(col_index, ws.max_column):
                cell_list = ws[f'{get_column_letter(col_iter)}{row}'].value
                if cell_list is not None and cell_list != 0 and cell_list != 1 and cell_list != '':
                    my_list.append(cell_list)
                else:
                    break
    return my_list


def create_column_unique_my_list(list_):
    set_res = set(list_)
    list_res = (list(set_res))
    wb = Workbook()
    ws = wb.active
    for index, item in enumerate(list_res):
        inx = index + 1
        ic(inx)
        ic(item)
        ws[f'C{inx}'].value = item
    wb.save(r'U:\ODU_DIR\СЭР\БАРС\ДопТС.xlsx')


#####################################################################


DIR_FILE_NAME = r'U:\ODU_DIR\СЭР\БАРС\ДопТС.xlsx'
NAME_SHEET = '12'
try:
    wb = load_workbook(filename=DIR_FILE_NAME)
    try:
        ws = wb[NAME_SHEET]
    except KeyError:
        ws = wb.active
    create_dop_ts(ws)
    list_ = create_list_ts(ws)
    create_column_unique_my_list(list_)
    wb.save(filename=DIR_FILE_NAME)
    print('The END')
except FileNotFoundError:
    print('Файл не найден!')
