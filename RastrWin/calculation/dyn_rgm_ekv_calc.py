# -*- coding: utf-8 -*-
from RastrWin.variables.variable_parametrs import VariableRowId
from RastrWin.object_rastr import RASTR
from time import time, localtime, strftime


class SteadyState:
    """
    Расчет режима, par – строка дополнительных параметров. Параметры могут быть следующими:
    "" – c параметрами по умолчанию;
    "p" – расчет с плоского старта;
    "z" – отключение стартового алгоритма;
    "c" – отключение контроля данных;
    "r" – отключение подготовки данных (можно использовать при повторном расчете режима
    с измененными значениями узловых мощностей и модулей напряжения).
    """

    def __init__(self, rastr_win=RASTR, par='', switch_command_line=False):
        self.rastr_win = rastr_win
        self.par = par
        self.switch_command_line = switch_command_line

    def __bool__(self):
        return self.switch_command_line

    def rgm(self):
        if self.switch_command_line is not False:
            start_time = time()
        else:
            start_time = 0
        print(f'Запуск "Расчет режима":')
        kod = self.rastr_win.Rgm(self.par)
        if self.switch_command_line is not False:
            print(f'\tСообщение о результатх расчета УР: {kod}')
            if kod != 0:
                print('\t\tРежим не сбалансирован!')
            elif kod == 0:
                print('\t\tРасчет УР завершен успешно!')
        if self.switch_command_line is not False:
            time_calc = time() - start_time
            print(
                f'\tВремя расчета режима: {strftime("M: %M [минут] S: %S [секунд]", localtime(time_calc))} (Seconds: {"%.2f" % (time_calc)} [секунд])')
        return kod


class Dynamic:
    """
    Функции для расчтета ЭМПП доступны в интерфейсе IFWDynamic. Интерфейс может быть
    получен с помощью свойства IRastr.FWDynamic.
    TimeReached - Вывод времени, достигнутого в расчете
    ResultMessage - Вывод сообщения о результатах расчета
    switch_result = False
    """

    def __init__(self, rastr_win=RASTR, calc_time=1, snap_max_count=1, switch_command_line=False):
        self.rastr_win = rastr_win
        self.calc_time = calc_time
        self.snap_max_count = snap_max_count
        self.FWDynamic = self.rastr_win.FWDynamic()
        self.TimeReached = self.FWDynamic.TimeReached  # Вывод времени, достигнутого в расчете
        self.ResultMessage = self.FWDynamic.ResultMessage  # Вывод сообщения о результатах расчета
        self.switch_command_line = switch_command_line

    def change_calc_time(self):
        settlement_time = VariableRowId(rastr_win=self.rastr_win, table='com_dynamics', column='Tras', row_id=0,
                                        switch_command_line=True)
        settlement_time.make_changes(value=self.calc_time)

    def change_snap_max_count(self):
        snap_max_count = VariableRowId(rastr_win=self.rastr_win, table='com_dynamics', column='SnapMaxCount', row_id=0,
                                       switch_command_line=True)
        snap_max_count.make_changes(value=self.snap_max_count)

    def run(self):
        if self.switch_command_line is not False:
            start_time = time()
        else:
            start_time = 0
        print(f'Запуск расчета ЭМПП:')
        self.FWDynamic.Run()
        if self.switch_command_line is not False:
            settlement_time = self.rastr_win.Tables('com_dynamics').Cols('Tras').Z(0)
            print(f'\tВремя расчета (T_расч): {float(settlement_time)}')
            print(f'\tСообщение о результатх расчета ЭМПП: {self.ResultMessage}')
            if self.ResultMessage == '':
                print('\t\tРасчет завершен успешно, потери синхронизма не выявлено.')
            elif self.ResultMessage == 0:
                print('\t\tРасчет завершен успешно, потери синхронизма не выявлено.')
            elif self.ResultMessage == 1:
                print('\t\tВыявлено превышение угла по ветви значения 180°.')
            elif self.ResultMessage == 2:
                print('\t\tВыявлено превышение угла по сопротивлению генератора значения 180°.')
            elif self.ResultMessage == 4:
                print('\t\tВыявлено превышение допустимой скорости вращения одного или нескольких генераторов.\n'
                      '\t\tДопустимая скорость вращения задается уставкой автомата безопасности в настройках динамики.')
        if self.switch_command_line is not False:
            time_calc = time() - start_time
            print(
                f'\tВремя расчета ЭМПП: {strftime("M: %M [минут] S: %S [секунд]", localtime(time_calc))} (Seconds: {"%.2f" % (time_calc)} [секунд])')


class Equivalent:
    """
    Эквивалентирование – упрощение электрической сети
    """

    def __init__(self, rastr_win=RASTR, switch_command_line=False):
        self.rastr_win = rastr_win
        self.switch_command_line = switch_command_line

    def __bool__(self):
        return self.switch_command_line

    def ekv(self, par=""):
        if self.switch_command_line is not False:
            start_time = time()
        else:
            start_time = 0
        print(f'Запуск "Эквивалентирование режима":')
        kod = self.rastr_win.Ekv(par)
        if self.switch_command_line is not False:
            print(f'\tСообщение о результатх расчета УР: {kod}')
            if kod != 0:
                print('\t\tРежим не сбалансирован!')
            elif kod == 0:
                print('\t\tРасчет УР завершен успешно!')
        if self.switch_command_line is not False:
            time_calc = time() - start_time
            print(
                f'\tВремя расчета режима: {strftime("M: %M [минут] S: %S [секунд]", localtime(time_calc))} (Seconds: {"%.2f" % (time_calc)} [секунд])')
        return kod


if __name__ == '__main__':
    import win32com.client
    from RastrWin.loading.load import load_file
    from RastrWin.directory_rastrwin.dir_test_rastr import file_RUSTab_9_rst as file_rst, \
        file_RUSTab_9_scn as file_scn
    from RastrWin.loading.shablon import shablon_file_dynamic, shablon_file_scenario
    from RastrWin.export_in_excel.export_data_rustab import ExportDataRUSTab
    from openpyxl import Workbook
    from RastrWin.variables.variable_parametrs import VariableSetSel
    from icecream import ic

    # from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active

    Rastr = win32com.client.Dispatch('Astra.Rastr')

    load_file(rastr_win=Rastr, file_path=file_rst, shablon=shablon_file_dynamic, switch_command_line=True)
    load_file(rastr_win=Rastr, file_path=file_scn, shablon=shablon_file_scenario, switch_command_line=True)
    load_file(rastr_win=Rastr, switch_command_line=True)

    din_one_run = Dynamic(rastr_win=Rastr, calc_time=2, switch_command_line=True)
    rgm_start = SteadyState(rastr_win=Rastr, par='', switch_command_line=True)
    rgm_start.rgm()

    print(f"До SnapMaxCount: {Rastr.Tables('com_dynamics').Cols('SnapMaxCount').Z(0)}")
    print(f"До Tras: {Rastr.Tables('com_dynamics').Cols('Tras').Z(0)}")

    din_one_run.change_calc_time()
    din_one_run.change_snap_max_count()

    print(f"После 1 SnapMaxCount: {Rastr.Tables('com_dynamics').Cols('SnapMaxCount').Z(0)}")
    print(f"После 1 Tras: {Rastr.Tables('com_dynamics').Cols('Tras').Z(0)}")

    din_one_run.run()

    data_dyn = ExportDataRUSTab(rastr_win=Rastr, table='Generator', switch_command_line=True)
    data_ = data_dyn.get_array(column='P', key='Num=2')
    ic(data_)
    for index, (param_arr, time_arr) in enumerate(data_):
        ws.cell(row=index + 2, column=1, value=time_arr)
        ws.cell(row=index + 2, column=2, value=param_arr)

    print(f"После 2 SnapMaxCount: {Rastr.Tables('com_dynamics').Cols('SnapMaxCount').Z(0)}")

    wb.save(filename='C:\\Users\\Ohrimenko_AG\\Desktop\\21\\1223.xlsx')
