# -*- coding: utf-8 -*-
from openpyxl import load_workbook, workbook


def delete_letter(str_one, str_two):
    list_simvol = ['-', '–', ':', 'ПС', 'пс', 'кВ', 'кв', 'вл', 'кл', '.', 'II', 'I']
    flag = True
    flag_one = True
    flag_two = True
    while flag:
        for j in list_simvol:
            if j in str_one:
                str_one.remove(j)
                flag_one = True
            else:
                flag_one = False
        for i in list_simvol:
            if i in str_two:
                str_two.remove(i)
                flag_two = True
            else:
                flag_two = False
        if flag_one == False and flag_two == False:
            flag = False


def percentage_of_coincidence_two_str(string_one, string_two, print_results=False):
    print(string_one)
    string_one = string_one.lower()
    string_two = string_two.lower()
    str_one = string_one.split(" ")
    str_two = string_two.split(" ")
    # print(f'do -> {str_one}')
    # print(f'do -> {str_two}')
    delete_letter(str_one, str_two)
    # print(f'after -> {str_one}')
    # print(f'after -> {str_two}')
    percent = 100
    percent_of_one_string = percent / len(str_one)
    for index, i in enumerate(str_one):
        if i in str_two:
            if print_results:
                print(f'{index + 1}.Percent = {round(percent)}')
        else:
            percent -= percent_of_one_string
            if print_results:
                print(f'{index + 1}.Percent = {round(percent)}')
    return round(percent)


# string_one = 'Соседка - Граждановская 1 сш'
# string_two = 'ПС 110 кВ Соседка : Нащёкинская - ВЛ 110 кВ Нащекинская – Соседка с отпайкой на ПС Граждановская'

wb = load_workbook(filename=r'C:\Users\Ohrimenko_AG\Desktop\BARS\test_vet.xlsx')
ws = wb.active
for i in range(2, 3000):
    name_bars_ex = str(ws[f"E{i}"].value)
    name_smzu_ex = str(ws[f"J{i}"].value)
    ws[f"K{i}"] = percentage_of_coincidence_two_str(string_one=name_bars_ex, string_two=name_smzu_ex,
                                                    print_results=False)

wb.save(filename=r'C:\Users\Ohrimenko_AG\Desktop\BARS\test_vet2.xlsx')
