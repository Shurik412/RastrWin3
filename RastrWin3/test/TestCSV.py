# -*- coding: utf-8 -*-
import csv
from openpyxl import load_workbook, Workbook


r'C:\Users\Ohrimenko_AG\Desktop\TestCSV\ogrsech.csv'


def open_file_csv(file_dir_name, delimiter=';', index_row=0, index_col=0):
    with open(file=file_dir_name) as f:
        reader = csv.reader(f, delimiter=delimiter)
        list_file = []
        for index, row in enumerate(reader):
            list_file.append(row)

        return list_file


def value_CSV(file, col, row):
    try:
        return open_file_csv(file_dir_name=file, delimiter=';', index_row=row, index_col=col)
    except FileNotFoundError:
        print('Файл не найден!')


# file_dir_name = r'C:\Users\Ohrimenko_AG\Desktop\TestCSV\23.csv'
#
# print(value_CSV(file=file_dir_name, col=6, row=3))

def convert_csv_to_xlsx(file, file_save):
    wb = Workbook()
    sheet = wb.active
    CSV_SEPARATOR = ";"
    with open(file) as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                    cell_exp = sheet.cell(row=r + 1, column=idx + 1)
                    cell_exp.value = val
    wb.save(file_save)


file = r'C:\Users\Ohrimenko_AG\Desktop\TestCSV\23.csv'
file_save = r'C:\Users\Ohrimenko_AG\Desktop\TestCSV\23.xlsx'
convert_csv_to_xlsx(file, file_save)